import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
import json
import os
import datetime
import warnings
from scipy.optimize import minimize
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.spatial.distance import squareform
try:
    from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
    from sklearn.preprocessing import StandardScaler
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False

# Assure que le dossier du script est dans sys.path
# (nécessaire sur Streamlit Cloud où le CWD peut différer)
import sys, os
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

try:
    from github_storage import (save_smf_data, load_smf_data,
                                 save_financial_db_github, load_financial_db_github,
                                 save_charia_results, load_charia_results,
                                 is_github_configured)
    GITHUB_STORAGE = True
except ImportError:
    GITHUB_STORAGE = False

try:
    from charia_screening import (parse_screening_file as parse_charia,
                                   screen_all_from_fin_data,
                                   get_charia_label,
                                   get_charia_compatible_tickers)
    CHARIA_AVAILABLE = True
except ImportError:
    CHARIA_AVAILABLE = False

try:
    from fs_parser import (parse_financial_file, merge_financial_data,
                           save_financial_db, load_financial_db,
                           validate_and_fix_units)
    from valuation_models import (valuation_pe, valuation_pb, valuation_ddm,
                                   valuation_dcf, combined_price, compute_beta,
                                   calibrate_params)
    VALUATION_AVAILABLE = True
except ImportError as _e:
    VALUATION_AVAILABLE = False
    _VALUATION_ERROR = str(_e)
warnings.filterwarnings("ignore")

st.set_page_config(page_title="CGF Gestion · SMF BRVM", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=JetBrains+Mono:wght@300;400;500&display=swap');
:root{--bg:#0b0f1a;--surface:#111827;--card:#161d2e;--border:#1e2d45;
      --accent:#3b82f6;--accent2:#06b6d4;--gold:#f59e0b;--green:#10b981;
      --red:#ef4444;--text:#e2e8f0;--muted:#64748b;}
html,body,[class*="css"]{background:var(--bg);color:var(--text);font-family:'JetBrains Mono',monospace;}
[data-testid="stSidebar"]{background:var(--surface)!important;border-right:1px solid var(--border);}
h1,h2,h3{font-family:'Syne',sans-serif!important;}
[data-testid="metric-container"]{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:16px;}
[data-testid="metric-container"] label{color:var(--muted)!important;font-size:11px!important;text-transform:uppercase;}
[data-testid="metric-container"] [data-testid="stMetricValue"]{color:var(--accent)!important;font-family:'JetBrains Mono'!important;}
.stButton>button{background:var(--accent)!important;color:white!important;border:none!important;border-radius:6px!important;transition:all .2s;}
.stButton>button:hover{background:var(--accent2)!important;transform:translateY(-1px);}
.stTabs [data-baseweb="tab-list"]{background:var(--surface);border-bottom:1px solid var(--border);}
.stTabs [data-baseweb="tab"]{color:var(--muted)!important;font-family:'JetBrains Mono'!important;font-size:13px;padding:10px 18px;}
.stTabs [aria-selected="true"]{color:var(--accent)!important;border-bottom:2px solid var(--accent)!important;}
.stInfo{background:rgba(59,130,246,.1)!important;border-left:3px solid var(--accent)!important;}
.stSuccess{background:rgba(16,185,129,.1)!important;border-left:3px solid var(--green)!important;}
.stWarning{background:rgba(245,158,11,.1)!important;border-left:3px solid var(--gold)!important;}
hr{border-color:var(--border)!important;}
.pill{display:inline-block;padding:2px 10px;border-radius:20px;font-size:11px;font-weight:600;
      letter-spacing:.08em;text-transform:uppercase;background:rgba(59,130,246,.15);
      color:var(--accent);border:1px solid rgba(59,130,246,.3);margin-bottom:8px;}
.sh{font-family:'Syne',sans-serif;font-size:22px;font-weight:700;color:#e2e8f0;margin:0 0 4px 0;}
.ss{font-family:'JetBrains Mono',monospace;font-size:12px;color:#64748b;margin-bottom:20px;}
.fbox{background:#0d1526;border:1px solid #1e3a5f;border-radius:8px;padding:14px 18px;
      margin:10px 0;font-family:'JetBrains Mono',monospace;font-size:13px;color:#93c5fd;}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTS ────────────────────────────────────────────────
BENCH_COLS = ['Unnamed: 0','ANNEE','JOUR','Unnamed: 62','.BRVMCI','BRVM30',
              'BRVM PREST','BRVM-PRINC','BRVM-C TR','BRVM-CB','BRVM-CD',
              'BRVM-ENER','BRVM-SFIN','BRVM-SPUB']

# ─── DATA LOADING ─────────────────────────────────────────────
@st.cache_data
def load_data(uploaded_bytes):
    xl = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    # COURS
    cours_raw = pd.read_excel(xl, sheet_name="Cours")
    cours = cours_raw.drop(columns=[c for c in BENCH_COLS if c in cours_raw.columns], errors='ignore')
    cours['Date'] = pd.to_datetime(cours['Date'], errors='coerce')
    cours = cours.dropna(subset=['Date']).set_index('Date').sort_index()
    tickers = [c for c in cours.columns]

    # VOLUMES
    vol_raw = pd.read_excel(xl, sheet_name="Volume moyen")
    vol_raw['Date'] = pd.to_datetime(vol_raw['Date'], errors='coerce')
    vol = vol_raw.dropna(subset=['Date']).set_index('Date').sort_index()
    # Keep only ticker columns that are numeric (exclude formula/text columns)
    vol_ticker_cols = [c for c in tickers if c in vol.columns]
    vol = vol[vol_ticker_cols].apply(pd.to_numeric, errors='coerce')

    # DIVIDENDES
    div_raw = pd.read_excel(xl, sheet_name="Historique_dividende")
    div = div_raw[['Date'] + [c for c in tickers if c in div_raw.columns]].copy()
    div = div.dropna(subset=['Date'])
    div['Date'] = div['Date'].astype(int)

    # MOYENNE COURS
    moy_raw = pd.read_excel(xl, sheet_name="Moyenne_cours")
    moy = moy_raw[['Date'] + [c for c in tickers if c in moy_raw.columns]].copy()
    moy = moy.dropna(subset=['Date'])
    moy['Date'] = moy['Date'].astype(int)

    # NB TITRES
    # Structure réelle (header=None) :
    #   Ligne 0 : [NaN, 'Date', 'CABC', 'FTSC', ...]
    #   Ligne 1 : [NaN, 'Titres flottants', val, val, ...]
    #   Ligne 2 : [NaN, 'Nombre de titres', val, val, ...]
    nb_raw = pd.read_excel(xl, sheet_name="Nb_titres", header=None)
    nb_titres, nb_flottant = {}, {}
    nb_ticker_col = {}
    for col_idx in range(2, nb_raw.shape[1]):
        val = nb_raw.iloc[0, col_idx]
        if pd.notna(val):
            nb_ticker_col[str(val).strip()] = col_idx
    for row_idx in range(1, nb_raw.shape[0]):
        label = str(nb_raw.iloc[row_idx, 1]).strip().lower()
        for t, col_idx in nb_ticker_col.items():
            v = nb_raw.iloc[row_idx, col_idx]
            try:
                v = float(v)
                if not np.isnan(v) and v > 0:
                    if 'flottant' in label:
                        nb_flottant[t] = v
                    elif 'nombre' in label:
                        nb_titres[t] = v
            except (TypeError, ValueError):
                pass

    # TABLEAU DE BORD (fondamentaux)
    # Structure réelle (header=None) :
    #   Ligne 0 : [NaN, 1000000, 'CABC', 'FTSC', ...]  ← tickers col 2+
    #   Col 1   : années (2019..2024) ou NaN
    #   Col 2   : label métrique ('CAPITAUX PROPRES') ou valeur numérique
    #   Lignes label : col1=NaN, col2='CAPITAUX PROPRES'
    #   Lignes data  : col1=année, col2+=valeurs par ticker
    tb_raw = pd.read_excel(xl, sheet_name="Tableau de bord", header=None)

    ticker_col = {}
    for col_idx in range(2, tb_raw.shape[1]):
        val = tb_raw.iloc[0, col_idx]
        if pd.notna(val) and str(val).strip() not in ["nan", ""]:
            ticker_col[str(val).strip()] = col_idx

    def extract_metric(label_keyword):
        result = {}
        in_block = False
        for row_idx in range(1, tb_raw.shape[0]):
            cell2 = str(tb_raw.iloc[row_idx, 2]).strip().upper()
            cell1 = tb_raw.iloc[row_idx, 1]

            # Détection du label de la métrique
            if label_keyword.upper() in cell2 and pd.isna(cell1):
                in_block = True
                continue

            if in_block:
                try:
                    year = int(float(cell1)) if pd.notna(cell1) else None
                except (ValueError, TypeError):
                    year = None

                if year is None:
                    # Ligne de tickers répétée (ex: row avec 'CABC','FTSC'...) → ignorer
                    if cell2 in ticker_col:
                        continue
                    # Nouveau label métrique → fin du bloc
                    if pd.isna(cell1) and cell2 not in ["NAN", ""]:
                        break
                    continue

                # Lire les valeurs pour chaque ticker (col 2+ de tb_raw)
                for t, col_idx in ticker_col.items():
                    if col_idx < tb_raw.shape[1]:
                        v = tb_raw.iloc[row_idx, col_idx]
                        try:
                            v = float(v)
                            if not np.isnan(v):
                                result.setdefault(t, {})[year] = v
                        except (TypeError, ValueError):
                            pass
        return result

    # Années disponibles dans les fondamentaux (exclut valeurs parasites > 9999)
    fundamental_years = sorted(
        {y for metric in [
            extract_metric("CAPITAUX PROPRES"),
            extract_metric("RESULTAT NET"),
        ] for t_data in metric.values() for y in t_data.keys()
         if isinstance(y, int) and 1990 <= y <= 2100},
        reverse=True
    )

    return {
        "cours": cours, "volumes": vol, "dividendes": div,
        "moyenne_cours": moy, "nb_titres": nb_titres, "nb_flottant": nb_flottant,
        "tickers": tickers,
        "fundamental_years": fundamental_years,
        "capitaux_propres": extract_metric("CAPITAUX PROPRES"),
        "resultat_net":     extract_metric("RESULTAT NET"),
        "flux_treso":       extract_metric("FLUX DE TRESORERIE"),
        "chiffre_affaires": extract_metric("CHIFFRE D'AFFAIRES"),
        "resultat_expl":    extract_metric("RESULTAT D'EXPLOITATION"),
    }

# ─── NORMALISATION HELPER ─────────────────────────────────────
# Note technique : F_i(t,T) = Σ w_ij * m_ij(t,T) / max_{E_t}(m_ij(t,T))
# La division est inconditionnelle — max peut être négatif (ex: rendements tous négatifs).
# Seul cas exclu : max == 0 exactement (division par zéro).
# Volatilité (cas inversé) : F_i = Σ w_ij * min_{E_t}(m_ij) / m_ij(t,T)

def _normalise_standard(col: pd.Series) -> pd.Series:
    """m_ij / max_{E_t}(m_ij)  — formule générale de la note technique."""
    col = col.replace([np.inf, -np.inf], np.nan)
    mx  = col.max()
    if pd.isna(mx) or mx == 0:
        return pd.Series(np.nan, index=col.index)
    return col / mx

def _normalise_volatility(col: pd.Series) -> pd.Series:
    """min_{E_t}(m_ij) / m_ij  — formule inversée pour la volatilité."""
    col = col.replace([np.inf, -np.inf], np.nan)
    mn  = col.min()
    if pd.isna(mn):
        return pd.Series(np.nan, index=col.index)
    # avoid division by zero for any individual m_ij == 0
    return (mn / col.replace(0, np.nan))

# ─── FACTOR FUNCTIONS ─────────────────────────────────────────

def compute_value_factor(data, year, weights, date_start=None, date_end=None):
    """
    F_value(t,T) = Σ w_j * m_j(t,T) / max_{E_t}(m_j(t,T))

    Métriques (toutes exprimées par action, divisées par le cours moyen) :

      1. B/P  = (Capitaux propres / Nb titres) / Cours moyen      [Book-to-Price]
                 → plus élevé = plus sous-évalué  ✓ déjà dans le bon sens

      2. E/P  = (Résultat net / Nb titres) / Cours moyen          [Earnings-to-Price]
                 → inverse de P/E : score élevé = titre bon marché ✓

      3. FCF/P  = (Flux nets trésorerie / Nb titres) / Cours moyen [FCF yield]

      4. CA/P   = (Chiffre d'affaires / Nb titres) / Cours moyen   [Sales yield]

      5. EBIT/P = (Résultat d'exploitation / Nb titres) / Cours moyen
                  → proxy de EV/EBIT inversé, rapporté au cours unitaire

    Le cours moyen est calculé sur la plage [date_start, date_end] à partir
    des cours journaliers (feuille Cours), ou sur l'année calendaire si non fournie.
    """
    nb = data["nb_titres"]

    # ── 1. Cours moyen sur la plage de dates choisie ──────────────
    cours = data["cours"].apply(pd.to_numeric, errors="coerce")
    if date_start is not None and date_end is not None:
        mask = (cours.index >= pd.to_datetime(date_start)) & \
               (cours.index <= pd.to_datetime(date_end))
        cours_period = cours.loc[mask]
        period_label = f"{date_start} → {date_end}"
    else:
        # fallback : année calendaire
        cours_period = cours[cours.index.year == year]
        period_label = str(year)

    if cours_period.empty:
        return None

    cours_moy = cours_period.mean(numeric_only=True)  # Series : ticker → cours moyen

    # ── 2. Fondamentaux annuels (année sélectionnée) ───────────────
    detail = {}
    for t in data["tickers"]:
        if t not in nb or nb[t] <= 0:
            continue
        cm = cours_moy.get(t, np.nan)
        if pd.isna(cm) or cm <= 0:
            continue

        n   = nb[t]                                               # nombre de titres
        cp  = data["capitaux_propres"].get(t, {}).get(year, np.nan)
        rn  = data["resultat_net"].get(t, {}).get(year, np.nan)
        fcf = data["flux_treso"].get(t, {}).get(year, np.nan)
        ca  = data["chiffre_affaires"].get(t, {}).get(year, np.nan)
        rex = data["resultat_expl"].get(t, {}).get(year, np.nan)

        # Ratios par action / cours moyen
        bp_val   = (cp  / n) / cm if pd.notna(cp)  else np.nan   # B/P  = (CP/n) / cours
        ep_val   = (rn  / n) / cm if pd.notna(rn)  else np.nan   # E/P  = (RN/n) / cours
        fcfp     = (fcf / n) / cm if pd.notna(fcf) else np.nan   # FCF/P
        cap_val  = (ca  / n) / cm if pd.notna(ca)  else np.nan   # CA/P
        ebitp    = (rex / n) / cm if pd.notna(rex) else np.nan   # EBIT/P

        detail[t] = {
            "B/P  (CP/n÷cours)":    bp_val,
            "E/P  (RN/n÷cours)":    ep_val,
            "FCF/P (FCF/n÷cours)":  fcfp,
            "CA/P  (CA/n÷cours)":   cap_val,
            "EBIT/P (Rex/n÷cours)": ebitp,
            "Cours moyen":          cm,
            "Nb titres":            n,
        }

    if not detail:
        return None

    df = pd.DataFrame(detail).T
    # Force float — DataFrame.T depuis un dict mixte donne dtype=object
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    metrics = ["B/P  (CP/n÷cours)", "E/P  (RN/n÷cours)",
               "FCF/P (FCF/n÷cours)", "CA/P  (CA/n÷cours)",
               "EBIT/P (Rex/n÷cours)"]

    # Garder uniquement les titres qui ont AU MOINS une métrique valide
    has_any = df[metrics].notna().any(axis=1)
    df = df[has_any]
    if df.empty:
        return None

    # ── 3. Normalisation : m_ij / max_{E_t}(m_ij) ─────────────────
    norm_df = pd.DataFrame({m: _normalise_standard(df[m]) for m in metrics})

    # ── 4. Score pondéré F_value = Σ w_j * (m_j / max) ───────────
    score = sum(norm_df[m].fillna(0) * weights.get(m, 0) for m in metrics)

    res = pd.DataFrame({
        "Score Value":           score,
        **{m: df[m] for m in metrics},
        "Cours moyen":           df["Cours moyen"],
        "Période":               period_label,
    })
    # Exclure les titres avec score strictement nul (aucune métrique disponible)
    res = res[res["Score Value"] > 0].sort_values("Score Value", ascending=False)
    return res if not res.empty else None


def compute_momentum_factor(data, horizon_ranges, weights):
    """
    F_mom(t,T) = Σ w_h * rdt_moyen_h(t,T) / max_{E_t}(rdt_moyen_h)

    horizon_ranges : dict  {label: (date_start, date_end)}
    Pour chaque horizon activé, le rendement moyen journalier est calculé
    sur la plage [date_start, date_end] choisie par l'utilisateur.
    """
    returns = data["cours"].apply(pd.to_numeric, errors='coerce').pct_change().dropna(how='all')

    mom_cols = {}
    active_horizons = []

    for h, (d_start, d_end) in horizon_ranges.items():
        if d_start is None or d_end is None:
            continue
        try:
            sd = pd.to_datetime(d_start)
            ed = pd.to_datetime(d_end)
        except Exception:
            continue
        sub = returns[(returns.index >= sd) & (returns.index <= ed)]
        if sub.empty:
            continue
        rdt_moy = sub.mean(numeric_only=True).replace([np.inf, -np.inf], np.nan)
        mom_cols[h] = rdt_moy
        active_horizons.append(h)

    if not active_horizons:
        return None

    mom_df = pd.DataFrame(mom_cols)

    # Normalisation : m_ij / max_{E_t}(m_ij)
    norm_df = pd.DataFrame({h: _normalise_standard(mom_df[h]) for h in active_horizons})

    # Score pondéré — renormalise les poids sur les horizons actifs
    active_weights = {h: weights.get(h, 0) for h in active_horizons}
    w_sum = sum(active_weights.values())
    if w_sum > 0:
        active_weights = {h: w/w_sum for h, w in active_weights.items()}

    score = sum(norm_df[h].fillna(0) * active_weights[h] for h in active_horizons)

    res = pd.DataFrame({
        "Score Momentum": score,
        **{f"Rdt {h}": mom_df[h] for h in active_horizons},
    })
    return res.dropna(subset=["Score Momentum"]).sort_values("Score Momentum", ascending=False)


def compute_volatility_factor(data, end_date, window=252):
    """
    F_vol(t,T) = w * min_{E_t}(σ) / σ(T)
    Cas particulier inversé — titre le moins volatile reçoit le score le plus élevé.
    """
    returns = data["cours"].apply(pd.to_numeric, errors='coerce').pct_change().dropna(how='all')
    try:
        ed = pd.to_datetime(end_date)
    except Exception:
        ed = returns.index[-1]

    # m_ij = écart-type des rendements sur la fenêtre
    vol = returns[returns.index <= ed].tail(window).std(numeric_only=True).dropna()

    # F_vol = min_{E_t}(σ) / σ(T)  — w = 1 (métrique unique)
    score = _normalise_volatility(vol).replace([np.inf, -np.inf], np.nan).dropna()

    return pd.DataFrame({
        "Score Volatilité": score,
        "Écart-type σ":     vol,
        "σ annualisée":     vol * np.sqrt(252),
    }).sort_values("Score Volatilité", ascending=False)


def compute_dividend_factor(data, year, date_start=None, date_end=None):
    """
    F_div(t,T) = 1.0 * DY(t,T) / max_{E_t}(DY(t,T))

    Métrique unique (w = 100%) : Dividend Yield = Dividende annuel / Cours moyen
    Le cours moyen est calculé sur la plage [date_start, date_end].
    """
    div = data["dividendes"]
    dr  = div[div['Date'] == year]
    if dr.empty:
        return None

    # Cours moyen sur la plage choisie
    cours = data["cours"].apply(pd.to_numeric, errors="coerce")
    if date_start is not None and date_end is not None:
        mask = (cours.index >= pd.to_datetime(date_start)) & \
               (cours.index <= pd.to_datetime(date_end))
        cours_period = cours.loc[mask]
    else:
        cours_period = cours[cours.index.year == year]

    if cours_period.empty:
        return None

    cours_moy = cours_period.mean(numeric_only=True)

    yields = {}
    for t in data["tickers"]:
        if t not in dr.columns:
            continue
        d = dr[t].values[0]
        p = cours_moy.get(t, np.nan)
        if pd.notna(d) and pd.notna(p) and p > 0 and d >= 0:
            yields[t] = float(d) / float(p)

    if not yields:
        return None

    s = pd.Series(yields)
    # F_div = (100% × DY) / max_{E_t}(DY)   [100% = poids unique w=1]
    score = _normalise_standard(s)

    return pd.DataFrame({
        "Score Dividende": score,
        "Dividend Yield":  s,
        "Cours moyen":     cours_moy.reindex(s.index),
    }).sort_values("Score Dividende", ascending=False)


def compute_liquidity_factor(data, date_start=None, date_end=None):
    """
    F_liq(t,T) = 1.0 * Vol_moy(T) / max_{E_t}(Vol_moy)

    Métrique unique (w = 100%) : volume moyen transigé sur la plage choisie.
    """
    vol_df = data["volumes"].apply(pd.to_numeric, errors="coerce")

    if date_start is not None and date_end is not None:
        mask = (vol_df.index >= pd.to_datetime(date_start)) & \
               (vol_df.index <= pd.to_datetime(date_end))
        vol_period = vol_df.loc[mask]
        period_label = f"{date_start} → {date_end}"
    else:
        vol_period = vol_df
        period_label = "Historique complet"

    if vol_period.empty:
        return None

    avg = vol_period.mean(numeric_only=True).replace([np.inf, -np.inf], np.nan).dropna()

    # F_liq = Vol_moy / max_{E_t}(Vol_moy)
    score = _normalise_standard(avg)

    return pd.DataFrame({
        "Score Liquidité": score,
        "Volume moyen":    avg,
        "Période":         period_label,
    }).sort_values("Score Liquidité", ascending=False)

# ─── OPTIMISATION DE L'UNIVERS ────────────────────────────────

def filter_liquidity(mf_scores, liq_factor_result, min_vol_pct):
    """
    Filtre 1 — Liquidité minimale.
    Garde les titres dont le volume moyen dépasse min_vol_pct% du volume max.
    """
    if liq_factor_result is None or min_vol_pct <= 0:
        return mf_scores.index.tolist()
    vol = liq_factor_result["Volume moyen"]
    threshold = vol.max() * (min_vol_pct / 100)
    liquid = vol[vol >= threshold].index.tolist()
    return [t for t in mf_scores.index if t in liquid]


def filter_mf_percentile(mf_scores, universe, top_pct):
    """
    Filtre 2 — Seuil Score MF.
    Garde les titres dans le top top_pct% de l'univers courant.
    """
    sub = mf_scores.reindex(universe).dropna()
    if sub.empty or top_pct >= 100:
        return universe
    threshold = sub.quantile(1 - top_pct / 100)
    return sub[sub >= threshold].index.tolist()


def filter_correlation(mf_scores, universe, cours_df, max_corr, window=252):
    """
    Filtre 3 — Diversification par corrélation.
    Clustering hiérarchique sur la matrice de corrélations.
    Dans chaque cluster, seul le titre avec le meilleur score MF est retenu.
    max_corr : seuil de corrélation (0→1). Plus bas = plus diversifié.
    """
    if len(universe) < 3 or max_corr >= 1.0:
        return universe

    ret = cours_df[universe].apply(pd.to_numeric, errors="coerce").pct_change().dropna(how="all")
    ret = ret.tail(window).dropna(axis=1, thresh=int(window * 0.5))
    valid = [t for t in universe if t in ret.columns]
    if len(valid) < 3:
        return universe

    corr = ret[valid].corr().fillna(0)
    dist = np.clip(1 - corr.values, 0, 2)
    np.fill_diagonal(dist, 0)
    condensed = squareform(dist)

    Z = linkage(condensed, method="ward")
    # Nombre de clusters : distance seuil = 1 - max_corr
    labels = fcluster(Z, t=(1 - max_corr), criterion="distance")

    # Pour chaque cluster, garder le titre au meilleur score MF
    selected = []
    sub_mf = mf_scores.reindex(valid)
    for cluster_id in np.unique(labels):
        members = [valid[i] for i, l in enumerate(labels) if l == cluster_id]
        best = sub_mf.reindex(members).idxmax()
        if pd.notna(best):
            selected.append(best)
    return selected


def optimize_markowitz(mf_scores, universe, cours_df,
                       window=252, risk_aversion=1.0,
                       mf_weight=0.5, min_w=0.0, max_w=1.0):
    """
    Filtre 4 — Optimisation Mean-Variance avec intégration du score MF.
    Maximise : mf_weight * MF_score_portefeuille - (1-mf_weight) * λ * variance
    Retourne les poids optimaux (Series ticker→poids).
    """
    ret = cours_df[universe].apply(pd.to_numeric, errors="coerce").pct_change().dropna(how="all")
    ret = ret.tail(window).dropna(axis=1, thresh=int(window * 0.5))
    valid = [t for t in universe if t in ret.columns]

    if len(valid) < 2:
        # Pas assez de données → revenir aux poids MF rank
        return None

    ret_valid = ret[valid]
    mu  = ret_valid.mean().values       # rendements moyens
    Sigma = ret_valid.cov().values      # matrice de covariance
    mf_vec = mf_scores.reindex(valid).fillna(0).values  # scores MF normalisés

    n = len(valid)
    w0 = np.ones(n) / n                # départ équipondéré

    def neg_utility(w):
        port_mf  = mf_weight * (w @ mf_vec)
        port_var = (1 - mf_weight) * risk_aversion * (w @ Sigma @ w)
        return -(port_mf - port_var)

    constraints = [{"type": "eq", "fun": lambda w: w.sum() - 1}]
    bounds = [(min_w, max_w)] * n

    res = minimize(neg_utility, w0, method="SLSQP",
                   bounds=bounds, constraints=constraints,
                   options={"maxiter": 500, "ftol": 1e-9})

    if not res.success:
        return None

    w_opt = pd.Series(res.x, index=valid)
    w_opt = w_opt[w_opt > 1e-4]        # drop poids négligeables
    w_opt = w_opt / w_opt.sum()        # renormalise
    return w_opt.sort_values(ascending=False)


def compute_portfolio_weights(mf, included=None):
    """
    α(T,t) = (n − r(T,t) + 1) / (n·(n+1)/2)
    Si included est fourni, seuls ces titres entrent dans le portefeuille.
    """
    if included:
        mf = mf.reindex(included).dropna()
    n = len(mf)
    if n == 0:
        return pd.Series(dtype=float)
    ranks = mf.rank(ascending=False, method='min')
    w = (n - ranks + 1) / (n * (n+1) / 2)
    return w.sort_values(ascending=False)


def run_optimization_pipeline(mf_scores, data, factor_results,
                              min_vol_pct, top_pct, max_corr,
                              use_markowitz, risk_aversion, mf_weight,
                              min_w, max_w, window,
                              use_sector_constraint=False, max_sector_pct=0.40,
                              use_corr_constraint=False, max_avg_corr=0.50):
    """
    Pipeline séquentiel — 4 filtres de base + 2 contraintes optionnelles.

    Contrainte ④ (optionnelle) — Sectorielle :
      Après pondération par rang MF, si un secteur dépasse max_sector_pct,
      on retire itérativement les titres les moins bien scorés du secteur
      en excès jusqu'à respecter le seuil.

    Contrainte ⑤ (optionnelle) — Corrélation effective :
      Parmi les titres restants, on sélectionne le sous-ensemble qui
      minimise la corrélation moyenne du portefeuille tout en conservant
      les meilleurs scores MF (greedy forward selection).
    """
    log = []
    universe = mf_scores.index.tolist()
    log.append(("Univers initial", len(universe), universe))

    # ── Filtre 1 — Liquidité ───────────────────────────────────
    liq = factor_results.get("Liquidité")
    universe = filter_liquidity(mf_scores, liq, min_vol_pct)
    log.append(("① Filtre Liquidité", len(universe), universe))

    # ── Filtre 2 — Score MF ────────────────────────────────────
    universe = filter_mf_percentile(mf_scores, universe, top_pct)
    log.append(("② Filtre Score MF", len(universe), universe))

    # ── Filtre 3 — Corrélation paires ─────────────────────────
    universe = filter_correlation(mf_scores, universe,
                                  data["cours"], max_corr, window)
    log.append(("③ Filtre Corrélation", len(universe), universe))

    if len(universe) == 0:
        return [], None, log

    # ── Contrainte ④ — Sectorielle (optionnelle) ───────────────
    if use_sector_constraint and SECTOR_MAP:
        universe_work = list(universe)
        changed = True
        while changed:
            changed = False
            # Calculer les poids provisoires par rang MF
            w_tmp = compute_portfolio_weights(mf_scores, included=universe_work)
            # Répartition sectorielle
            sect_w = {}
            for t, w in w_tmp.items():
                s = SECTOR_MAP.get(t.upper(), "Autre")
                sect_w[s] = sect_w.get(s, 0) + float(w)
            # Trouver les secteurs en excès
            for sec, sw in sect_w.items():
                if sw > max_sector_pct:
                    # Titres de ce secteur dans l'univers, triés par score MF (desc)
                    sec_tickers = sorted(
                        [t for t in universe_work
                         if SECTOR_MAP.get(t.upper(), "Autre") == sec],
                        key=lambda t: float(mf_scores.get(t, 0)),
                        reverse=True
                    )
                    # Retirer le moins bien scoré du secteur
                    if len(sec_tickers) > 1:
                        universe_work.remove(sec_tickers[-1])
                        changed = True
                        break  # recalculer
        universe = universe_work
        log.append(("④ Contrainte Sectorielle", len(universe), universe))

    if len(universe) == 0:
        return [], None, log

    # ── Contrainte ⑤ — Corrélation effective (optionnelle) ─────
    if use_corr_constraint and len(universe) > 2:
        try:
            BENCH = {'Unnamed: 0','ANNEE','JOUR','Unnamed: 62','.BRVMCI',
                     'BRVM30','BRVM PREST','BRVM-PRINC','BRVM-C TR',
                     'BRVM-CB','BRVM-CD','BRVM-ENER','BRVM-SFIN','BRVM-SPUB','Date'}
            cours_cl = data["cours"].apply(pd.to_numeric, errors="coerce")
            ticker_cl = [c for c in cours_cl.columns if c not in BENCH]
            ret_cl = cours_cl[ticker_cl].tail(window).ffill().pct_change()\
                              .clip(-0.20, 0.30).dropna(how="all")

            avail = [t for t in universe if t in ret_cl.columns]
            if len(avail) > 3:
                corr_full = ret_cl[avail].corr()

                # Greedy forward selection : commence par le meilleur scorer
                # Ajoute le titre suivant qui minimise la corrélation moyenne
                ranked = [t for t in mf_scores.loc[avail]
                           .sort_values(ascending=False).index]
                selected = [ranked[0]]

                for candidate in ranked[1:]:
                    # Corrélation moyenne du candidat avec les titres déjà sélectionnés
                    corrs_with = [float(corr_full.loc[candidate, s])
                                  for s in selected
                                  if candidate in corr_full.index
                                  and s in corr_full.columns]
                    avg_new = np.mean(corrs_with) if corrs_with else 0

                    if avg_new <= max_avg_corr:
                        selected.append(candidate)

                # Garder au moins 3 titres
                if len(selected) >= 3:
                    universe = selected

        except Exception:
            pass  # En cas d'erreur : garder l'univers tel quel

        log.append(("⑤ Corrélation effective", len(universe), universe))

    if len(universe) == 0:
        return [], None, log

    # ── Pondération finale ─────────────────────────────────────
    if use_markowitz:
        weights = optimize_markowitz(mf_scores, universe, data["cours"],
                                     window=window, risk_aversion=risk_aversion,
                                     mf_weight=mf_weight, min_w=min_w, max_w=max_w)
        if weights is None:
            # Fallback : poids MF rank
            weights = compute_portfolio_weights(mf_scores, included=universe)
            log.append(("④ Markowitz (fallback → rang MF)", len(weights), weights.index.tolist()))
        else:
            log.append(("④ Markowitz", len(weights), weights.index.tolist()))
    else:
        weights = compute_portfolio_weights(mf_scores, included=universe)
        log.append(("④ Poids rang MF", len(weights), weights.index.tolist()))

    return universe, weights, log


# ─── MULTIFACTOR & PORTFOLIO ──────────────────────────────────

def compute_multifactor(factor_results, betas):
    all_t = set()
    for df in factor_results.values():
        if df is not None:
            all_t.update(df.index)
    mf = pd.Series(0.0, index=list(all_t))
    for fname, df in factor_results.items():
        if df is None:
            continue
        sc = [c for c in df.columns if "Score" in c]
        if not sc:
            continue
        mf = mf.add(df[sc[0]] * betas.get(fname, 0), fill_value=0)
    return mf.sort_values(ascending=False)

def optimize_betas_ols(data, train_start, train_end,
                       target_start, target_end, year):
    """
    Approche 1 — Régression OLS.
    β_i = coefficients de régression linéaire qui minimisent :
    ||Rendement_T - Σ β_i * F_i(T)||²
    Les β obtenus représentent la contribution marginale de chaque facteur
    pour expliquer les rendements historiques.
    """
    FACTOR_NAMES = ["Value","Momentum","Volatilité","Dividende","Liquidité"]
    X, y, tickers = build_ml_dataset(
        data, train_start, train_end, target_start, target_end, year
    )
    if X is None or len(X) < 6:
        return None, None

    from numpy.linalg import lstsq
    X_arr = X.values
    y_arr = y.values

    # OLS : β = (XᵀX)⁻¹ Xᵀy
    coeffs, _, _, _ = lstsq(X_arr, y_arr, rcond=None)
    coeffs_s = pd.Series(coeffs, index=FACTOR_NAMES)

    # Normalisation : on prend les valeurs absolues et on renormalise
    # (un coefficient négatif = facteur inversement lié → on garde l'info mais β ≥ 0)
    abs_coeffs = coeffs_s.abs().clip(lower=0)
    total = abs_coeffs.sum()
    if total <= 0:
        return None, None
    betas = (abs_coeffs / total).to_dict()

    # R² manuel
    y_pred = X_arr @ coeffs
    ss_res = ((y_arr - y_pred)**2).sum()
    ss_tot = ((y_arr - y_arr.mean())**2).sum()
    r2 = 1 - ss_res/ss_tot if ss_tot > 0 else 0

    return betas, {
        "coefficients": coeffs_s.to_dict(),
        "r2": float(r2),
        "n_obs": len(y),
    }


def optimize_betas_walkforward(data, train_start, train_end,
                                target_start, target_end, year,
                                n_windows=4):
    """
    Approche 2 — Walk-forward (fenêtres glissantes + maximisation Sharpe).
    Divise [train_start → train_end] en n_windows sous-périodes.
    Pour chaque sous-période, calcule les scores factoriels et les rendements
    de la période suivante. Retourne les β moyens qui ont donné le meilleur
    Sharpe ratio cumulé.
    """
    FACTOR_NAMES = ["Value","Momentum","Volatilité","Dividende","Liquidité"]

    ts = pd.to_datetime(train_start)
    te = pd.to_datetime(train_end)
    tgs = pd.to_datetime(target_start)
    tge = pd.to_datetime(target_end)

    total_days = (tge - ts).days
    if total_days < 365 or n_windows < 2:
        return None, None

    # Découper en n_windows fenêtres glissantes
    window_days = total_days // (n_windows + 1)
    all_betas = []
    window_results = []

    for i in range(n_windows):
        w_train_start = ts + pd.Timedelta(days=i * window_days)
        w_train_end   = w_train_start + pd.Timedelta(days=window_days)
        w_target_start= w_train_end
        w_target_end  = w_target_start + pd.Timedelta(days=window_days)

        if w_target_end > tge:
            break

        X, y, _ = build_ml_dataset(
            data,
            w_train_start.date(), w_train_end.date(),
            w_target_start.date(), w_target_end.date(),
            year
        )
        if X is None or len(X) < 5:
            continue

        # Sharpe simple de chaque facteur sur la fenêtre
        sharpes = {}
        for f in FACTOR_NAMES:
            if f not in X.columns:
                sharpes[f] = 0
                continue
            scores = X[f]
            # Rendement du facteur = corrélation score × rendement réalisé
            corr = scores.corr(y)
            sharpes[f] = max(float(corr), 0)

        total_sh = sum(sharpes.values())
        if total_sh > 0:
            betas_w = {f: v/total_sh for f, v in sharpes.items()}
            all_betas.append(betas_w)
            window_results.append({
                "fenetre": i+1,
                "train": f"{w_train_start.date()} → {w_train_end.date()}",
                "target": f"{w_target_start.date()} → {w_target_end.date()}",
                "sharpes": sharpes,
            })

    if not all_betas:
        return None, None

    # Moyenne des β sur toutes les fenêtres
    avg_betas = {}
    for f in FACTOR_NAMES:
        avg_betas[f] = float(np.mean([b.get(f, 0) for b in all_betas]))

    total = sum(avg_betas.values())
    if total > 0:
        avg_betas = {f: v/total for f, v in avg_betas.items()}

    return avg_betas, {"n_windows": len(all_betas), "detail": window_results}


def vote_majority_betas(results_dict):
    """
    Vote majoritaire par facteur entre les 3 approches.
    Pour chaque facteur : β_final = médiane des β des approches disponibles.
    Puis renormalisation pour que Σβ = 1.
    """
    FACTOR_NAMES = ["Value","Momentum","Volatilité","Dividende","Liquidité"]
    available = {name: betas for name, betas in results_dict.items()
                 if betas is not None}
    if not available:
        return {f: 1/5 for f in FACTOR_NAMES}

    median_betas = {}
    for f in FACTOR_NAMES:
        vals = [b.get(f, 0) for b in available.values()]
        median_betas[f] = float(np.median(vals))

    total = sum(median_betas.values())
    if total > 0:
        median_betas = {f: v/total for f, v in median_betas.items()}
    return median_betas
    """
    α(T,t) = (n − r(T,t) + 1) / (n·(n+1)/2)
    Si included est fourni, seuls ces titres entrent dans le portefeuille.
    """
    if included:
        mf = mf.reindex(included).dropna()
    n = len(mf)
    if n == 0:
        return pd.Series(dtype=float)
    ranks = mf.rank(ascending=False, method='min')
    w = (n - ranks + 1) / (n * (n+1) / 2)
    return w.sort_values(ascending=False)


# ─── ML BETA OPTIMISATION ─────────────────────────────────────

def compute_scores_on_window(data, train_start, train_end, year):
    """
    Recalcule les 5 scores factoriels sur la plage [train_start, train_end]
    spécifiquement pour le ML — indépendant des onglets facteurs.

    Retourne un DataFrame (tickers × facteurs) avec les scores normalisés.
    """
    FACTOR_NAMES = ["Value", "Momentum", "Volatilité", "Dividende", "Liquidité"]
    scores = {}

    cours  = data["cours"].apply(pd.to_numeric, errors="coerce")
    ts, te = pd.to_datetime(train_start), pd.to_datetime(train_end)
    mask   = (cours.index >= ts) & (cours.index <= te)
    sub    = cours.loc[mask]
    if sub.empty:
        return None

    tickers = [t for t in data["tickers"] if t in sub.columns]

    # ── VALUE ─────────────────────────────────────────────────
    nb  = data["nb_titres"]
    cms = sub.mean(numeric_only=True)
    val_scores = {}
    for t in tickers:
        if t not in nb or nb[t] <= 0:
            continue
        cm = cms.get(t, np.nan)
        if pd.isna(cm) or cm <= 0:
            continue
        n   = nb[t]
        cp  = data["capitaux_propres"].get(t, {}).get(year, np.nan)
        rn  = data["resultat_net"].get(t, {}).get(year, np.nan)
        fcf = data["flux_treso"].get(t, {}).get(year, np.nan)
        ca  = data["chiffre_affaires"].get(t, {}).get(year, np.nan)
        rex = data["resultat_expl"].get(t, {}).get(year, np.nan)
        metrics = {
            "bp":  (cp/n)/cm  if pd.notna(cp)  else np.nan,
            "ep":  (rn/n)/cm  if pd.notna(rn)  else np.nan,
            "fp":  (fcf/n)/cm if pd.notna(fcf) else np.nan,
            "cp":  (ca/n)/cm  if pd.notna(ca)  else np.nan,
            "ebp": (rex/n)/cm if pd.notna(rex) else np.nan,
        }
        val_scores[t] = metrics
    if val_scores:
        vdf = pd.DataFrame(val_scores).T
        vdf = vdf.apply(pd.to_numeric, errors="coerce")
        norm_v = {}
        for m in vdf.columns:
            mx = vdf[m].max()
            norm_v[m] = vdf[m]/mx if (pd.notna(mx) and mx != 0) else vdf[m]*0
        norm_vdf = pd.DataFrame(norm_v)
        scores["Value"] = norm_vdf.mean(axis=1).fillna(0)

    # ── MOMENTUM (rendement moyen sur la fenêtre) ──────────────
    ret = sub.pct_change().dropna(how="all")
    if not ret.empty:
        mom = ret.mean(numeric_only=True).replace([np.inf,-np.inf], np.nan).dropna()
        mx  = mom.max()
        if pd.notna(mx) and mx != 0:
            scores["Momentum"] = (mom/mx).clip(0)
        else:
            scores["Momentum"] = mom*0

    # ── VOLATILITÉ (inversée) ───────────────────────────────────
    vol = ret.std(numeric_only=True).dropna()
    if not vol.empty:
        mn = vol.min()
        if pd.notna(mn) and mn > 0:
            scores["Volatilité"] = (mn/vol).replace([np.inf,-np.inf], np.nan).fillna(0)

    # ── DIVIDENDE ──────────────────────────────────────────────
    div = data["dividendes"]
    dr  = div[div["Date"] == year]
    if not dr.empty:
        yields = {}
        for t in tickers:
            if t not in dr.columns:
                continue
            d = dr[t].values[0]
            p = cms.get(t, np.nan)
            if pd.notna(d) and pd.notna(p) and p > 0:
                yields[t] = float(d)/float(p)
        if yields:
            s  = pd.Series(yields)
            mx = s.max()
            scores["Dividende"] = (s/mx if mx > 0 else s).clip(0)

    # ── LIQUIDITÉ ──────────────────────────────────────────────
    vol_df = data["volumes"].apply(pd.to_numeric, errors="coerce")
    vmask  = (vol_df.index >= ts) & (vol_df.index <= te)
    vsub   = vol_df.loc[vmask]
    if not vsub.empty:
        avg = vsub.mean(numeric_only=True).replace([np.inf,-np.inf], np.nan).dropna()
        mx  = avg.max()
        if pd.notna(mx) and mx > 0:
            scores["Liquidité"] = (avg/mx).clip(0)

    if not scores:
        return None

    score_df = pd.DataFrame(scores).reindex(columns=FACTOR_NAMES).fillna(0)
    return score_df


def build_ml_dataset(data, train_start, train_end, target_start, target_end, year):
    """
    Construit le dataset ML avec recalcul complet des scores sur la fenêtre ML :

      Features X  : F_i(T) recalculés sur [train_start → train_end]
                    → indépendant des plages définies dans les onglets facteurs
      Cible Y     : rendement total du titre T sur [target_start → target_end]

    Retourne (X_df, y_series, tickers) ou (None, None, None).
    """
    # ── Features : recalcul des scores sur la fenêtre d'entraînement ──
    X_df = compute_scores_on_window(data, train_start, train_end, year)
    if X_df is None or X_df.empty:
        return None, None, None

    # ── Cible : rendement total sur la fenêtre cible ────────────────
    cours = data["cours"].apply(pd.to_numeric, errors="coerce")
    mask  = (cours.index >= pd.to_datetime(target_start)) & \
            (cours.index <= pd.to_datetime(target_end))
    sub   = cours.loc[mask]

    if sub.empty or len(sub) < 2:
        return None, None, None

    returns = {}
    for t in X_df.index:
        if t not in sub.columns:
            continue
        s = sub[t].dropna()
        if len(s) < 2:
            continue
        returns[t] = (s.iloc[-1] - s.iloc[0]) / s.iloc[0]

    if len(returns) < 5:
        return None, None, None

    y      = pd.Series(returns)
    common = X_df.index.intersection(y.index)
    if len(common) < 5:
        return None, None, None

    return X_df.loc[common], y.loc[common], list(common)


def optimize_betas_ml(data, train_start, train_end,
                      target_start, target_end, year, n_estimators=200):
    """
    Pipeline complet :
      1. Recalcule les scores factoriels sur [train_start → train_end]
      2. Cible Y = rendements réalisés sur [target_start → target_end]
      3. Entraîne RF + GB, retourne les importances normalisées comme β_i
    """
    FACTOR_NAMES = ["Value", "Momentum", "Volatilité", "Dividende", "Liquidité"]

    X, y, tickers = build_ml_dataset(
        data, train_start, train_end,
        target_start, target_end, year
    )
    if X is None:
        return None, None, None

    scaler = StandardScaler()
    X_sc   = scaler.fit_transform(X)

    results = {}

    # ── Random Forest ──────────────────────────────────────────
    rf = RandomForestRegressor(
        n_estimators=n_estimators,
        max_features="sqrt",
        min_samples_leaf=2,
        random_state=42,
        n_jobs=-1
    )
    rf.fit(X_sc, y)
    rf_imp = pd.Series(rf.feature_importances_, index=FACTOR_NAMES)
    results["Random Forest"] = {
        "importances": rf_imp,
        "r2": rf.score(X_sc, y),
    }

    # ── Gradient Boosting ──────────────────────────────────────
    gb = GradientBoostingRegressor(
        n_estimators=n_estimators,
        max_depth=3,
        learning_rate=0.05,
        subsample=0.8,
        random_state=42
    )
    gb.fit(X_sc, y)
    gb_imp = pd.Series(gb.feature_importances_, index=FACTOR_NAMES)
    results["Gradient Boosting"] = {
        "importances": gb_imp,
        "r2": gb.score(X_sc, y),
    }

    # ── β combinés ────────────────────────────────────────────
    combined  = (rf_imp * 0.5 + gb_imp * 0.5).clip(lower=0)
    total     = combined.sum()
    betas_opt = (combined / total).to_dict() if total > 0 \
                else {f: 1/5 for f in FACTOR_NAMES}

    return betas_opt, results, {
        "X": X, "y": y, "tickers": tickers,
        "train_window": f"{train_start} → {train_end}",
        "target_window": f"{target_start} → {target_end}",
        "year_fundamentals": year,
    }

# ─── PLOT HELPERS ─────────────────────────────────────────────
PLOT_LAYOUT = dict(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                   font=dict(color="#94a3b8", family="JetBrains Mono"),
                   margin=dict(l=10, r=10, t=20, b=60),
                   xaxis=dict(gridcolor="#1e2d45", tickangle=-45),
                   yaxis=dict(gridcolor="#1e2d45"))

def score_bar(series, color_end, height=360, fmt=".4f", yformat=None):
    top = series.head(25)
    fig = go.Figure(go.Bar(
        x=top.index, y=top.values,
        marker=dict(color=top.values,
                    colorscale=[[0,"#1e2d45"],[1, color_end]], line=dict(width=0)),
        text=[f"{v:{fmt}}" for v in top.values],
        textposition="outside",
        textfont=dict(size=9, color="#94a3b8"),
    ))
    layout = {**PLOT_LAYOUT, "height": height}
    if yformat:
        layout["yaxis"] = dict(gridcolor="#1e2d45", tickformat=yformat)
    fig.update_layout(**layout)
    return fig

# ─── SESSION STATE ────────────────────────────────────────────
for k in ["data","factor_results","mf_scores","pw"]:
    if k not in st.session_state:
        st.session_state[k] = {} if k == "factor_results" else None

# β sliders — initialisés inconditionnellement au démarrage
for k, default in [("sv_val",0.20),("sv_mom",0.20),("sv_vol",0.20),
                   ("sv_div",0.20),("sv_liq",0.20)]:
    if k not in st.session_state:
        st.session_state[k] = default

# ── Chargement persistant depuis GitHub ───────────────────────
# Au démarrage : on tente de charger les données depuis GitHub
# Si disponibles → l'utilisateur n'a pas besoin de recharger ses fichiers
if "github_loaded" not in st.session_state:
    st.session_state.github_loaded = False

if not st.session_state.github_loaded and GITHUB_STORAGE:
    # Guard: ne charger qu'une seule fois par session
    st.session_state.github_loaded = True  # marquer immédiatement pour éviter la boucle
    try:
        # 1. Données SMF
        if st.session_state.data is None:
            smf_gh = load_smf_data()
            if smf_gh:
                st.session_state.data = smf_gh

        # 2. États financiers
        if not st.session_state.get("fin_data"):
            fin_gh = load_financial_db_github()
            if fin_gh:
                st.session_state.fin_data = fin_gh

        # 3. Charia (stocké temporairement, fusionné avec exclusions fixes ensuite)
        charia_gh = load_charia_results()
        if charia_gh:
            st.session_state["_charia_from_github"] = charia_gh
    except Exception:
        pass  # silencieux — l'app fonctionne sans persistance

# Screening Charia — exclusions fixes + données GitHub
if "charia_results" not in st.session_state:
    if CHARIA_AVAILABLE:
        from charia_screening import BANK_TICKERS, ILLICIT_SECTOR_TICKERS
        _init_charia = {}
        for t in BANK_TICKERS:
            _init_charia[t] = {
                "compatible": False, "n_standards": 0,
                "halal_sector": False, "excluded": True,
                "raison": "Banque conventionnelle — incompatible Charia (modèle Riba)",
                "standards": {s: {"pass": False} for s in ["DJIM","FTSE","S&P","AAOIFI"]},
            }
        for t, raison in ILLICIT_SECTOR_TICKERS.items():
            _init_charia[t] = {
                "compatible": False, "n_standards": 0,
                "halal_sector": False, "excluded": True,
                "raison": f"Secteur illicite — {raison}",
                "standards": {s: {"pass": False} for s in ["DJIM","FTSE","S&P","AAOIFI"]},
            }
        # Fusionner avec les données GitHub (les exclusions fixes ont priorité)
        charia_from_gh = st.session_state.pop("_charia_from_github", {})
        merged_charia = {**charia_from_gh, **_init_charia}
        st.session_state.charia_results = merged_charia
    else:
        st.session_state.charia_results = {}

# Valuation DB path
VALUATION_DB_PATH = "financial_db.json"
if "fin_data" not in st.session_state:
    st.session_state.fin_data = load_financial_db(VALUATION_DB_PATH) \
                                 if VALUATION_AVAILABLE else {}

# ── Chargement de la classification sectorielle ───────────────
@st.cache_data
def load_sector_mapping():
    """Charge sectors.json et retourne {ticker: secteur}."""
    candidates = [
        os.path.join(_APP_DIR, "sectors.json"),
        "sectors.json",
    ]
    for path in candidates:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
            mapping = {}
            for sector, tickers in raw.items():
                for t in tickers:
                    mapping[t.strip().upper()] = sector
            return mapping
    return {}

SECTOR_MAP = load_sector_mapping()

# ─── SIDEBAR ──────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style='text-align:center;padding:14px 0 8px 0;'>
      <div style='font-family:Syne,sans-serif;font-size:20px;font-weight:800;
           background:linear-gradient(135deg,#3b82f6,#06b6d4);
           -webkit-background-clip:text;-webkit-text-fill-color:transparent;'>CGF GESTION</div>
      <div style='font-size:10px;color:#475569;letter-spacing:.12em;'>STRATÉGIE MULTIFACTORIELLE BRVM</div>
    </div><hr style='margin:8px 0 14px 0;'>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("📁 Base_de_données_-_SMF.xlsx",
                                 type=["xlsx","xls"], label_visibility="visible")

    # Indicateur persistance GitHub
    if GITHUB_STORAGE and is_github_configured():
        loaded_from_gh = st.session_state.github_loaded
        if loaded_from_gh:
            st.caption("☁️ Données chargées depuis GitHub")
        elif st.session_state.data:
            st.caption("💾 Données en session · non persistées")
        else:
            st.caption("☁️ GitHub configuré · chargement en cours...")
    else:
        if not GITHUB_STORAGE:
            st.caption("⚠️ github_storage.py non trouvé")
        else:
            st.caption("⚙️ Configurez GitHub Secrets pour la persistance")
    if uploaded:
        try:
            st.session_state.data = load_data(uploaded.read())
            n_tickers = len(st.session_state.data['tickers'])
            st.success(f"✅ {n_tickers} titres chargés")
            # Auto-sauvegarde GitHub silencieuse (pas de spinner = pas de rerun)
            if GITHUB_STORAGE and is_github_configured():
                ok = save_smf_data(st.session_state.data)
                if ok:
                    st.caption("☁️ Sauvegardé sur GitHub")
        except Exception as e:
            st.error(f"Erreur : {e}")

    data = st.session_state.data
    if data:
        st.markdown("---")
        # Années fondamentaux disponibles (CP, RN, etc.) — exclut 2025 si absent
        fund_years = data.get("fundamental_years", [])
        if fund_years:
            selected_year = st.selectbox(
                "📅 Année fondamentaux (Value / Dividende)",
                fund_years,
                help="Années pour lesquelles les données de Capitaux propres et Résultat net sont disponibles dans le Tableau de bord."
            )
            st.caption(f"Données fondamentaux disponibles : {', '.join(str(y) for y in fund_years)}")
        else:
            selected_year = 2024
            st.warning("Aucune année de fondamentaux détectée.")
        date_ref = st.date_input("📅 Date ref. (Momentum / Vol.)", value=data["cours"].index[-1].date())
        st.markdown("---")
        st.markdown("**⚖️ Poids des facteurs β_i**")
        st.caption("Calibrez chaque facteur · Σβ doit = 1.0")

        b_val = st.slider("💰 Value",      0.0, 1.0, st.session_state.sv_val, 0.01, key="b_val")
        b_mom = st.slider("🚀 Momentum",   0.0, 1.0, st.session_state.sv_mom, 0.01, key="b_mom")
        b_vol = st.slider("📉 Volatilité", 0.0, 1.0, st.session_state.sv_vol, 0.01, key="b_vol")
        b_div = st.slider("💸 Dividende",  0.0, 1.0, st.session_state.sv_div, 0.01, key="b_div")
        b_liq = st.slider("💧 Liquidité",  0.0, 1.0, st.session_state.sv_liq, 0.01, key="b_liq")

        bs = round(b_val + b_mom + b_vol + b_div + b_liq, 4)
        if abs(bs - 1.0) <= 0.01:
            st.success(f"✅ Σβ = {bs:.2f}")
        else:
            st.warning(f"⚠️ Σβ = {bs:.2f} ≠ 1.0")

        if st.button("⚖️ Égaliser (20% chacun)"):
            # Écriture dans les clés de stockage (pas les clés widget)
            for k in ["sv_val","sv_mom","sv_vol","sv_div","sv_liq"]:
                st.session_state[k] = 0.20
            st.rerun()

        betas = {
            "Value":      b_val,
            "Momentum":   b_mom,
            "Volatilité": b_vol,
            "Dividende":  b_div,
            "Liquidité":  b_liq,
        }
        st.session_state.betas = betas
    else:
        selected_year, date_ref = 2024, None
        betas = st.session_state.get("betas",
                {"Value":0.20,"Momentum":0.20,"Volatilité":0.20,"Dividende":0.20,"Liquidité":0.20})

    st.markdown("---")
    st.markdown("""<div style='font-size:10px;color:#475569;line-height:1.6;'>
    Note Technique CGF Gestion · 10/05/2024<br>Auteur : A.B.A.M. Gueye</div>""", unsafe_allow_html=True)

    # ── Manuel d'utilisation intégré ──────────────────────────
    st.markdown("---")
    st.markdown("**📖 Manuel d'utilisation**")

    @st.cache_data(show_spinner=False)
    def generate_manual_pdf():
        """Génère le PDF du manuel et retourne les bytes."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable, KeepTogether
        )
        import datetime, io as _io

        BLUE   = colors.HexColor("#3b82f6")
        CYAN   = colors.HexColor("#06b6d4")
        GREEN  = colors.HexColor("#10b981")
        AMBER  = colors.HexColor("#f59e0b")
        RED    = colors.HexColor("#ef4444")
        PURPLE = colors.HexColor("#8b5cf6")
        LIGHT  = colors.HexColor("#f0f4ff")
        MUTED  = colors.HexColor("#64748b")
        DARK   = colors.HexColor("#1e293b")
        WHITE  = colors.white
        PW     = A4[0]

        def sty(name, **kw):
            return ParagraphStyle(name, **kw)

        S = {
            "cover_title": sty("ct", fontName="Helvetica-Bold", fontSize=26,
                               textColor=DARK, alignment=TA_CENTER, spaceAfter=6, leading=32),
            "cover_sub":   sty("cs", fontName="Helvetica", fontSize=12,
                               textColor=BLUE, alignment=TA_CENTER, spaceAfter=4),
            "cover_date":  sty("cd", fontName="Helvetica", fontSize=9,
                               textColor=MUTED, alignment=TA_CENTER, spaceAfter=10),
            "h1":  sty("h1", fontName="Helvetica-Bold", fontSize=15,
                       textColor=BLUE, spaceBefore=16, spaceAfter=7, leading=19),
            "h2":  sty("h2", fontName="Helvetica-Bold", fontSize=12,
                       textColor=DARK, spaceBefore=12, spaceAfter=5, leading=15),
            "h3":  sty("h3", fontName="Helvetica-Bold", fontSize=10,
                       textColor=MUTED, spaceBefore=8, spaceAfter=4, leading=13),
            "body": sty("body", fontName="Helvetica", fontSize=9.5,
                        textColor=DARK, alignment=TA_JUSTIFY, spaceAfter=5, leading=14),
            "bl":   sty("bl", fontName="Helvetica", fontSize=9.5,
                        textColor=DARK, spaceAfter=4, leading=14, leftIndent=14),
            "cap":  sty("cap", fontName="Helvetica-Oblique", fontSize=8.5,
                        textColor=MUTED, alignment=TA_CENTER, spaceAfter=4),
            "frm":  sty("frm", fontName="Courier-Bold", fontSize=9.5,
                        textColor=colors.HexColor("#1e3a5f"),
                        backColor=colors.HexColor("#dbeafe"),
                        alignment=TA_CENTER, spaceAfter=7, leading=15, borderPad=8),
            "toc":  sty("toc", fontName="Helvetica", fontSize=10.5,
                        textColor=DARK, spaceAfter=5, leading=17),
            "tocs": sty("tocs", fontName="Helvetica", fontSize=9.5,
                        textColor=MUTED, spaceAfter=4, leading=15, leftIndent=18),
        }

        def sp(n=1): return Spacer(1, n*0.32*cm)
        def hr(c=colors.HexColor("#e2e8f0"), w=0.5):
            return HRFlowable(width="100%", thickness=w, color=c, spaceAfter=6, spaceBefore=3)

        def badge(text, col=BLUE):
            t = Table([[Paragraph(f'<font color="white"><b>{text}</b></font>', S["bl"])]],
                      colWidths=[PW-4*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),col),
                ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
                ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
            ]))
            return t

        def ibox(text, bg=LIGHT, border=BLUE):
            t = Table([[Paragraph(text, S["bl"])]], colWidths=[PW-4*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),bg),
                ("LINEBEFORE",(0,0),(0,-1),3,border),
                ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                ("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),10),
            ]))
            return t

        def tbl(headers, rows, col_w=None):
            data = [headers] + rows
            n = len(headers)
            cw = col_w or ([(PW-4*cm)/n]*n)
            t = Table(data, colWidths=cw, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),BLUE),
                ("TEXTCOLOR",(0,0),(-1,0),WHITE),
                ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,-1),8.5),
                ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LIGHT]),
                ("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#cbd5e1")),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                ("LEFTPADDING",(0,0),(-1,-1),7),
            ]))
            return t

        def ph(text): return Paragraph(f"<b>{text}</b>", S["bl"])
        def pb(text): return Paragraph(text, S["bl"])

        buf = _io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2.2*cm, bottomMargin=2.2*cm,
            title="Manuel CGF Gestion — SMF BRVM")

        story = []
        today = datetime.date.today().strftime("%d/%m/%Y")

        # ── COVER ────────────────────────────────────────────────
        story += [
            Spacer(1, 2.5*cm),
            Paragraph("MANUEL D'UTILISATION", S["cover_title"]),
            sp(0.3),
            Paragraph("Moteur d'Allocation Multifactoriel BRVM", S["cover_sub"]),
            Paragraph(f"CGF Gestion · Note Technique 10/05/2024 · v1.0 · {today}", S["cover_date"]),
            sp(),
            hr(BLUE, 1.5), sp(0.5),
        ]
        cov = Table([[pb("Application : Streamlit"), pb("Auteur : A.B.A.M. Gueye"),
                      pb("Marché : BRVM — 48 titres"), pb("Facteurs : 5 (Value · Mom · Vol · Div · Liq)")]],
                    colWidths=[(PW-4*cm)/4]*4)
        cov.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),LIGHT),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#cbd5e1")),
            ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("LEFTPADDING",(0,0),(-1,-1),8),("VALIGN",(0,0),(-1,-1),"TOP"),
            ("FONTSIZE",(0,0),(-1,-1),8.5)]))
        story += [cov, sp(2), hr(BLUE,1.5), Spacer(1,3*cm),
            Paragraph("Ce document décrit l'utilisation complète de l'application de gestion "
                "de portefeuille multifactorielle CGF Gestion, couvrant le chargement des "
                "données, le calcul des 5 facteurs, l'indice MF et l'optimisation du "
                "portefeuille selon la Note Technique du 10/05/2024.", S["body"]),
            PageBreak()]

        # ── TOC ──────────────────────────────────────────────────
        story += [Paragraph("TABLE DES MATIÈRES", S["h1"]), hr(), sp(0.3)]
        toc = [
            ("1.", "Présentation générale et workflow", False),
            ("2.", "Démarrage et chargement des données", False),
            ("2.1", "Chargement du fichier Excel", True),
            ("2.2", "Structure des 6 feuilles", True),
            ("2.3", "Mise à jour des données", True),
            ("3.", "Barre latérale — Paramètres globaux", False),
            ("4.", "Onglet Value — Facteur de valorisation", False),
            ("4.1", "Formules et 5 métriques", True),
            ("4.2", "Paramétrage et interprétation", True),
            ("5.", "Onglet Momentum — Dynamique des cours", False),
            ("5.1", "Formule et 6 horizons", True),
            ("5.2", "Plages de dates libres par horizon", True),
            ("6.", "Onglet Volatilité — Facteur de risque", False),
            ("7.", "Onglet Dividende — Rendement", False),
            ("8.", "Onglet Liquidité — Volume transigé", False),
            ("9.", "Onglet Indice Multifactoriel", False),
            ("10.", "Onglet Portefeuille Cible", False),
            ("10.1", "Pipeline automatique — 4 filtres", True),
            ("10.2", "Mode sélection manuelle", True),
            ("10.3", "Pondération finale (Rang MF / Markowitz)", True),
            ("11.", "Export des résultats", False),
            ("12.", "Référence des formules mathématiques", False),
            ("13.", "Résolution des problèmes fréquents", False),
        ]
        for num, title, sub in toc:
            style = S["tocs"] if sub else S["toc"]
            indent = "    " if sub else ""
            story.append(Paragraph(f"{indent}<b>{num}</b>  {title}", style))
        story.append(PageBreak())

        # ── 1. PRÉSENTATION ──────────────────────────────────────
        story += [badge("1. PRÉSENTATION GÉNÉRALE ET WORKFLOW", BLUE), sp(),
            Paragraph("L'application implémente la Note Technique CGF Gestion du 10/05/2024 "
                "sur la stratégie multifactorielle. Elle suit un processus en 3 étapes :", S["body"])]

        steps = [
            ("1", BLUE, "Calcul des indices factoriels (Étape 1)",
             "Calculer F_i(t,T) pour chaque facteur et chaque titre. "
             "5 onglets dédiés : Value, Momentum, Volatilité, Dividende, Liquidité."),
            ("2", PURPLE, "Calcul de l'indice multifactoriel (Étape 2)",
             "Agréger : MF(t,T) = Σ β_i · F_i(t,T). Les β sont calibrés dans la sidebar."),
            ("3", GREEN, "Construction du portefeuille cible (Étape 3)",
             "Pondération par rang MF ou optimisation Markowitz sur l'univers filtré."),
        ]
        for num, col, title, desc in steps:
            r = Table([[Paragraph(f'<font color="white"><b>{num}</b></font>', S["bl"]),
                        Paragraph(f'<b>{title}</b><br/>{desc}', S["bl"])]],
                      colWidths=[0.9*cm, PW-4*cm-0.9*cm])
            r.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(0,0),col),("BACKGROUND",(1,0),(1,0),LIGHT),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("ALIGN",(0,0),(0,0),"CENTER"),
                ("TOPPADDING",(0,0),(-1,-1),9),("BOTTOMPADDING",(0,0),(-1,-1),9),
                ("LEFTPADDING",(0,0),(-1,-1),9),
            ]))
            story += [sp(0.3), r]
        story += [sp(), PageBreak()]

        # ── 2. DONNÉES ───────────────────────────────────────────
        story += [badge("2. DÉMARRAGE ET CHARGEMENT DES DONNÉES", BLUE), sp(),
            Paragraph("2.1  Chargement du fichier Excel", S["h2"]),
            Paragraph("Dans la barre latérale, cliquez sur Browse files et sélectionnez "
                "Base_de_données_-_SMF.xlsx. Le message ✅ 48 titres chargés confirme "
                "le succès.", S["body"]),
            sp(0.3),
            Paragraph("2.2  Structure des 6 feuilles requises", S["h2"]),
            tbl([ph("Feuille"), ph("Contenu"), ph("Utilisée pour")],
                [[pb("Cours"), pb("Cours journaliers de clôture"), pb("Momentum · Volatilité")],
                 [pb("Volume moyen"), pb("Volumes journaliers transigés"), pb("Liquidité")],
                 [pb("Historique_dividende"), pb("Dividendes annuels versés"), pb("Dividende")],
                 [pb("Moyenne_cours"), pb("Cours moyen annuel"), pb("Référence interne")],
                 [pb("Nb_titres"), pb("Nombre de titres en circulation"), pb("Value (unitaire)")],
                 [pb("Tableau de bord"), pb("5 blocs : CP · RN · FCF · CA · Rex"), pb("Value")]],
                col_w=[(PW-4*cm)*0.28, (PW-4*cm)*0.42, (PW-4*cm)*0.30]),
            sp(0.5),
            Paragraph("2.3  Mise à jour des données (ex. ajout 2025)", S["h2"]),
            Paragraph("Ajoutez une ligne 2025 dans chacun des 5 blocs du Tableau de bord "
                "(CAPITAUX PROPRES, RESULTAT NET, FLUX DE TRESORERIE, CHIFFRE D'AFFAIRES, "
                "RESULTAT D'EXPLOITATION) en respectant le format des années 2019–2024. "
                "L'année 2025 apparaîtra automatiquement dans le sélecteur de la sidebar.", S["body"]),
            sp(0.3),
            ibox("⚠️  Les années proposées dans la sidebar sont limitées aux années disposant "
                "simultanément de Capitaux propres ET de Résultat net dans le Tableau de bord.",
                colors.HexColor("#fef3c7"), AMBER),
            PageBreak()]

        # ── 3. SIDEBAR ───────────────────────────────────────────
        story += [badge("3. BARRE LATÉRALE — PARAMÈTRES GLOBAUX", MUTED), sp(),
            Paragraph("La sidebar est permanente et accessible depuis tous les onglets.", S["body"]),
            Paragraph("Année fondamentaux", S["h2"]),
            Paragraph("Sélectionne l'année des données CP/RN/FCF/CA/Rex pour Value et Dividende. "
                "Limitée aux années réellement disponibles dans votre fichier.", S["body"]),
            Paragraph("Poids des facteurs β_i  (Σβ = 1.0)", S["h2"]),
            tbl([ph("Curseur"), ph("Facteur"), ph("Rôle")],
                [[pb("💰 β Value"), pb("Valorisation"), pb("Poids du score Value dans MF")],
                 [pb("🚀 β Momentum"), pb("Dynamique"), pb("Poids du score Momentum dans MF")],
                 [pb("📉 β Volatilité"), pb("Risque"), pb("Poids du score Volatilité dans MF")],
                 [pb("💸 β Dividende"), pb("Rendement"), pb("Poids du score Dividende dans MF")],
                 [pb("💧 β Liquidité"), pb("Négociabilité"), pb("Poids du score Liquidité dans MF")]]),
            sp(0.3),
            ibox("💡  Le bouton ⚖️ Égaliser remet tous les β à 0.20. Les poids sont persistés "
                "entre les onglets via le session_state de Streamlit.", LIGHT, BLUE),
            PageBreak()]

        # ── 4. VALUE ─────────────────────────────────────────────
        story += [badge("4. ONGLET VALUE — FACTEUR DE VALORISATION", BLUE), sp(),
            Paragraph("4.1  Formules et 5 métriques", S["h2"]),
            Paragraph("F_value(t,T) = w1·(B/P) + w2·(E/P) + w3·(FCF/P) + w4·(CA/P) + w5·(EBIT/P)", S["frm"]),
            Paragraph("Chaque métrique est normalisée : m_ij / max_E(m_ij). Score élevé = titre sous-évalué.", S["cap"]),
            sp(0.3),
            tbl([ph("Métrique"), ph("Formule"), ph("Poids défaut"), ph("Interprétation")],
                [[pb("B/P"), pb("(CP/n) / Cours moyen"), pb("20%"), pb("Valeur comptable/prix — ↑ = sous-évalué")],
                 [pb("E/P"), pb("(RN/n) / Cours moyen"), pb("20%"), pb("Inverse du P/E — ↑ = bon marché")],
                 [pb("FCF/P"), pb("(FCF/n) / Cours moyen"), pb("20%"), pb("Rendement des flux — ↑ = générateur de cash")],
                 [pb("CA/P"), pb("(CA/n) / Cours moyen"), pb("20%"), pb("Revenus/prix — ↑ = chiffre d'affaires élevé")],
                 [pb("EBIT/P"), pb("(Rex/n) / Cours moyen"), pb("20%"), pb("Profitabilité opérationnelle/prix")]],
                col_w=[(PW-4*cm)*0.12,(PW-4*cm)*0.30,(PW-4*cm)*0.16,(PW-4*cm)*0.42]),
            sp(0.5),
            Paragraph("4.2  Paramétrage et interprétation", S["h2"]),
            Paragraph("→  Période du cours moyen : choisissez Date début et Date fin. "
                "Le cours moyen est la moyenne arithmétique des cours journaliers sur la période. "
                "Les fondamentaux proviennent de l'année sélectionnée en sidebar.", S["bl"]),
            Paragraph("→  Pondérations : ajustez les 5 poids (somme = 1.0). "
                "Un poids à 0 exclut la métrique du calcul.", S["bl"]),
            Paragraph("→  Cliquez ⚙️ Calculer le Facteur Value pour lancer. "
                "Les résultats incluent KPIs, graphique de classement et table exportable.", S["bl"]),
            sp(0.3),
            ibox("⚠️  Score Value = 0.0000 : aucune métrique disponible pour ce titre "
                "sur l'année sélectionnée. Vérifiez le Tableau de bord dans votre fichier Excel.",
                colors.HexColor("#fef3c7"), AMBER),
            PageBreak()]

        # ── 5. MOMENTUM ──────────────────────────────────────────
        story += [badge("5. ONGLET MOMENTUM — DYNAMIQUE DES COURS", PURPLE), sp(),
            Paragraph("5.1  Formule", S["h2"]),
            Paragraph("F_mom(t,T) = Σ w_h · rdt_moy_h(t,T) / max_E(rdt_moy_h)", S["frm"]),
            Paragraph("rdt_moy_h = rendement journalier moyen sur la plage [debut_h, fin_h]", S["cap"]),
            sp(0.3),
            Paragraph("5.2  Plages de dates libres par horizon", S["h2"]),
            Paragraph("Chaque horizon dispose de son propre sélecteur de plage de dates :", S["body"]),
            sp(0.2),
            tbl([ph("Horizon"), ph("Plage par défaut"), ph("Exemple avancé")],
                [[pb("Journalier"), pb("J-1 → Aujourd'hui"), pb("Toujours 1 jour")],
                 [pb("Hebdo"), pb("J-7 → Aujourd'hui"), pb("Semaine de référence")],
                 [pb("Mensuel"), pb("J-30 → Aujourd'hui"), pb("Mois de référence")],
                 [pb("Trimestriel"), pb("J-91 → Aujourd'hui"), pb("Dernier trimestre")],
                 [pb("Semestriel"), pb("J-182 → Aujourd'hui"), pb("6 derniers mois")],
                 [pb("Annuel"), pb("J-365 → Aujourd'hui"), pb("5 ans : 20/05/2021 → 20/05/2026")]]),
            sp(0.3),
            ibox("💡  Pour désactiver un horizon, mettez son poids à 0. Les poids restants "
                "sont automatiquement renormalisés pour que leur somme reste égale à 1.0.", LIGHT, BLUE),
            PageBreak()]

        # ── 6·7·8 ────────────────────────────────────────────────
        story += [badge("6. VOLATILITÉ  |  7. DIVIDENDE  |  8. LIQUIDITÉ", MUTED), sp(),
            Paragraph("Facteur Volatilité (onglet 📉)", S["h2"]),
            Paragraph("F_vol(t,T) = min_E(sigma) / sigma(T)    [formule inversée]", S["frm"]),
            Paragraph("sigma = ecart-type des rendements journaliers sur la fenetre choisie (60–504 jours)", S["cap"]),
            Paragraph("Un titre peu volatil reçoit un score élevé. La fenêtre est ajustable "
                "via un curseur (défaut = 252 jours). Le graphique compare la volatilité "
                "annualisée (sigma × sqrt(252)) de tous les titres.", S["body"]),
            sp(0.5),
            Paragraph("Facteur Dividende (onglet 💸)", S["h2"]),
            Paragraph("F_div(t,T) = 1.0 x DY(t,T) / max_E(DY)    [poids unique = 100%]", S["frm"]),
            Paragraph("DY = Dividende annuel verse / Cours moyen sur la periode choisie", S["cap"]),
            Paragraph("Paramètres : année du dividende (issu de Historique_dividende) + "
                "plage de dates pour le cours moyen (indépendante). "
                "Un titre non-payeur reçoit DY = 0 et score = 0.", S["body"]),
            sp(0.5),
            Paragraph("Facteur Liquidité (onglet 💧)", S["h2"]),
            Paragraph("F_liq(t,T) = 1.0 x Vol_moy(T) / max_E(Vol_moy)    [poids unique = 100%]", S["frm"]),
            Paragraph("Vol_moy = volume moyen journalier sur la periode choisie", S["cap"]),
            Paragraph("Plage de dates entièrement libre sur l'historique disponible. "
                "Par défaut : historique complet.", S["body"]),
            PageBreak()]

        # ── 9. INDICE MF ─────────────────────────────────────────
        story += [badge("9. ONGLET INDICE MULTIFACTORIEL", PURPLE), sp(),
            Paragraph("MF(t,T) = beta_Value·F_Value + beta_Mom·F_Mom + beta_Vol·F_Vol + beta_Div·F_Div + beta_Liq·F_Liq", S["frm"]),
            Paragraph("Avec Σ beta_i = 1  ·  Les beta sont calibres dans la barre laterale", S["cap"]),
            sp(0.3),
            Paragraph("L'onglet affiche les β actifs avec leur statut (✓ calculé / ✗ non calculé). "
                "Il faut cliquer sur 🔢 Calculer l'Indice MF après chaque modification des β. "
                "Résultats : classement en barres, décomposition factorielle empilée Top 15 "
                "(contribution β_i·F_i de chaque facteur), table exportable.", S["body"]),
            sp(0.3),
            ibox("💡  Recalculez l'indice MF après toute modification des β en sidebar ou "
                "après recalcul d'un facteur dans ses onglets dédiés.", LIGHT, BLUE),
            PageBreak()]

        # ── 10. PORTEFEUILLE ─────────────────────────────────────
        story += [badge("10. ONGLET PORTEFEUILLE CIBLE", GREEN), sp(),
            Paragraph("10.1  Pipeline automatique — 4 filtres séquentiels", S["h2"]),
            tbl([ph("Filtre"), ph("Paramètre"), ph("Objectif"), ph("Défaut")],
                [[pb("① Liquidité"), pb("Volume ≥ X% du max"), pb("Exclut les titres peu liquides"), pb("0%")],
                 [pb("② Score MF"), pb("Top X% des scores"), pb("Ne garde que les mieux classés"), pb("60%")],
                 [pb("③ Corrélation"), pb("Corr. max entre titres"), pb("Clustering — 1 par groupe"), pb("0.75")],
                 [pb("④ Poids"), pb("Rang MF ou Markowitz"), pb("Pondération optimale"), pb("Rang MF")]],
                col_w=[(PW-4*cm)*0.18,(PW-4*cm)*0.27,(PW-4*cm)*0.33,(PW-4*cm)*0.22]),
            sp(0.3),
            Paragraph("La trace du pipeline s'affiche sous forme de cartes montrant le "
                "nombre de titres survivant à chaque filtre. Si le résultat est 0 titres, "
                "élargissez les seuils (Liquidité → 0%, Top MF → 100%, Corrélation → 1.0).", S["body"]),
            sp(0.5),
            Paragraph("10.2  Mode sélection manuelle", S["h2"]),
            Paragraph("Activez l'expander 🔧 Sélection manuelle pour choisir les titres "
                "titre par titre, court-circuitant les 3 premiers filtres. "
                "La pondération finale reste au choix.", S["body"]),
            sp(0.5),
            Paragraph("10.3  Pondération finale — Rang MF vs Markowitz", S["h2"]),
            tbl([ph("Méthode"), ph("Formule"), ph("Quand l'utiliser")],
                [[pb("Rang MF (défaut)"),
                  pb("alpha(T) = (n-r+1) / (n(n+1)/2)"),
                  pb("Approche conforme à la Note Technique · Simple · Reproductible")],
                 [pb("Markowitz MF-augmenté"),
                  pb("max{ lambda·Score_MF - (1-lambda)·variance }"),
                  pb("Optimise rendement/risque · Prend en compte les corrélations")]],
                col_w=[(PW-4*cm)*0.25,(PW-4*cm)*0.38,(PW-4*cm)*0.37]),
            sp(0.3),
            Paragraph("Markowitz : paramètres supplémentaires = poids Score MF dans l'utilité "
                "(0→1), aversion au risque λ, poids min/max par titre.", S["body"]),
            PageBreak()]

        # ── 11. EXPORT ───────────────────────────────────────────
        story += [badge("11. EXPORT DES RÉSULTATS", BLUE), sp(),
            tbl([ph("Onglet"), ph("Fichier Excel exporté"), ph("Contenu")],
                [[pb("Value"), pb("value_{annee}_{debut}_{fin}.xlsx"), pb("Scores + 5 métriques + cours moyen")],
                 [pb("Momentum"), pb("momentum.xlsx"), pb("Scores + rendements moyens par horizon")],
                 [pb("Dividende"), pb("dividende_{annee}.xlsx"), pb("Scores + Dividend Yield + cours moyen")],
                 [pb("Liquidité"), pb("liquidite_{debut}_{fin}.xlsx"), pb("Scores + volumes moyens")],
                 [pb("Indice MF"), pb("classement_MF.xlsx"), pb("Rang + Score MF par titre")],
                 [pb("Portefeuille"), pb("portefeuille_optimal_BRVM.xlsx"), pb("Rang MF + poids alpha + poids %")]],
                col_w=[(PW-4*cm)*0.18,(PW-4*cm)*0.40,(PW-4*cm)*0.42]),
            PageBreak()]

        # ── 12. FORMULES ─────────────────────────────────────────
        story += [badge("12. RÉFÉRENCE DES FORMULES MATHÉMATIQUES", DARK), sp(),
            Paragraph("Indice factoriel — formule générale", S["h2"]),
            Paragraph("F_i(t,T) = SUM(j=1..k_i) w_ij * m_ij(t,T) / max_Et(m_ij(t,T))", S["frm"]),
            Paragraph("Indice factoriel — cas Volatilité (inversé)", S["h2"]),
            Paragraph("F_vol(t,T) = SUM(j=1..k_i) w_ij * min_Et(m_ij(t,T)) / m_ij(t,T)", S["frm"]),
            Paragraph("Indice Multifactoriel", S["h2"]),
            Paragraph("MF(t,T) = SUM(i=1..5) beta_i * F_i(t,T)    avec SUM(beta_i) = 1", S["frm"]),
            Paragraph("Pondération portefeuille cible", S["h2"]),
            Paragraph("alpha(T,t) = (n(t) - r(T,t) + 1) / ( n(t) * (n(t)+1) / 2 )", S["frm"]),
            sp(0.3),
            tbl([ph("Variable"), ph("Définition")],
                [[pb("k_i"), pb("Nombre de métriques composant le facteur i")],
                 [pb("m_ij(t,T)"), pb("Valeur de la métrique j du facteur i pour le titre T à la date t")],
                 [pb("w_ij"), pb("Poids de la métrique j dans le facteur i  (Σw_ij = 1)")],
                 [pb("E_t"), pb("Ensemble des titres admissibles à la date t")],
                 [pb("beta_i"), pb("Poids du facteur i dans l'indice MF  (Σβ_i = 1)")],
                 [pb("r(T,t)"), pb("Rang du titre T selon MF (r=1 = meilleur score)")],
                 [pb("n(t)"), pb("Nombre total de titres admissibles à la date t")],
                 [pb("alpha(T,t)"), pb("Pondération du titre T dans le portefeuille cible")]],
                col_w=[(PW-4*cm)*0.20, (PW-4*cm)*0.80]),
            PageBreak()]

        # ── 13. PROBLÈMES ────────────────────────────────────────
        story += [badge("13. RÉSOLUTION DES PROBLÈMES FRÉQUENTS", RED), sp()]
        problems = [
            ("❌ \"Données insuffisantes\" dans Value",
             ['L\'année choisie (ex: 2025) n\'a pas de fondamentaux dans le Tableau de bord.',
              'Solution : choisissez une année disponible (2019–2024) ou renseignez les données 2025.']),
            ("❌ Score Value = 0.0000 pour tous les titres",
             ['Toutes les métriques affichent None : les fondamentaux ne sont pas lus.',
              'Solution : vérifiez la structure du Tableau de bord (5 blocs, tickers en ligne 0 col 2+).']),
            ("❌ Erreur TypeError sur la Liquidité",
             ['La feuille Volume moyen contient des colonnes non numériques (totaux, formules).',
              'Solution : supprimez les colonnes parasites et rechargez le fichier.']),
            ("❌ \"Aucun horizon valide\" dans Momentum",
             ['Toutes les plages de Momentum ont Date début >= Date fin.',
              'Solution : vérifiez que la date début est strictement antérieure à la date fin.']),
            ("❌ Pipeline Portefeuille retourne 0 titres",
             ['Les filtres sont trop restrictifs.',
              'Solution : réduisez Liquidité → 0%, montez Top MF → 100%, Corrélation → 1.0.']),
            ("❌ StreamlitAPIException sur les sélecteurs de dates",
             ['La valeur par défaut dépasse les bornes min/max de la date.',
              'Solution : rechargez la page (F5) — ce cas est normalement géré automatiquement.']),
        ]
        for title, items in problems:
            story += [KeepTogether(
                [Paragraph(title, S["h3"])] +
                [Paragraph(f"→  {item}", S["bl"]) for item in items] +
                [sp(0.4)]
            )]

        story += [hr(BLUE,1), sp(0.3),
            Paragraph(f"CGF Gestion · Moteur Multifactoriel BRVM · v1.0 · {today}", S["cap"])]

        doc.build(story)
        buf.seek(0)
        return buf.read()

    # Génère le PDF (mis en cache)
    pdf_bytes = generate_manual_pdf()

    # Bouton de téléchargement
    st.download_button(
        label="⬇️ Télécharger le manuel (PDF)",
        data=pdf_bytes,
        file_name="Manuel_CGF_SMF_BRVM.pdf",
        mime="application/pdf",
        use_container_width=True,
    )

    # Viewer inline dans un expander
    with st.expander("👁️ Consulter le manuel", expanded=False):
        import base64
        b64 = base64.b64encode(pdf_bytes).decode("utf-8")
        st.markdown(
            f'<iframe src="data:application/pdf;base64,{b64}" '
            f'width="100%" height="600px" style="border:1px solid #1e2d45;border-radius:6px;">'
            f'</iframe>',
            unsafe_allow_html=True,
        )

    # ── Rapport de résultats ───────────────────────────────────
    st.markdown("---")
    st.markdown("**📋 Rapport de résultats**")

    def generate_rapport_pdf(data, factor_results, mf_scores, pw,
                             betas, charia_results):
        """
        Génère un PDF de rapport complet avec :
        - Page de garde CGF Gestion
        - Définition et résultats de chaque facteur
        - Classement MF et méthode de calcul
        - Allocation du portefeuille cible
        - Portefeuille Charia si disponible
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable, KeepTogether
        )
        import datetime, io as _io

        buf = _io.BytesIO()
        W, H = A4
        today = datetime.date.today().strftime("%d/%m/%Y")

        # ── Palette couleurs ──────────────────────────────────
        C_BG    = colors.HexColor("#0b0f1a")
        C_BLUE  = colors.HexColor("#3b82f6")
        C_CYAN  = colors.HexColor("#06b6d4")
        C_GREEN = colors.HexColor("#10b981")
        C_AMBER = colors.HexColor("#f59e0b")
        C_TEXT  = colors.HexColor("#1e293b")
        C_LIGHT = colors.HexColor("#f8fafc")
        C_GRAY  = colors.HexColor("#64748b")
        C_LINE  = colors.HexColor("#e2e8f0")

        FACTOR_COLORS = {
            "Value":      colors.HexColor("#3b82f6"),
            "Momentum":   colors.HexColor("#8b5cf6"),
            "Volatilité": colors.HexColor("#10b981"),
            "Dividende":  colors.HexColor("#f59e0b"),
            "Liquidité":  colors.HexColor("#06b6d4"),
        }

        FACTOR_ICONS = {
            "Value":      "Value",
            "Momentum":   "Momentum",
            "Volatilite": "Volatilite",
            "Dividende":  "Dividende",
            "Liquidite":  "Liquidite",
        }

        FACTOR_DEFINITIONS = {
            "Value": (
                "Le facteur Value identifie les titres sous-évalués par le marché "
                "en comparant leur prix à leurs fondamentaux financiers.",
                [
                    ("B/P",    "Valeur comptable / Cours",           "10%"),
                    ("E/P",    "Résultat net / Cours (inverse PER)", "50%"),
                    ("FCF/P",  "Free Cash-Flow / Cours",             "20%"),
                    ("CA/P",   "Chiffre d'affaires / Cours",         "10%"),
                    ("EBIT/EV","Résultat exploitation / Val. entrep.","10%"),
                ]
            ),
            "Momentum": (
                "Le facteur Momentum capture la persistance des tendances de cours "
                "en mesurant les rendements moyens sur plusieurs horizons temporels.",
                [
                    ("Rdt Journalier",    "Rendement moyen sur 1 jour",       "16.67%"),
                    ("Rdt Hebdomadaire",  "Rendement moyen sur 1 semaine",     "16.67%"),
                    ("Rdt Mensuel",       "Rendement moyen sur 1 mois",        "16.67%"),
                    ("Rdt Trimestriel",   "Rendement moyen sur 3 mois",        "16.67%"),
                    ("Rdt Semestriel",    "Rendement moyen sur 6 mois",        "16.67%"),
                    ("Rdt Annuel",        "Rendement moyen sur 12 mois",       "16.67%"),
                ]
            ),
            "Volatilite": (
                "Le facteur Volatilité favorise les titres à faible risque en mesurant "
                "l'écart-type des rendements. L'indice est calculé de façon inversée "
                "(min/valeur) pour récompenser la faible volatilité.",
                [("Ecart-type", "Ecart-type des rendements journaliers", "100%")]
            ),
            "Dividende": (
                "Le facteur Dividende identifie les titres offrant un rendement "
                "locatif attractif, mesuré par le ratio dividende / cours.",
                [("Dividend Yield", "Dividende par action / Cours", "100%")]
            ),
            "Liquidite": (
                "Le facteur Liquidité favorise les titres facilement négociables "
                "sur le marché BRVM, mesurée par le volume moyen transigé.",
                [("Volume moyen", "Volume moyen de transactions", "100%")]
            ),
        }

        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=1.8*cm, bottomMargin=1.8*cm,
            leftMargin=2*cm, rightMargin=2*cm
        )

        # ── Styles ────────────────────────────────────────────
        SS = getSampleStyleSheet()
        def S(name, **kw):
            base = SS.get(name, SS["Normal"])
            return ParagraphStyle(name+"_custom", parent=base, **kw)

        sTitle    = S("Title",    fontSize=22, textColor=C_BLUE,
                      fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=6)
        sSubtitle = S("Normal",   fontSize=12, textColor=C_GRAY,
                      fontName="Helvetica", alignment=TA_CENTER, spaceAfter=4)
        sH1       = S("Heading1", fontSize=14, textColor=C_BLUE,
                      fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
        sH2       = S("Heading2", fontSize=11, textColor=C_TEXT,
                      fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=4)
        sBody     = S("Normal",   fontSize=9,  textColor=C_TEXT,
                      fontName="Helvetica", leading=14, spaceAfter=4,
                      alignment=TA_JUSTIFY)
        sCap      = S("Normal",   fontSize=8,  textColor=C_GRAY,
                      fontName="Helvetica-Oblique", spaceAfter=2)
        sCell     = S("Normal",   fontSize=8,  textColor=C_TEXT,
                      fontName="Helvetica", leading=11)
        sCellB    = S("Normal",   fontSize=8,  textColor=C_TEXT,
                      fontName="Helvetica-Bold", leading=11)
        sCellC    = S("Normal",   fontSize=8,  textColor=C_BLUE,
                      fontName="Helvetica-Bold", leading=11)

        def table_style(header_color=C_BLUE):
            return TableStyle([
                ("BACKGROUND",  (0,0), (-1,0),  header_color),
                ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
                ("FONTNAME",    (0,0), (-1,0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0),  8),
                ("ALIGN",       (0,0), (-1,-1), "LEFT"),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[C_LIGHT, colors.white]),
                ("GRID",        (0,0), (-1,-1), 0.3, C_LINE),
                ("LEFTPADDING", (0,0), (-1,-1), 6),
                ("RIGHTPADDING",(0,0), (-1,-1), 6),
                ("TOPPADDING",  (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("FONTSIZE",    (0,1), (-1,-1), 8),
            ])

        story = []

        # ══════════════════════════════════════════════════════
        # PAGE DE GARDE
        # ══════════════════════════════════════════════════════
        story += [
            Spacer(1, 2*cm),
            Paragraph("CGF GESTION", S("Title", fontSize=28, textColor=C_BLUE,
                       fontName="Helvetica-Bold", alignment=TA_CENTER)),
            Spacer(1, 0.3*cm),
            HRFlowable(width="80%", thickness=2, color=C_BLUE, spaceAfter=8),
            Spacer(1, 0.5*cm),
            Paragraph("RAPPORT D'ALLOCATION MULTIFACTORIELLE", sTitle),
            Paragraph("Bourse Régionale des Valeurs Mobilières (BRVM)", sSubtitle),
            Spacer(1, 0.3*cm),
            Paragraph(f"Généré le {today}", sCap),
            Spacer(1, 1.5*cm),
        ]

        # Bloc résumé
        n_tickers = len(data.get("tickers", [])) if data else 0
        n_factors = sum(1 for v in factor_results.values() if v is not None)
        n_ptf     = len(pw) if pw is not None and not pw.empty else 0
        n_charia  = len([t for t,r in charia_results.items()
                         if r and r.get("compatible")]) if charia_results else 0

        summary_data = [
            ["Univers BRVM", "Facteurs calculés",
             "Titres en portefeuille", "Titres Charia compatibles"],
            [str(n_tickers), str(n_factors), str(n_ptf), str(n_charia)],
        ]
        t_sum = Table(summary_data, colWidths=[4*cm]*4)
        t_sum.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,0), C_BLUE),
            ("TEXTCOLOR",    (0,0),(-1,0), colors.white),
            ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0),(-1,0), 8),
            ("BACKGROUND",   (0,1),(-1,1), colors.HexColor("#dbeafe")),
            ("FONTNAME",     (0,1),(-1,1), "Helvetica-Bold"),
            ("FONTSIZE",     (0,1),(-1,1), 16),
            ("TEXTCOLOR",    (0,1),(-1,1), C_BLUE),
            ("ALIGN",        (0,0),(-1,-1),"CENTER"),
            ("VALIGN",       (0,0),(-1,-1),"MIDDLE"),
            ("GRID",         (0,0),(-1,-1), 0.3, C_LINE),
            ("TOPPADDING",   (0,0),(-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ]))
        story += [t_sum, Spacer(1, 1*cm)]

        # β actifs
        if betas:
            beta_data = [["Facteur", "Beta (b_i)", "Contribution"],
                         *[[k, f"{v:.2f}", f"{v*100:.1f}%"]
                           for k,v in betas.items()]]
            t_beta = Table(beta_data, colWidths=[5*cm, 4*cm, 4*cm])
            t_beta.setStyle(table_style(C_CYAN))
            story += [
                Paragraph("Pondérations des facteurs (b_i)", sH2),
                t_beta,
            ]

        story.append(PageBreak())

        # ══════════════════════════════════════════════════════
        # SECTION 1 — FACTEURS
        # ══════════════════════════════════════════════════════
        story.append(Paragraph("1. ANALYSE FACTORIELLE", sH1))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=C_LINE, spaceAfter=8))
        story.append(Paragraph(
            "La stratégie multifactorielle repose sur l'exploitation de cinq facteurs "
            "de risque et de performance. Chaque facteur est calculé comme une "
            "combinaison pondérée de métriques financières et de marché, normalisées "
            "par la valeur maximale observée dans l'univers BRVM à la date de calcul "
            "(formule Note Technique CGF Gestion, 10/05/2024).", sBody))
        story.append(Spacer(1, 0.3*cm))

        FACTOR_ORDER = ["Value","Momentum","Volatilite","Dividende","Liquidite"]
        FACTOR_NAMES_MAP = {
            "Value":"Value","Momentum":"Momentum",
            "Volatilite":"Volatilité","Dividende":"Dividende","Liquidite":"Liquidité"
        }

        for fkey in FACTOR_ORDER:
            fname = FACTOR_NAMES_MAP[fkey]
            fcolor= FACTOR_COLORS.get(fname, C_BLUE)
            defn, metrics = FACTOR_DEFINITIONS[fkey]

            # En-tête facteur
            story.append(KeepTogether([
                Paragraph(fname, S("Heading2", fontSize=12,
                          textColor=fcolor, fontName="Helvetica-Bold",
                          spaceBefore=10, spaceAfter=4)),
                Paragraph(defn, sBody),
                Spacer(1, 0.2*cm),
            ]))

            # Tableau des métriques
            m_data = [["Métrique", "Description", "Poids"],
                      *[[m[0], m[1], m[2]] for m in metrics]]
            t_m = Table(m_data, colWidths=[3.5*cm, 9*cm, 2*cm])
            t_m.setStyle(table_style(fcolor))
            story.append(t_m)

            # Résultats si disponibles
            fr_key = fname
            if fr_key in factor_results and factor_results[fr_key] is not None:
                df_f = factor_results[fr_key]
                sc_cols = [c for c in df_f.columns if "Score" in c]
                if sc_cols:
                    top10 = df_f[sc_cols[0]].nlargest(10)
                    res_data = [["Rang", "Ticker", f"Score {fname}"]] + \
                               [[str(i+1), t, f"{v:.6f}"]
                                for i, (t, v) in enumerate(top10.items())]
                    t_res = Table(res_data, colWidths=[1.5*cm, 4*cm, 9*cm])
                    t_res.setStyle(table_style(fcolor))
                    story += [
                        Spacer(1, 0.2*cm),
                        Paragraph(f"Top 10 — Scores {fname}", sCap),
                        t_res,
                    ]
            story.append(Spacer(1, 0.4*cm))

        story.append(PageBreak())

        # ══════════════════════════════════════════════════════
        # SECTION 2 — INDICE MULTIFACTORIEL
        # ══════════════════════════════════════════════════════
        story.append(Paragraph("2. INDICE MULTIFACTORIEL (MF)", sH1))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=C_LINE, spaceAfter=8))
        story.append(Paragraph(
            "L'indice multifactoriel MF(t,T) est calculé comme la somme pondérée "
            "des indices factoriels : MF(t,T) = somme des b_i x F_i(t,T) "
            "pour i allant de 1 a 7. Les b_i représentent les poids "
            "attribués à chaque facteur.", sBody))

        if mf_scores is not None and not mf_scores.empty:
            top20 = mf_scores.head(20)
            n_total = len(mf_scores)
            mf_data = [["Rang", "Ticker", "Score MF", "Centile"]] + \
                      [[str(i+1), t, f"{v:.6f}",
                        f"Top {(i+1)/n_total*100:.0f}%"]
                       for i, (t, v) in enumerate(top20.items())]
            t_mf = Table(mf_data, colWidths=[1.5*cm, 3*cm, 7*cm, 3*cm])
            t_mf.setStyle(table_style(C_BLUE))
            story += [
                Paragraph("Classement MF — Top 20 titres", sH2),
                t_mf,
                Spacer(1, 0.3*cm),
                Paragraph(
                    f"Univers complet : {n_total} titres classés. "
                    "Formule : alpha(T,t) = (n - r(T,t) + 1) / (n x (n+1) / 2)",
                    sCap),
            ]
        else:
            story.append(Paragraph(
                "Score MF non encore calculé — lancez l'Indice MF dans l'application.",
                sBody))

        story.append(PageBreak())

        # ══════════════════════════════════════════════════════
        # SECTION 3 — PORTEFEUILLE CIBLE
        # ══════════════════════════════════════════════════════
        story.append(Paragraph("3. PORTEFEUILLE CIBLE", sH1))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=C_LINE, spaceAfter=8))
        story.append(Paragraph(
            "Le portefeuille cible est construit en appliquant la formule de "
            "pondération par rang MF : alpha(T,t) = (n - r(T,t) + 1) / (n x (n+1) / 2). "
            "Les titres sont sélectionnés après application des filtres "
            "de liquidité, de score MF et de corrélation.", sBody))

        if pw is not None and not pw.empty and mf_scores is not None:
            ranks = mf_scores.reindex(pw.index).rank(
                ascending=False, method="min")
            ptf_data = [["Rang MF", "Ticker", "Secteur",
                         "Score MF", "Poids alpha", "Poids (%)"]]
            for t, w in pw.items():
                rk = int(ranks.get(t, 0)) if not pd.isna(ranks.get(t, 0)) else "-"
                sc = f"{mf_scores.get(t, 0):.6f}" \
                     if t in mf_scores.index else "-"
                sec = SECTOR_MAP.get(t.upper(), "Autre") if SECTOR_MAP else "-"
                # Charia label
                ch = charia_results.get(t, {})
                ch_ok = "oui" if ch.get("compatible") else \
                        ("exclu" if ch.get("excluded") else "non")
                ptf_data.append([str(rk), t, sec[:20],
                                  sc, f"{w:.6f}", f"{w*100:.2f}%"])

            t_ptf = Table(ptf_data,
                          colWidths=[1.5*cm, 2.5*cm, 4.5*cm, 4*cm, 2.5*cm, 1.5*cm])
            t_ptf.setStyle(table_style(C_BLUE))
            story += [
                Paragraph(f"Allocation — {len(pw)} titres retenus", sH2),
                t_ptf,
                Spacer(1, 0.3*cm),
                Paragraph(
                    f"Somme des poids : {pw.sum():.6f} · "
                    f"Titre en tête : {pw.index[0]} ({pw.iloc[0]*100:.2f}%) · "
                    f"Titre en queue : {pw.index[-1]} ({pw.iloc[-1]*100:.2f}%)",
                    sCap),
            ]
        else:
            story.append(Paragraph(
                "Portefeuille non encore construit — lancez le pipeline "
                "dans l'onglet Portefeuille.",
                sBody))

        # ══════════════════════════════════════════════════════
        # SECTION 4 — SCREENING CHARIA
        # ══════════════════════════════════════════════════════
        if charia_results:
            story.append(PageBreak())
            story.append(Paragraph("4. SCREENING CHARIATIQUE", sH1))
            story.append(HRFlowable(width="100%", thickness=1,
                                    color=C_LINE, spaceAfter=8))
            story.append(Paragraph(
                "Le screening Charia évalue la conformité de chaque titre selon "
                "4 standards islamiques reconnus : DJIM, FTSE, S&P et AAOIFI/Malaisie. "
                "Un titre est considéré compatible s'il qualifie sur au moins 3 "
                "standards sur 4. Les banques conventionnelles et les sociétés "
                "opérant dans des secteurs illicites sont automatiquement exclues.",
                sBody))

            compat   = [(t, r) for t, r in charia_results.items()
                        if r and r.get("compatible")]
            excluded = [(t, r) for t, r in charia_results.items()
                        if r and r.get("excluded")]
            non_compat = [(t, r) for t, r in charia_results.items()
                          if r and not r.get("compatible") and not r.get("excluded")]

            # Résumé Charia
            ch_sum_data = [
                ["Compatibles", "Exclus (secteur)", "Non conformes", "Total"],
                [str(len(compat)), str(len(excluded)),
                 str(len(non_compat)), str(len(charia_results))],
            ]
            t_ch_sum = Table(ch_sum_data, colWidths=[4*cm]*4)
            t_ch_sum.setStyle(TableStyle([
                ("BACKGROUND",   (0,0),(-1,0), C_GREEN),
                ("TEXTCOLOR",    (0,0),(-1,0), colors.white),
                ("FONTNAME",     (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0),(-1,0), 8),
                ("BACKGROUND",   (0,1),(-1,1), colors.HexColor("#d1fae5")),
                ("FONTNAME",     (0,1),(-1,1), "Helvetica-Bold"),
                ("FONTSIZE",     (0,1),(-1,1), 16),
                ("TEXTCOLOR",    (0,1),(-1,1), C_GREEN),
                ("ALIGN",        (0,0),(-1,-1),"CENTER"),
                ("VALIGN",       (0,0),(-1,-1),"MIDDLE"),
                ("GRID",         (0,0),(-1,-1), 0.3, C_LINE),
                ("TOPPADDING",   (0,0),(-1,-1), 8),
                ("BOTTOMPADDING",(0,0),(-1,-1), 8),
            ]))
            story += [t_ch_sum, Spacer(1, 0.4*cm)]

            # Tableau des compatibles
            if compat:
                ch_data = [["Ticker", "Standards", "RE/Actif",
                            "RC/Actif", "RL/Actif"]]
                for t, r in sorted(compat, key=lambda x: x[0]):
                    stds = "/".join(
                        s for s, v in r.get("standards", {}).items()
                        if v.get("pass"))
                    rats = r.get("ratios", {})
                    ch_data.append([
                        t, stds,
                        f"{rats.get('RE_actif',0):.3f}" if rats.get('RE_actif') else "-",
                        f"{rats.get('RC_actif',0):.3f}" if rats.get('RC_actif') else "-",
                        f"{rats.get('RL_actif',0):.3f}" if rats.get('RL_actif') else "-",
                    ])
                t_ch = Table(ch_data,
                             colWidths=[2.5*cm, 5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
                t_ch.setStyle(table_style(C_GREEN))
                story += [
                    Paragraph("Titres compatibles Charia", sH2),
                    t_ch,
                ]

            # Portefeuille Charia si disponible
            if pw is not None and not pw.empty and mf_scores is not None:
                charia_tickers = [t for t, r in charia_results.items()
                                  if r and r.get("compatible")]
                pw_ch = pw[pw.index.isin(charia_tickers)]
                if not pw_ch.empty:
                    story += [Spacer(1, 0.4*cm),
                              Paragraph("Portefeuille Charia (intersection)",sH2)]
                    pw_ch_norm = pw_ch / pw_ch.sum()
                    ch_ptf_data = [["Titre","Poids rebasé (%)"]]
                    for t, w in pw_ch_norm.items():
                        ch_ptf_data.append([t, f"{w*100:.2f}%"])
                    t_ch_ptf = Table(ch_ptf_data, colWidths=[5*cm, 5*cm])
                    t_ch_ptf.setStyle(table_style(C_GREEN))
                    story.append(t_ch_ptf)

        # ── Pied de page ──────────────────────────────────────
        story += [
            Spacer(1, 1*cm),
            HRFlowable(width="100%", thickness=0.5, color=C_LINE),
            Spacer(1, 0.2*cm),
            Paragraph(
                f"CGF Gestion · Rapport SMF BRVM · Généré le {today} · "
                "Note Technique 10/05/2024 · Confidentiel",
                S("Normal", fontSize=7, textColor=C_GRAY,
                  alignment=TA_CENTER, fontName="Helvetica-Oblique")),
        ]

        doc.build(story)
        buf.seek(0)
        return buf.read()

    # Bouton rapport
    if st.button("📋 Générer le rapport PDF", type="primary",
                 use_container_width=True):
        with st.spinner("Génération du rapport..."):
            try:
                rapport_bytes = generate_rapport_pdf(
                    data=st.session_state.data,
                    factor_results=st.session_state.factor_results,
                    mf_scores=st.session_state.mf_scores,
                    pw=st.session_state.pw,
                    betas=betas if data else {},
                    charia_results=st.session_state.charia_results,
                )
                st.session_state.rapport_pdf = rapport_bytes
                st.success("✅ Rapport généré")
            except Exception as e:
                st.error(f"Erreur génération PDF : {e}")

    if "rapport_pdf" in st.session_state and st.session_state.rapport_pdf:
        import datetime
        today_str = datetime.date.today().strftime("%Y%m%d")
        st.download_button(
            label="⬇️ Télécharger le rapport",
            data=st.session_state.rapport_pdf,
            file_name=f"Rapport_SMF_BRVM_{today_str}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )

# ─── HEADER ───────────────────────────────────────────────────
st.markdown("""
<div style='padding:6px 0 20px 0;'>
  <span class='pill'>Smart Beta · BRVM · 48 Titres</span>
  <p class='sh'>Moteur d'Allocation Multifactoriel</p>
  <p class='ss'>CGF Gestion · Note Technique 10/05/2024 · Facteurs : Value · Momentum · Volatilité · Dividende · Liquidité</p>
</div>""", unsafe_allow_html=True)

if not data:
    st.warning("⬅️ Veuillez charger **Base_de_données_-_SMF.xlsx** dans la barre latérale.")
    st.stop()

fr = st.session_state.factor_results

def run_backtest(cours_df, factor_results, betas,
                 start_date, end_date,
                 rebal_freq="Q",
                 charia_results=None,
                 fundamental_years=None,
                 data=None):
    """
    Backtesting du portefeuille multifactoriel BRVM.

    Stratégie :
      - À chaque rebalancement (trimestriel), recalcule les scores MF
        depuis les cours disponibles, puis pondère par rang MF.
      - Entre deux rebalancements : rendements des poids fixes.

    Benchmarks :
      - BRVM Composite : moyenne équipondérée de tous les titres
      - Portefeuille Charia : même stratégie mais filtré sur les
        titres compatibles Charia

    Retourne un dict avec les séries de performance et les métriques.
    """
    BENCH_COLS = {'Unnamed: 0','ANNEE','JOUR','Unnamed: 62',
                  '.BRVMCI','BRVM30','BRVM PREST','BRVM-PRINC',
                  'BRVM-C TR','BRVM-CB','BRVM-CD','BRVM-ENER',
                  'BRVM-SFIN','BRVM-SPUB','Date'}

    # Nettoyage des cours — forcer float, exclure colonnes indices
    cours = cours_df.apply(pd.to_numeric, errors="coerce")
    ticker_cols = [c for c in cours.columns if c not in BENCH_COLS]
    cours = cours[ticker_cols]

    # Filtrage période
    sd = pd.to_datetime(start_date)
    ed = pd.to_datetime(end_date)
    cours = cours[(cours.index >= sd) & (cours.index <= ed)]
    if cours.empty or len(cours) < 20:
        return None

    # Garder seulement les titres avec au moins 50% de données
    n_rows = len(cours)
    coverage = cours.notna().sum() / n_rows
    cours = cours[coverage[coverage >= 0.50].index]

    # Forward-fill puis backward-fill pour les trous ponctuels
    cours = cours.ffill().bfill()

    # Remplacer les zéros/négatifs par NaN (cours invalides)
    cours = cours.where(cours > 0, np.nan).ffill().bfill()

    # Rendements journaliers — clippés à [-20%, +30%] pour éviter les aberrations
    ret = cours.pct_change()
    ret = ret.clip(lower=-0.20, upper=0.30).fillna(0)

    # Dates de rebalancement
    rebal_dates = pd.date_range(
        start=cours.index[0], end=cours.index[-1], freq=rebal_freq
    )
    rebal_dates = [d for d in rebal_dates if d in cours.index or
                   cours.index[cours.index >= d].size > 0]

    def get_rebal_date(target):
        """Trouve la prochaine date de trading disponible."""
        future = cours.index[cours.index >= target]
        return future[0] if len(future) > 0 else None

    rebal_dates = [get_rebal_date(d) for d in rebal_dates
                   if get_rebal_date(d) is not None]
    rebal_dates = sorted(set(rebal_dates))

    def compute_weights_at(date, eligible=None):
        """Calcule les poids MF depuis les cours jusqu'à `date`."""
        sub = cours[cours.index <= date]
        if sub.empty:
            return None
        # Proxy score MF simple : momentum 1 an + inverse volatilité
        # (scores factoriels complets non disponibles pour dates passées)
        lookback = min(252, len(sub))
        sub_r = sub.tail(lookback).pct_change().dropna(how="all")
        if sub_r.empty:
            return None

        mom   = sub_r.mean()
        vol   = sub_r.std().replace(0, np.nan)
        score = mom / vol  # ratio Sharpe simplifié

        # Appliquer les betas si facteurs disponibles
        if factor_results:
            mf_score = pd.Series(0.0, index=score.index)
            for fname, df_f in factor_results.items():
                if df_f is None:
                    continue
                sc_cols = [c for c in df_f.columns if "Score" in c]
                if not sc_cols:
                    continue
                sc = df_f[sc_cols[0]]
                beta_i = betas.get(fname, 0.2)
                for t in mf_score.index:
                    if t in sc.index:
                        mf_score[t] += sc[t] * beta_i
            score = mf_score

        if eligible:
            score = score.reindex([t for t in eligible if t in score.index]).dropna()

        score = score.dropna()
        if score.empty:
            return None

        n = len(score)
        ranks = score.rank(ascending=False, method='min')
        w = (n - ranks + 1) / (n * (n+1) / 2)
        return w / w.sum()

    # Charia tickers
    charia_tickers = None
    if charia_results:
        from charia_screening import get_charia_compatible_tickers
        charia_tickers = get_charia_compatible_tickers(charia_results)

    # ── Simulation jour par jour ───────────────────────────────
    pf_values   = []   # portefeuille MF
    ch_values   = []   # portefeuille Charia
    bm_values   = []   # BRVM Composite (éqiupondéré)

    w_pf = None
    w_ch = None
    next_rebal_idx = 0
    val_pf = val_ch = val_bm = 100.0

    for date in cours.index:
        # Rebalancement
        if (next_rebal_idx < len(rebal_dates) and
                date >= rebal_dates[next_rebal_idx]):
            w_pf = compute_weights_at(date)
            if charia_tickers:
                w_ch = compute_weights_at(date, eligible=charia_tickers)
            next_rebal_idx += 1

        # Rendements du jour — extraction scalaire garantie
        day_ret_raw = ret.loc[date]
        if isinstance(day_ret_raw, pd.DataFrame):
            day_ret = day_ret_raw.iloc[0].apply(pd.to_numeric, errors="coerce").fillna(0)
        else:
            day_ret = pd.to_numeric(day_ret_raw, errors="coerce").fillna(0)

        def port_ret(weights):
            if weights is None:
                return float(day_ret.mean())
            common = weights.index.intersection(day_ret.index)
            if common.empty:
                return 0.0
            w = pd.to_numeric(weights[common], errors="coerce").fillna(0)
            r = pd.to_numeric(day_ret[common],  errors="coerce").fillna(0)
            return float((w * r).sum())

        r_pf = port_ret(w_pf)
        r_ch = port_ret(w_ch) if (w_ch is not None and charia_tickers) else r_pf
        r_bm = float(day_ret.mean())

        val_pf *= (1 + r_pf)
        val_ch *= (1 + r_ch)
        val_bm *= (1 + r_bm)

        pf_values.append(val_pf)
        ch_values.append(val_ch)
        bm_values.append(val_bm)

    # ── Séries de performance ──────────────────────────────────
    idx = cours.index
    perf = pd.DataFrame({
        "Portefeuille MF":      pd.array(pf_values, dtype=float),
        "Portefeuille Charia":  pd.array(ch_values, dtype=float),
        "BRVM Composite":       pd.array(bm_values, dtype=float),
    }, index=idx)

    # ── Métriques ─────────────────────────────────────────────
    rf_daily = 0.06 / 252

    def metrics(series):
        s = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
        if len(s) < 2 or s.iloc[0] == 0:
            return {}
        rets = s.pct_change().replace([np.inf, -np.inf], np.nan).dropna()
        rets = rets.clip(-0.20, 0.30)
        if len(rets) < 5:
            return {}
        n_years = len(rets) / 252
        if n_years <= 0 or s.iloc[0] <= 0 or s.iloc[-1] <= 0:
            return {}
        cagr    = (s.iloc[-1] / s.iloc[0]) ** (1/n_years) - 1
        vol     = float(rets.std()) * np.sqrt(252)
        std_    = float(rets.std())
        sharpe  = float((rets.mean() - rf_daily) / std_ * np.sqrt(252)) \
                  if std_ > 0 else 0.0
        roll_max = s.cummax().replace(0, np.nan)
        dd       = ((s - roll_max) / roll_max).replace([np.inf,-np.inf], np.nan)
        max_dd   = float(dd.min()) if not dd.isna().all() else 0.0
        total_ret = float(s.iloc[-1] / s.iloc[0]) - 1

        # Sanity check
        if not np.isfinite(cagr) or not np.isfinite(vol):
            return {}

        return {
            "Rendement total":     f"{total_ret:+.2%}",
            "CAGR":                f"{cagr:+.2%}",
            "Volatilité annuelle": f"{vol:.2%}",
            "Ratio de Sharpe":     f"{sharpe:.3f}",
            "Max Drawdown":        f"{max_dd:.2%}",
            "_cagr":  cagr, "_vol":   vol,
            "_sharpe":sharpe, "_maxdd":max_dd, "_total":total_ret,
        }

    mets = {k: metrics(perf[k]) for k in perf.columns}

    # Alpha / Beta vs BRVM Composite
    def alpha_beta(port_series, bm_series):
        pr = pd.to_numeric(port_series, errors="coerce").pct_change().replace(
            [np.inf,-np.inf], np.nan).dropna().clip(-0.20, 0.30)
        br = pd.to_numeric(bm_series,   errors="coerce").pct_change().replace(
            [np.inf,-np.inf], np.nan).dropna().clip(-0.20, 0.30)
        common = pr.index.intersection(br.index)
        if len(common) < 20:
            return None, None
        pr, br = pr.loc[common], br.loc[common]
        bvar = float(br.var())
        if bvar <= 0 or not np.isfinite(bvar):
            return None, None
        beta  = float(pr.cov(br) / bvar)
        alpha = float((pr.mean() - beta * br.mean()) * 252)
        if not np.isfinite(alpha) or not np.isfinite(beta):
            return None, None
        return alpha, beta

    for k in ["Portefeuille MF", "Portefeuille Charia"]:
        a, b = alpha_beta(perf[k], perf["BRVM Composite"])
        if a is not None:
            mets[k]["Alpha annualisé"] = f"{a:+.2%}"
            mets[k]["Beta"]            = f"{b:.3f}"
            mets[k]["_alpha"]          = a
            mets[k]["_beta"]           = b

    return {"perf": perf, "metrics": mets, "rebal_dates": rebal_dates}


def compute_correlation_matrix(cours_df, window=252, min_coverage=0.5):
    """
    Matrice de corrélations des rendements sur une fenêtre glissante.
    Exclut les indices BRVM et les titres avec trop de données manquantes.
    """
    BENCH = {'Unnamed: 0','ANNEE','JOUR','Unnamed: 62','.BRVMCI','BRVM30',
             'BRVM PREST','BRVM-PRINC','BRVM-C TR','BRVM-CB','BRVM-CD',
             'BRVM-ENER','BRVM-SFIN','BRVM-SPUB','Date'}
    cours = cours_df.apply(pd.to_numeric, errors="coerce")
    cols  = [c for c in cours.columns if c not in BENCH]
    sub   = cours[cols].tail(window).ffill()
    # Filtre couverture
    cov   = sub.notna().mean()
    sub   = sub[cov[cov >= min_coverage].index]
    ret   = sub.pct_change().clip(-0.20, 0.30).dropna(how="all")
    return ret.corr()


def cluster_tickers(corr_matrix, max_corr=0.70):
    """
    Clustering hiérarchique Ward sur la matrice de corrélations.
    Retourne un dict {ticker: cluster_id} et la liste ordonnée des tickers.
    """
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance import squareform

    # Nettoyer la matrice : remplacer NaN par 0, forcer symétrie exacte
    corr = corr_matrix.copy()
    corr = corr.fillna(0)
    corr_vals = corr.values.astype(float)

    # Forcer symétrie parfaite : moyenne des deux triangles
    corr_sym = (corr_vals + corr_vals.T) / 2
    np.fill_diagonal(corr_sym, 1.0)

    # Distance = 1 - corrélation, bornée [0, 2]
    dist = np.clip(1 - corr_sym, 0, 2)
    np.fill_diagonal(dist, 0.0)

    # Forcer symétrie exacte à la précision flottante
    dist = (dist + dist.T) / 2
    np.fill_diagonal(dist, 0.0)

    tickers = corr_matrix.columns.tolist()

    try:
        condensed = squareform(dist, checks=False)
        Z         = linkage(condensed, method="ward")
        labels    = fcluster(Z, t=(1 - max_corr), criterion="distance")
    except Exception:
        # Fallback : un seul cluster
        labels = np.ones(len(tickers), dtype=int)
        Z      = None

    order = np.argsort(labels)
    return {tickers[i]: int(labels[i]) for i in range(len(tickers))}, \
           [tickers[i] for i in order], Z


def sector_allocation(weights, sector_map, all_tickers):
    """
    Calcule la répartition sectorielle d'un portefeuille.
    weights : Series {ticker: poids}
    Retourne Series {secteur: poids_total}
    """
    alloc = {}
    for t, w in weights.items():
        sec = sector_map.get(t.upper(), "Autre")
        alloc[sec] = alloc.get(sec, 0) + float(w)
    return pd.Series(alloc).sort_values(ascending=False)


def benchmark_sector_weights(sector_map, cours_df, date=None):
    """
    Poids sectoriels du benchmark (BRVM Composite équipondéré).
    Retourne Series {secteur: poids}
    """
    BENCH = {'Unnamed: 0','ANNEE','JOUR','Unnamed: 62','.BRVMCI','BRVM30',
             'BRVM PREST','BRVM-PRINC','BRVM-C TR','BRVM-CB','BRVM-CD',
             'BRVM-ENER','BRVM-SFIN','BRVM-SPUB','Date'}
    cols = [c for c in cours_df.columns if c not in BENCH]
    # Filtrer les titres avec cours actif
    if date:
        sub = cours_df.loc[cours_df.index <= pd.to_datetime(date), cols]
    else:
        sub = cours_df[cols]
    active = sub.tail(5).notna().any()
    active_tickers = active[active].index.tolist()
    n = len(active_tickers)
    if n == 0:
        return pd.Series()
    w_each = 1.0 / n
    alloc  = {}
    for t in active_tickers:
        sec = sector_map.get(t.upper(), "Autre")
        alloc[sec] = alloc.get(sec, 0) + w_each
    return pd.Series(alloc).sort_values(ascending=False)


def sector_performance(cours_df, sector_map, start_date, end_date):
    """
    Performance de chaque secteur sur la période (rendement équipondéré).
    """
    BENCH = {'Unnamed: 0','ANNEE','JOUR','Unnamed: 62','.BRVMCI','BRVM30',
             'BRVM PREST','BRVM-PRINC','BRVM-C TR','BRVM-CB','BRVM-CD',
             'BRVM-ENER','BRVM-SFIN','BRVM-SPUB','Date'}
    cours = cours_df.apply(pd.to_numeric, errors="coerce")
    cols  = [c for c in cours.columns if c not in BENCH]
    sub   = cours[cols]
    sub   = sub[(sub.index >= pd.to_datetime(start_date)) &
                (sub.index <= pd.to_datetime(end_date))]
    sub   = sub.ffill()

    sect_ret = {}
    for sec in set(sector_map.values()):
        tickers = [t for t, s in sector_map.items() if s == sec
                   and t in sub.columns]
        if not tickers:
            continue
        prices  = sub[tickers].dropna(how="all")
        if len(prices) < 2:
            continue
        rets    = prices.pct_change().clip(-0.20, 0.30).mean(axis=1)
        cumret  = (1 + rets).cumprod()
        sect_ret[sec] = cumret

    return pd.DataFrame(sect_ret) if sect_ret else pd.DataFrame()


# ─── TABS ─────────────────────────────────────────────────────
t1, t2, t3, t4, t5, t6, t7, t8, t9, t10, t11 = st.tabs([
    "💰 Value", "🚀 Momentum", "📉 Volatilité", "💸 Dividende",
    "💧 Liquidité", "🔢 Indice MF", "📂 Portefeuille",
    "📈 Valorisation", "📊 Backtesting",
    "🔗 Corrélations & Diversification", "ℹ️ Données"])

# ══ VALUE ══════════════════════════════════════════════════════
with t1:
    st.markdown("<span class='pill'>Étape 1 · Facteur Value</span><p class='sh'>Valorisation relative au cours</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    F_value(t,T) = Σ w_j · m_j(t,T) / max_{E_t}(m_j(t,T))<br><br>
    B/P  = (Capitaux propres / Nb titres) / Cours moyen &nbsp;→&nbsp; score ↑ = sous-évalué ✓<br>
    E/P  = (Résultat net / Nb titres) / Cours moyen &nbsp;&nbsp;&nbsp;&nbsp;→&nbsp; inverse de P/E  ✓<br>
    FCF/P = (FCF / Nb titres) / Cours moyen<br>
    CA/P  = (CA / Nb titres) / Cours moyen
    </div>""", unsafe_allow_html=True)

    # ── Paramètres de période ──────────────────────────────────
    st.markdown("**📅 Période de calcul du cours moyen**")
    vd1, vd2, vd3 = st.columns([1,1,1])
    with vd1:
        cours_min = data["cours"].index.min().date()
        cours_max = data["cours"].index.max().date()
        v_start_default = max(cours_min, min(cours_max,
                          pd.Timestamp(f"{selected_year}-01-01").date()))
        v_date_start = st.date_input("Date début", value=v_start_default,
                                      min_value=cours_min, max_value=cours_max, key="v_d1")
    with vd2:
        v_end_default = max(cours_min, min(cours_max,
                        pd.Timestamp(f"{selected_year}-12-31").date()))
        v_date_end   = st.date_input("Date fin", value=v_end_default,
                                      min_value=cours_min, max_value=cours_max, key="v_d2")
    with vd3:
        st.markdown(f"<br><span style='font-size:11px;color:#64748b;'>"
                    f"Fondamentaux année : <b>{selected_year}</b></span>", unsafe_allow_html=True)

    # ── Pondérations ───────────────────────────────────────────
    st.markdown("**⚖️ Pondérations des métriques**")
    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
    w_bp   = mc1.number_input("B/P  (CP/n÷cours)",    0.0, 1.0, 0.20, 0.05, key="wbp")
    w_eps  = mc2.number_input("E/P  (RN/n÷cours)",    0.0, 1.0, 0.20, 0.05, key="wep")
    w_fcf  = mc3.number_input("FCF/P (FCF/n÷cours)",  0.0, 1.0, 0.20, 0.05, key="wfcf")
    w_ca   = mc4.number_input("CA/P  (CA/n÷cours)",   0.0, 1.0, 0.20, 0.05, key="wca")
    w_ebit = mc5.number_input("EBIT/P (Rex/n÷cours)", 0.0, 1.0, 0.20, 0.05, key="webit")

    ws = round(w_bp + w_eps + w_fcf + w_ca + w_ebit, 4)
    if abs(ws - 1.0) > 0.01:
        st.warning(f"⚠️ Somme des poids = {ws:.2f} ≠ 1.0")
    else:
        st.success(f"✅ Somme des poids = {ws:.2f}")

    if st.button("⚙️ Calculer le Facteur Value", type="primary"):
        metrics_v = ["B/P  (CP/n÷cours)", "E/P  (RN/n÷cours)",
                     "FCF/P (FCF/n÷cours)", "CA/P  (CA/n÷cours)",
                     "EBIT/P (Rex/n÷cours)"]
        weights_v = dict(zip(metrics_v, [w_bp, w_eps, w_fcf, w_ca, w_ebit]))
        res = compute_value_factor(data, selected_year, weights_v,
                                   date_start=v_date_start, date_end=v_date_end)
        if res is not None:
            fr["Value"] = res
            st.success(f"✅ {len(res)} titres scorés · cours moyen {v_date_start} → {v_date_end} · fondamentaux {selected_year}")
        else:
            st.error("Données insuffisantes pour la période et l'année sélectionnées.")

    if "Value" in fr:
        res = fr["Value"]
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Titres scorés",  len(res))
        m2.metric("Score max",      f"{res['Score Value'].max():.4f}")
        m3.metric("N°1 Value",      res.index[0])
        m4.metric("Score moyen",    f"{res['Score Value'].mean():.4f}")

        st.plotly_chart(score_bar(res["Score Value"], "#06b6d4"), width="stretch")

        metrics_v = ["B/P  (CP/n÷cours)", "E/P  (RN/n÷cours)",
                     "FCF/P (FCF/n÷cours)", "CA/P  (CA/n÷cours)",
                     "EBIT/P (Rex/n÷cours)"]
        all_possible = ["Score Value"] + metrics_v + ["Cours moyen", "Période"]
        show_cols = [c for c in all_possible if c in res.columns]
        disp = res[show_cols].reset_index().rename(columns={"index": "Ticker"})
        disp.insert(0, "Rang", range(1, len(disp) + 1))
        fmt = {"Score Value": "{:.4f}", "Cours moyen": "{:,.1f}"}
        fmt.update({m: "{:.6f}" for m in metrics_v})
        st.dataframe(disp.style.format(fmt).bar(subset=["Score Value"],
                     color=["#1e3a5f", "#3b82f6"]),
                     width="stretch", hide_index=True)

        buf = io.BytesIO()
        disp.to_excel(buf, index=False)
        st.download_button("⬇️ Exporter Value (Excel)", buf.getvalue(),
                           f"value_{selected_year}_{v_date_start}_{v_date_end}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ MOMENTUM ═══════════════════════════════════════════════════
with t2:
    st.markdown("<span class='pill'>Étape 1 · Facteur Momentum</span><p class='sh'>Dynamique des cours — 6 horizons</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    F_mom(t,T) = Σ w_h · rdt_moyen_h(t,T) / max_{E_t}(rdt_moyen_h)<br>
    Une seule plage globale · les 6 horizons calculent leur rendement moyen sur cette fenêtre
    </div>""", unsafe_allow_html=True)

    c_min = data["cours"].index.min().date()
    c_max = data["cours"].index.max().date()

    # ── Plage globale unique ──────────────────────────────────
    st.markdown("**📅 Plage de dates globale**")
    st.caption("Tous les horizons (Journalier, Hebdo, Mensuel...) calculent leur rendement moyen sur cette fenêtre.")
    mg1, mg2, mg3 = st.columns([1, 1, 1])
    with mg1:
        mom_global_start = st.date_input(
            "Date début", key="mom_global_start",
            value=max(c_min, min(c_max, (pd.Timestamp(c_max) - pd.Timedelta(days=365)).date())),
            min_value=c_min, max_value=c_max
        )
    with mg2:
        mom_global_end = st.date_input(
            "Date fin", key="mom_global_end",
            value=c_max, min_value=c_min, max_value=c_max
        )
    with mg3:
        if mom_global_start < mom_global_end:
            n_cal = (mom_global_end - mom_global_start).days
            # Nombre de jours de trading sur la période
            cours_tmp = data["cours"]
            n_trd = ((cours_tmp.index >= pd.to_datetime(mom_global_start)) &
                     (cours_tmp.index <= pd.to_datetime(mom_global_end))).sum()
            st.markdown("<br>", unsafe_allow_html=True)
            st.success(f"✅ **{n_cal}** jours cal. · **{n_trd}** jours trading")
        else:
            st.markdown("<br>", unsafe_allow_html=True)
            st.error("⚠️ Date début ≥ Date fin")

    # ── Pondérations par horizon ──────────────────────────────
    st.markdown("**⚖️ Poids par horizon**")
    HORIZON_LABELS = ["Journalier","Hebdo","Mensuel","Trimestriel","Semestriel","Annuel"]
    wh_cols = st.columns(6)
    horizon_weights = {}
    for i, h in enumerate(HORIZON_LABELS):
        w = wh_cols[i].number_input(h, 0.0, 1.0, round(1/6, 4), 0.01, key=f"mom_w_{h}")
        horizon_weights[h] = w

    w_mom_sum = round(sum(horizon_weights.values()), 4)
    if abs(w_mom_sum - 1.0) > 0.01:
        st.warning(f"⚠️ Σ poids = {w_mom_sum:.2f} ≠ 1.0 — renormalisés automatiquement")
    else:
        st.success(f"✅ Σ poids = {w_mom_sum:.2f}")

    if st.button("⚙️ Calculer le Momentum", type="primary"):
        if mom_global_start >= mom_global_end:
            st.error("Plage de dates invalide.")
        else:
            # La même plage pour tous les horizons
            horizon_ranges = {h: (mom_global_start, mom_global_end) for h in HORIZON_LABELS}
            res = compute_momentum_factor(data, horizon_ranges, horizon_weights)
            if res is not None:
                fr["Momentum"] = res
                st.success(f"✅ {len(res)} titres scorés · plage {mom_global_start} → {mom_global_end}")
            else:
                st.error("Données insuffisantes pour la période sélectionnée.")

    if "Momentum" in fr:
        res = fr["Momentum"]
        m1, m2, m3 = st.columns(3)
        m1.metric("Titres",    len(res))
        m2.metric("N°1",       res.index[0])
        m3.metric("Score max", f"{res['Score Momentum'].max():.4f}")

        st.plotly_chart(score_bar(res["Score Momentum"], "#8b5cf6"), width="stretch")

        rdt_c = [c for c in res.columns if "Rdt" in c]
        disp2 = res[["Score Momentum"] + rdt_c].head(20).reset_index().rename(columns={"index": "Ticker"})
        disp2.insert(0, "Rang", range(1, len(disp2)+1))
        st.dataframe(disp2.style.format(
            {"Score Momentum": "{:.4f}"} | {c: "{:.4%}" for c in rdt_c}
        ), width="stretch", hide_index=True)

        buf = io.BytesIO()
        disp2.to_excel(buf, index=False)
        st.download_button("⬇️ Exporter Momentum (Excel)", buf.getvalue(),
                           "momentum.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ VOLATILITÉ ═════════════════════════════════════════════════
with t3:
    st.markdown("<span class='pill'>Étape 1 · Facteur Volatilité</span><p class='sh'>Faible volatilité — Inversion</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>F_vol(t,T) = min_{E}(σ) / σ(T) &nbsp;·&nbsp; Score élevé = titre stable</div>""", unsafe_allow_html=True)
    window_v = st.slider("Fenêtre (jours de trading)", 60, 504, 252, 21)

    if st.button("⚙️ Calculer la Volatilité", type="primary"):
        res = compute_volatility_factor(data, str(date_ref), window_v)
        fr["Volatilité"] = res
        st.success(f"✅ {len(res)} titres scorés")

    if "Volatilité" in fr:
        res = fr["Volatilité"]
        m1,m2,m3 = st.columns(3)
        m1.metric("Titre le + stable", res.index[0])
        m2.metric("σ min", f"{res['Écart-type σ'].min():.4%}")
        m3.metric("σ moy", f"{res['Écart-type σ'].mean():.4%}")

        l,r = st.columns(2)
        with l:
            st.markdown("**Score (inversé) — Top 25**")
            st.plotly_chart(score_bar(res["Score Volatilité"], "#ef4444"), width="stretch")
        with r:
            st.markdown("**Volatilité annualisée (σ×√252)**")
            rv = res.sort_values("Écart-type σ")
            fig2 = go.Figure(go.Bar(x=rv.index, y=rv["σ annualisée"],
                                    marker=dict(color=rv["σ annualisée"],
                                                colorscale=[[0,"#10b981"],[0.5,"#f59e0b"],[1,"#ef4444"]])))
            fig2.update_layout(**{**PLOT_LAYOUT,"height":360,"yaxis":dict(gridcolor="#1e2d45",tickformat=".1%")})
            st.plotly_chart(fig2, width="stretch")

# ══ DIVIDENDE ══════════════════════════════════════════════════
with t4:
    st.markdown("<span class='pill'>Étape 1 · Facteur Dividende</span><p class='sh'>Dividend Yield = Dividende annuel / Cours moyen</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    F_div(t,T) = 1.0 × DY(t,T) / max_{E_t}(DY(t,T))<br>
    Métrique unique (w = 100%) : DY = Dividende annuel versé / Cours moyen sur la période
    </div>""", unsafe_allow_html=True)

    div_years = sorted(data["dividendes"]["Date"].dropna().astype(int).unique(), reverse=True)
    dd1, dd2, dd3 = st.columns([1, 1, 1])
    with dd1:
        div_yr = st.selectbox("Année du dividende", div_years, key="div_yr")
    with dd2:
        d_cours_min = data["cours"].index.min().date()
        d_cours_max = data["cours"].index.max().date()
        d_start_default = max(d_cours_min, min(d_cours_max,
                          pd.Timestamp(f"{div_years[0]}-01-01").date()))
        d_date_start = st.date_input("Cours moyen — Date début",
                                      value=d_start_default,
                                      min_value=d_cours_min,
                                      max_value=d_cours_max, key="dd1")
    with dd3:
        d_end_default = max(d_cours_min, min(d_cours_max,
                        pd.Timestamp(f"{div_years[0]}-12-31").date()))
        d_date_end = st.date_input("Cours moyen — Date fin",
                                    value=d_end_default,
                                    min_value=d_cours_min,
                                    max_value=d_cours_max, key="dd2")

    if st.button("⚙️ Calculer le Dividende", type="primary"):
        res = compute_dividend_factor(data, div_yr,
                                       date_start=d_date_start, date_end=d_date_end)
        if res is not None:
            fr["Dividende"] = res
            paying = len(res[res["Dividend Yield"] > 0])
            st.success(f"✅ {len(res)} titres · {paying} payeurs · cours moyen {d_date_start} → {d_date_end}")
        else:
            st.error("Données insuffisantes pour la période sélectionnée.")

    if "Dividende" in fr:
        res = fr["Dividende"]
        paying = res[res["Dividend Yield"] > 0]
        m1, m2, m3 = st.columns(3)
        m1.metric("Payeurs de dividende", len(paying))
        m2.metric("Yield max", f"{res['Dividend Yield'].max():.2%}")
        m3.metric("N°1 Dividende", res.index[0])

        fig_d = go.Figure(go.Bar(
            x=res.index, y=res["Dividend Yield"],
            marker=dict(color=res["Dividend Yield"],
                        colorscale=[[0,"#1e2d45"],[1,"#f59e0b"]]),
            text=[f"{v:.2%}" for v in res["Dividend Yield"]],
            textposition="outside", textfont=dict(size=9, color="#94a3b8")))
        fig_d.update_layout(**{**PLOT_LAYOUT, "height": 360,
                               "yaxis": dict(gridcolor="#1e2d45", tickformat=".1%")})
        st.plotly_chart(fig_d, width="stretch")

        disp_d = res.reset_index().rename(columns={"index": "Ticker"})
        disp_d.insert(0, "Rang", range(1, len(disp_d)+1))
        st.dataframe(disp_d.style.format({
            "Score Dividende": "{:.4f}",
            "Dividend Yield":  "{:.4%}",
            "Cours moyen":     "{:,.1f}",
        }), width="stretch", hide_index=True)

        buf = io.BytesIO()
        disp_d.to_excel(buf, index=False)
        st.download_button("⬇️ Exporter Dividende (Excel)", buf.getvalue(),
                           f"dividende_{div_yr}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ LIQUIDITÉ ══════════════════════════════════════════════════
with t5:
    st.markdown("<span class='pill'>Étape 1 · Facteur Liquidité</span><p class='sh'>Volume moyen transigé sur la période</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    F_liq(t,T) = 1.0 × Vol_moy(T) / max_{E_t}(Vol_moy)<br>
    Métrique unique (w = 100%) : volume moyen transigé calculé sur la plage de dates choisie
    </div>""", unsafe_allow_html=True)

    ll1, ll2 = st.columns(2)
    with ll1:
        l_date_start = st.date_input("Date début",
                                      value=data["volumes"].index.min().date(),
                                      min_value=data["volumes"].index.min().date(),
                                      max_value=data["volumes"].index.max().date(), key="ld1")
    with ll2:
        l_date_end   = st.date_input("Date fin",
                                      value=data["volumes"].index.max().date(),
                                      min_value=data["volumes"].index.min().date(),
                                      max_value=data["volumes"].index.max().date(), key="ld2")

    if st.button("⚙️ Calculer la Liquidité", type="primary"):
        res = compute_liquidity_factor(data, date_start=l_date_start, date_end=l_date_end)
        if res is not None:
            fr["Liquidité"] = res
            st.success(f"✅ {len(res)} titres scorés · période {l_date_start} → {l_date_end}")
        else:
            st.error("Données insuffisantes pour la période sélectionnée.")

    if "Liquidité" in fr:
        res = fr["Liquidité"]
        m1, m2, m3 = st.columns(3)
        m1.metric("Titre le + liquide", res.index[0])
        m2.metric("Vol. moyen max",  f"{res['Volume moyen'].max()/1e6:.1f} M FCFA")
        m3.metric("Score max",       f"{res['Score Liquidité'].max():.4f}")

        fig_l = go.Figure(go.Bar(
            x=res.index, y=res["Volume moyen"] / 1e6,
            marker=dict(color=res["Score Liquidité"],
                        colorscale=[[0,"#1e2d45"],[1,"#06b6d4"]]),
            text=(res["Volume moyen"]/1e6).round(1),
            textposition="outside", textfont=dict(size=9, color="#94a3b8"),
        ))
        fig_l.update_layout(**{**PLOT_LAYOUT, "height": 360,
                               "yaxis": dict(gridcolor="#1e2d45", title="Volume moy. (M FCFA)")})
        st.plotly_chart(fig_l, width="stretch")

        disp_l = res[["Score Liquidité","Volume moyen"]].reset_index().rename(columns={"index":"Ticker"})
        disp_l.insert(0, "Rang", range(1, len(disp_l)+1))
        st.dataframe(disp_l.style.format({
            "Score Liquidité": "{:.4f}",
            "Volume moyen":    "{:,.0f}",
        }), width="stretch", hide_index=True)

        buf = io.BytesIO()
        disp_l.to_excel(buf, index=False)
        st.download_button("⬇️ Exporter Liquidité (Excel)", buf.getvalue(),
                           f"liquidite_{l_date_start}_{l_date_end}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ INDICE MULTIFACTORIEL ══════════════════════════════════════
with t6:
    st.markdown("<span class='pill'>Étape 2</span><p class='sh'>Indice Multifactoriel</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>MF(t,T) = Σ_{i=1}^{5} β_i · F_i(t,T) &nbsp;·&nbsp; Σ β_i = 1<br>
    Calibrez les β manuellement via la sidebar, ou laissez le ML les optimiser automatiquement</div>""",
    unsafe_allow_html=True)

    betas_mf = st.session_state.get("betas",
               {"Value":0.20,"Momentum":0.20,"Volatilité":0.20,"Dividende":0.20,"Liquidité":0.20})
    computed = list(fr.keys())
    CLRS  = {"Value":"#3b82f6","Momentum":"#8b5cf6","Volatilité":"#ef4444",
             "Dividende":"#f59e0b","Liquidité":"#06b6d4"}
    ICONS = {"Value":"💰","Momentum":"🚀","Volatilité":"📉","Dividende":"💸","Liquidité":"💧"}

    if not computed:
        st.info("👈 Calculez au moins un facteur (onglets 💰 🚀 📉 💸 💧) avant de continuer.")
    else:
        st.success(f"✅ Facteurs calculés : {', '.join(computed)}")

        # ── Choix du mode de calibration des β ────────────────────
        st.markdown("---")
        st.markdown("**⚖️ Mode de calibration des β_i**")
        beta_mode = st.radio(
            "Mode β",
            ["🎛️ Manuel", "🤖 Automatique"],
            horizontal=True,
            label_visibility="collapsed",
        )

        # ══════════════════════════════════════════════════════════
        # MODE A — MANUEL
        # ══════════════════════════════════════════════════════════
        if beta_mode == "🎛️ Manuel":
            st.caption("Les β sont ceux définis dans la barre latérale ←")
            bc = st.columns(5)
            for i, fname in enumerate(["Value","Momentum","Volatilité","Dividende","Liquidité"]):
                with bc[i]:
                    st.markdown(
                        f"<div style='background:#161d2e;border:1px solid {CLRS[fname]};"
                        f"border-radius:8px;padding:10px;text-align:center;'>"
                        f"<div style='font-size:16px'>{ICONS[fname]}</div>"
                        f"<div style='font-size:10px;color:#94a3b8;'>{fname}</div>"
                        f"<div style='font-size:20px;font-weight:800;color:{CLRS[fname]};'>"
                        f"{betas_mf.get(fname,0):.2f}</div>"
                        f"<div style='font-size:10px;color:#64748b;'>"
                        f"{'✓' if fname in computed else '✗'}</div></div>",
                        unsafe_allow_html=True
                    )
            bs_mf = round(sum(betas_mf.values()), 4)
            if abs(bs_mf - 1.0) > 0.01:
                st.warning(f"⚠️ Σβ = {bs_mf:.2f} — ajustez les curseurs dans la sidebar")

        # ══════════════════════════════════════════════════════════
        # MODE B — AUTOMATIQUE (3 méthodes distinctes)
        # ══════════════════════════════════════════════════════════
        else:
            if not ML_AVAILABLE:
                st.error("❌ scikit-learn non installé. Ajoutez `scikit-learn>=1.3.0` dans requirements.txt.")
            else:
                # ── Sélection de la méthode ────────────────────────────
                st.markdown("**🔬 Choisissez une méthode d'optimisation**")
                methode = st.radio(
                    "Méthode",
                    [
                        "① OLS — Régression sur rendements historiques",
                        "② Walk-forward — Sharpe glissant",
                        "③ ML — Random Forest + Gradient Boosting",
                    ],
                    label_visibility="collapsed",
                )

                METHODE_INFO = {
                    "① OLS — Régression sur rendements historiques": (
                        "Les β_i sont les coefficients d'une régression linéaire (moindres carrés) "
                        "entre les scores factoriels F_i(T) et les rendements réalisés. "
                        "β élevé = facteur qui explique le mieux les rendements historiques."
                    ),
                    "② Walk-forward — Sharpe glissant": (
                        "La période est découpée en fenêtres glissantes. Dans chaque fenêtre, "
                        "la corrélation score↔rendement mesure la prédictivité de chaque facteur. "
                        "β_i = moyenne des corrélations normalisées sur toutes les fenêtres."
                    ),
                    "③ ML — Random Forest + Gradient Boosting": (
                        "Deux modèles ML (Random Forest et Gradient Boosting) apprennent quels "
                        "facteurs ont le mieux prédit les rendements. "
                        "β_i = moyenne des importances RF et GB, normalisée."
                    ),
                }
                st.info(METHODE_INFO[methode])

                # ── Paramètres communs ─────────────────────────────────
                st.markdown("---")
                st.markdown("**📅 Fenêtres de calcul**")
                ml_c1, ml_c2 = st.columns(2)
                c_min = data["cours"].index.min().date()
                c_max = data["cours"].index.max().date()

                with ml_c1:
                    st.markdown("**📐 Features X — Scores factoriels**")
                    st.caption("Scores F_i(T) recalculés sur cette plage")
                    ml_ts = st.date_input("Début features", key="ml_ts",
                        value=max(c_min, min(c_max, pd.Timestamp("2019-01-01").date())),
                        min_value=c_min, max_value=max(c_max, datetime.date.today()))
                    ml_te = st.date_input("Fin features", key="ml_te",
                        value=max(c_min, min(c_max, pd.Timestamp("2023-12-31").date())),
                        min_value=c_min, max_value=max(c_max, datetime.date.today()))
                    fund_years = data.get("fundamental_years", [2024])
                    ml_year = st.selectbox("Année fondamentaux", fund_years, key="ml_year")

                with ml_c2:
                    st.markdown("**🎯 Cible Y — Rendements réalisés**")
                    st.caption("Rendement total (P_fin - P_deb) / P_deb")
                    ml_tgs = st.date_input("Début cible", key="ml_tgs",
                        value=max(c_min, min(c_max, pd.Timestamp("2024-01-01").date())),
                        min_value=c_min, max_value=max(c_max, datetime.date.today()))
                    ml_tge = st.date_input("Fin cible", key="ml_tge",
                        value=c_max, min_value=c_min, max_value=max(c_max, datetime.date.today()))
                    st.markdown("<br>", unsafe_allow_html=True)
                    if ml_tgs < ml_tge:
                        gap = (ml_tgs - ml_te).days if ml_te < ml_tgs else 0
                        if gap > 0:
                            st.success(f"✅ Écart features→cible : {gap} jours")
                        else:
                            st.warning("⚠️ Fenêtres qui se chevauchent")

                # Paramètre spécifique ML
                n_trees = 200
                if "③ ML" in methode:
                    n_trees = st.slider("Nombre d'arbres (RF + GB)", 50, 500, 200, 50,
                                        key="ml_ntrees")

                apply_auto = st.toggle(
                    "Appliquer β à la sidebar automatiquement",
                    value=True, key="ml_auto"
                )

                # ── Bouton de lancement ────────────────────────────────
                lbl_btn = {
                    "① OLS":  "📐 Lancer la régression OLS",
                    "② Walk": "📈 Lancer le Walk-forward",
                    "③ ML":   "🤖 Lancer le ML (RF + GB)",
                }
                btn_label = next(v for k,v in lbl_btn.items() if k[:4] in methode)

                if st.button(btn_label, type="primary"):
                    if ml_ts >= ml_te:
                        st.error("Fenêtre features invalide (début ≥ fin).")
                    elif ml_tgs >= ml_tge:
                        st.error("Fenêtre cible invalide (début ≥ fin).")
                    else:
                        betas_opt = None
                        extra_info = None

                        if "① OLS" in methode:
                            with st.spinner("Régression OLS en cours..."):
                                betas_opt, extra_info = optimize_betas_ols(
                                    data, ml_ts, ml_te, ml_tgs, ml_tge, ml_year)
                            if betas_opt:
                                r2 = extra_info.get("r2", 0) if extra_info else 0
                                st.success(f"✅ OLS terminé · R² = {r2:.4f} · "
                                           f"{extra_info.get('n_obs',0)} observations")

                        elif "② Walk" in methode:
                            with st.spinner("Walk-forward Sharpe en cours..."):
                                betas_opt, extra_info = optimize_betas_walkforward(
                                    data, ml_ts, ml_te, ml_tgs, ml_tge, ml_year)
                            if betas_opt:
                                nw = extra_info.get("n_windows", 0) if extra_info else 0
                                st.success(f"✅ Walk-forward terminé · {nw} fenêtres calculées")

                        elif "③ ML" in methode:
                            with st.spinner("Random Forest + Gradient Boosting en cours..."):
                                betas_opt, ml_res, ml_ds = optimize_betas_ml(
                                    data=data,
                                    train_start=ml_ts, train_end=ml_te,
                                    target_start=ml_tgs, target_end=ml_tge,
                                    year=ml_year, n_estimators=n_trees
                                )
                                extra_info = ml_res
                                st.session_state.ml_results = ml_res
                                st.session_state.ml_dataset = ml_ds
                            if betas_opt and ml_res:
                                r2_rf = ml_res["Random Forest"]["r2"]
                                r2_gb = ml_res["Gradient Boosting"]["r2"]
                                st.success(f"✅ ML terminé · R² RF={r2_rf:.4f} · "
                                           f"R² GB={r2_gb:.4f} · "
                                           f"{len(ml_ds['tickers']) if ml_ds else 0} titres")

                        if betas_opt:
                            st.session_state.ml_betas   = betas_opt
                            st.session_state.ml_methode = methode
                            if apply_auto:
                                km = {"Value":"sv_val","Momentum":"sv_mom",
                                      "Volatilité":"sv_vol","Dividende":"sv_div",
                                      "Liquidité":"sv_liq"}
                                for f, b in betas_opt.items():
                                    if f in km:
                                        st.session_state[km[f]] = round(float(b), 4)
                                st.rerun()
                        else:
                            st.error("Données insuffisantes. Vérifiez les fenêtres "
                                     "et que les facteurs sont calculés.")

                # ── Résultats de la méthode choisie ───────────────────
                if "ml_betas" in st.session_state and st.session_state.ml_betas:
                    betas_opt    = st.session_state.ml_betas
                    methode_used = st.session_state.get("ml_methode", methode)

                    st.markdown("---")
                    st.markdown(f"**β optimaux — {methode_used}**")

                    FACTORS = ["Value","Momentum","Volatilité","Dividende","Liquidité"]

                    # Cards β
                    b_cols = st.columns(5)
                    for i, fname in enumerate(FACTORS):
                        bv = betas_opt.get(fname, 0)
                        with b_cols[i]:
                            st.markdown(
                                f"<div style='background:#161d2e;border:1px solid {CLRS[fname]};"
                                f"border-radius:8px;padding:10px;text-align:center;'>"
                                f"<div style='font-size:16px'>{ICONS[fname]}</div>"
                                f"<div style='font-size:10px;color:#94a3b8;'>{fname}</div>"
                                f"<div style='font-size:22px;font-weight:800;color:{CLRS[fname]};'>"
                                f"{bv:.3f}</div>"
                                f"<div style='font-size:10px;color:#64748b;'>{bv*100:.1f}%</div>"
                                f"</div>", unsafe_allow_html=True
                            )

                    # Graphique β optimaux vs sidebar actuels
                    st.markdown("**β optimaux vs β sidebar actuels**")
                    cur = st.session_state.get("betas", {f: 0.20 for f in FACTORS})
                    x_lbl = [f"{ICONS.get(f,'')} {f}" for f in FACTORS]
                    fig_cmp = go.Figure()
                    fig_cmp.add_trace(go.Bar(
                        name="β optimaux", x=x_lbl,
                        y=[betas_opt.get(f,0) for f in FACTORS],
                        marker_color="#3b82f6", opacity=0.85
                    ))
                    fig_cmp.add_trace(go.Bar(
                        name="β sidebar", x=x_lbl,
                        y=[cur.get(f,0) for f in FACTORS],
                        marker_color="#475569", opacity=0.6
                    ))
                    fig_cmp.update_layout(
                        barmode="group",
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color="#94a3b8", family="JetBrains Mono"), height=300,
                        legend=dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)"),
                        margin=dict(l=10,r=10,t=40,b=30),
                        xaxis=dict(gridcolor="#1e2d45"),
                        yaxis=dict(gridcolor="#1e2d45", title="β_i"),
                    )
                    st.plotly_chart(fig_cmp, width="stretch")

                    # Détail ML si disponible
                    if "③ ML" in methode_used and st.session_state.get("ml_results"):
                        ml_res = st.session_state.ml_results
                        st.markdown("**Détail ML — Importance RF vs GB**")
                        factors_ord = FACTORS
                        rf_imp = ml_res["Random Forest"]["importances"]
                        gb_imp = ml_res["Gradient Boosting"]["importances"]
                        fig_ml = go.Figure()
                        fig_ml.add_trace(go.Bar(name="Random Forest", x=x_lbl,
                            y=[rf_imp.get(f,0) for f in factors_ord],
                            marker_color="#3b82f6", opacity=0.8))
                        fig_ml.add_trace(go.Bar(name="Gradient Boosting", x=x_lbl,
                            y=[gb_imp.get(f,0) for f in factors_ord],
                            marker_color="#8b5cf6", opacity=0.8))
                        fig_ml.update_layout(barmode="group",
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color="#94a3b8", family="JetBrains Mono"), height=280,
                            legend=dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)"),
                            margin=dict(l=10,r=10,t=40,b=30),
                            xaxis=dict(gridcolor="#1e2d45"),
                            yaxis=dict(gridcolor="#1e2d45"))
                        st.plotly_chart(fig_ml, width="stretch")

                    # Bouton appliquer si toggle off
                    if not apply_auto:
                        if st.button("📥 Appliquer ces β à la sidebar",
                                     key="apply_auto_btn"):
                            km = {"Value":"sv_val","Momentum":"sv_mom",
                                  "Volatilité":"sv_vol","Dividende":"sv_div",
                                  "Liquidité":"sv_liq"}
                            for f, b in betas_opt.items():
                                if f in km:
                                    st.session_state[km[f]] = round(float(b), 4)
                            st.rerun()

                    # Mise à jour betas_mf pour le calcul MF
                    betas_mf = betas_opt

                    # Mise à jour de betas_mf pour le calcul MF ci-dessous
                    betas_mf = betas_opt

        # ── Calcul MF (commun aux deux modes) ─────────────────────
        st.markdown("---")
        if st.button("🔢 Calculer l'Indice MF", type="primary"):
            mf = compute_multifactor(fr, betas_mf)
            st.session_state.mf_scores = mf
            st.success(f"✅ {len(mf)} titres classés · β = { {k: round(v,3) for k,v in betas_mf.items()} }")

        if st.session_state.mf_scores is not None:
            mf = st.session_state.mf_scores
            st.markdown("---")
            st.plotly_chart(score_bar(mf, "#3b82f6", height=380), width="stretch")

            st.markdown("**Décomposition factorielle — Top 15**")
            top15 = mf.head(15).index
            fig_s = go.Figure()
            for fname in computed:
                sc = [c for c in fr[fname].columns if "Score" in c][0]
                vals = [fr[fname].loc[t, sc] * betas_mf.get(fname, 0)
                        if t in fr[fname].index else 0 for t in top15]
                fig_s.add_trace(go.Bar(
                    name=f"{ICONS.get(fname,'')} {fname}",
                    x=list(top15), y=vals,
                    marker_color=CLRS.get(fname, "#64748b"), opacity=0.85
                ))
            fig_s.update_layout(barmode="stack",
                **{**PLOT_LAYOUT, "height":340,
                   "legend":dict(orientation="h",y=1.08,bgcolor="rgba(0,0,0,0)"),
                   "margin":dict(l=10,r=10,t=50,b=50)})
            st.plotly_chart(fig_s, width="stretch")

            tbl_mf = mf.reset_index()
            tbl_mf.columns = ["Ticker","Score MF"]
            tbl_mf.insert(0,"Rang",range(1,len(tbl_mf)+1))
            tbl_mf["Score MF"] = tbl_mf["Score MF"].round(6)
            st.dataframe(tbl_mf, width="stretch", hide_index=True)

            buf = io.BytesIO()
            tbl_mf.to_excel(buf, index=False)
            st.download_button("⬇️ Exporter classement MF", buf.getvalue(),
                "classement_MF.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ PORTEFEUILLE ═══════════════════════════════════════════
with t7:
    st.markdown("<span class='pill'>Étape 3</span><p class='sh'>Construction du Portefeuille Cible</p>", unsafe_allow_html=True)

    if st.session_state.mf_scores is None:
        st.info("👈 Calculez d'abord l'Indice MF (onglet 🔢).")
    else:
        mf = st.session_state.mf_scores
        all_tickers = mf.index.tolist()

        # ── Choix du mode ─────────────────────────────────────────
        st.markdown("**🎯 Mode de construction du portefeuille**")
        mode = st.radio(
            label="Mode",
            options=[
                "🤖 Optimisation automatique (pipeline de filtres)",
                "🖐️ Sélection manuelle des titres",
            ],
            label_visibility="collapsed",
            horizontal=True,
        )
        st.markdown("---")

        # ══════════════════════════════════════════════════════════
        # MODE A — PIPELINE AUTOMATIQUE
        # ══════════════════════════════════════════════════════════
        if mode == "🤖 Optimisation automatique (pipeline de filtres)":
            st.markdown("""<div class='fbox'>
            Pipeline séquentiel — chaque filtre réduit l'univers :<br>
            Univers complet → ① Liquidité → ② Score MF → ③ Corrélation → ④ Poids finaux
            </div>""", unsafe_allow_html=True)

            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                st.markdown("**① Filtre Liquidité**")
                liq_ok = "Liquidité" in fr
                min_vol_pct = st.slider("Volume moyen ≥ X% du max", 0, 50, 10, 1,
                    help="0% = pas de filtre", disabled=not liq_ok)
                if not liq_ok:
                    st.caption("⚠️ Calculez le facteur Liquidité pour activer")
            with col_f2:
                st.markdown("**② Filtre Score MF**")
                top_pct = st.slider("Garder le top X% des scores MF", 10, 100, 60, 5,
                    help="60% = garde les 60% de titres avec les meilleurs scores")
            with col_f3:
                st.markdown("**③ Filtre Corrélation**")
                max_corr = st.slider("Corrélation max entre titres", 0.3, 1.0, 0.75, 0.05,
                    help="Deux titres corrélés au-delà de ce seuil → on garde le meilleur")
                window_corr = st.number_input("Fenêtre (jours)", 60, 504, 252, 21, key="win_corr_a")

            st.markdown("---")
            st.markdown("**④ Pondération finale**")
            use_mkz = st.toggle("Optimisation Markowitz MF-augmentée", value=False,
                help="OFF = poids par rang MF (note technique) · ON = Mean-Variance avec score MF")
            if use_mkz:
                mk1,mk2,mk3,mk4 = st.columns(4)
                mf_weight     = mk1.slider("Poids Score MF",    0.0,1.0,0.50,0.05)
                risk_aversion = mk2.slider("Aversion risque λ", 0.1,10.0,2.0,0.1)
                min_w         = mk3.slider("Poids min/titre (%)",0,20,0,1)/100
                max_w         = mk4.slider("Poids max/titre (%)",5,100,40,5)/100
            else:
                mf_weight,risk_aversion,min_w,max_w = 0.5,2.0,0.0,1.0

            # ── Contraintes optionnelles ④ et ⑤ ──────────────────
            st.markdown("---")
            st.markdown("**⚙️ Contraintes optionnelles de diversification**")
            st.caption("Ces contraintes s'appliquent après les filtres de base "
                       "et affinent la construction du portefeuille.")

            copt1, copt2 = st.columns(2)

            with copt1:
                use_sector_c = st.toggle(
                    "④ Contrainte sectorielle",
                    value=False, key="use_sect_c",
                    help="Limite la concentration maximale dans un seul secteur"
                )
                if use_sector_c:
                    max_sect_pct = st.slider(
                        "Poids max par secteur (%)", 10, 80, 40, 5,
                        key="max_sect_pct",
                        help="Si un secteur dépasse ce seuil après pondération, "
                             "les titres les moins bien scorés de ce secteur sont retirés"
                    ) / 100
                    st.caption(f"Aucun secteur ne dépassera **{max_sect_pct:.0%}** du portefeuille")
                else:
                    max_sect_pct = 1.0

            with copt2:
                use_corr_c = st.toggle(
                    "⑤ Corrélation effective du portefeuille",
                    value=False, key="use_corr_c",
                    help="Sélectionne les titres qui minimisent la corrélation "
                         "moyenne du portefeuille (greedy forward selection)"
                )
                if use_corr_c:
                    max_avg_corr = st.slider(
                        "Corrélation moyenne max du portefeuille", 0.1, 0.9, 0.50, 0.05,
                        key="max_avg_corr",
                        help="Un titre n'est ajouté que si sa corrélation moyenne "
                             "avec les titres déjà sélectionnés est ≤ ce seuil"
                    )
                    st.caption(f"Corrélation moyenne cible : **≤ {max_avg_corr:.2f}**")
                else:
                    max_avg_corr = 1.0

            if not use_sector_c and not use_corr_c:
                st.info("💡 Pipeline standard (4 filtres) · Activez les contraintes "
                        "ci-dessus pour enrichir l'optimisation")

            if st.button("🚀 Lancer le pipeline", type="primary"):
                with st.spinner("Optimisation en cours..."):
                    inc,pw,pipeline_log = run_optimization_pipeline(
                        mf_scores=mf, data=data, factor_results=fr,
                        min_vol_pct=min_vol_pct if liq_ok else 0,
                        top_pct=top_pct, max_corr=max_corr,
                        use_markowitz=use_mkz, risk_aversion=risk_aversion,
                        mf_weight=mf_weight, min_w=min_w, max_w=max_w,
                        window=int(window_corr),
                        use_sector_constraint=use_sector_c,
                        max_sector_pct=max_sect_pct,
                        use_corr_constraint=use_corr_c,
                        max_avg_corr=max_avg_corr,
                    )
                if pw is None or len(pw)==0:
                    st.error("❌ Aucun titre retenu — élargissez les seuils des filtres.")
                else:
                    st.session_state.pw         = pw
                    st.session_state.pw_log     = pipeline_log
                    st.session_state.pw_use_mkz = use_mkz
                    st.session_state.pw_mode    = "auto"
                    st.success(f"✅ {len(pw)} titres retenus · Σα = {pw.sum():.6f}")

        # ══════════════════════════════════════════════════════════
        # MODE B — SÉLECTION MANUELLE
        # ══════════════════════════════════════════════════════════
        else:
            st.markdown("""<div class='fbox'>
            Vous choisissez librement les titres à inclure, en vous appuyant sur le classement MF.<br>
            Les pondérations α sont ensuite calculées sur cet univers restreint.
            </div>""", unsafe_allow_html=True)

            with st.expander("📊 Classement MF — aide à la sélection", expanded=True):
                mf_disp = mf.reset_index()
                mf_disp.columns = ["Ticker","Score MF"]
                mf_disp.insert(0,"Rang",range(1,len(mf_disp)+1))
                mf_disp["Score MF"] = mf_disp["Score MF"].round(6)
                st.dataframe(
                    mf_disp.style.format({"Score MF":"{:.6f}"})
                           .bar(subset=["Score MF"],color=["#1e3a5f","#3b82f6"]),
                    width="stretch", hide_index=True, height=300
                )

            included = st.multiselect(
                "✅ Titres à inclure dans le portefeuille",
                options=all_tickers, default=all_tickers,
                help="Consultez le classement MF ci-dessus pour guider votre sélection"
            )
            n_inc = len(included)
            if n_inc == 0:
                st.warning("⚠️ Sélectionnez au moins un titre.")
            elif n_inc < len(all_tickers):
                st.info(f"ℹ️ {n_inc} titres sélectionnés sur {len(all_tickers)}")
            else:
                st.success(f"✅ Univers complet ({n_inc} titres)")

            st.markdown("---")
            st.markdown("**④ Pondération finale**")
            use_mkz_m = st.toggle("Optimisation Markowitz MF-augmentée", value=False,
                key="mkz_manual",
                help="OFF = poids par rang MF · ON = Mean-Variance avec score MF")
            if use_mkz_m:
                mm1,mm2,mm3,mm4 = st.columns(4)
                mf_weight_m     = mm1.slider("Poids Score MF",    0.0,1.0,0.50,0.05,key="mfw_m")
                risk_aversion_m = mm2.slider("Aversion risque λ", 0.1,10.0,2.0,0.1,key="rav_m")
                min_w_m         = mm3.slider("Poids min/titre (%)",0,20,0,1,key="mnw_m")/100
                max_w_m         = mm4.slider("Poids max/titre (%)",5,100,40,5,key="mxw_m")/100
            else:
                mf_weight_m,risk_aversion_m,min_w_m,max_w_m = 0.5,2.0,0.0,1.0

            if st.button("📂 Construire le portefeuille", type="primary") and n_inc > 0:
                with st.spinner("Calcul des poids..."):
                    if use_mkz_m:
                        pw = optimize_markowitz(mf, included, data["cours"],
                            window=252, risk_aversion=risk_aversion_m,
                            mf_weight=mf_weight_m, min_w=min_w_m, max_w=max_w_m)
                        if pw is None:
                            pw = compute_portfolio_weights(mf, included=included)
                            st.warning("⚠️ Markowitz indisponible — poids rang MF appliqués")
                    else:
                        pw = compute_portfolio_weights(mf, included=included)
                st.session_state.pw         = pw
                st.session_state.pw_log     = [("Sélection manuelle",n_inc,included),
                                               ("Poids finaux",len(pw),pw.index.tolist())]
                st.session_state.pw_use_mkz = use_mkz_m
                st.session_state.pw_mode    = "manuel"
                st.success(f"✅ {len(pw)} titres · Σα = {pw.sum():.6f}")

        # ══════════════════════════════════════════════════════════
        # RÉSULTATS COMMUNS AUX DEUX MODES
        # ══════════════════════════════════════════════════════════
        if st.session_state.pw is not None:
            pw        = st.session_state.pw
            log       = st.session_state.get("pw_log",[])
            mode_used = st.session_state.get("pw_mode","")

            st.markdown("---")

            # Trace pipeline (mode auto uniquement)
            if log and mode_used == "auto":
                st.markdown("**🔍 Trace du pipeline**")
                log_cols  = st.columns(len(log))
                clr_steps = ["#475569","#3b82f6","#8b5cf6","#06b6d4","#10b981"]
                for i,(step,n_step,_) in enumerate(log):
                    with log_cols[i]:
                        st.markdown(
                            f"<div style='background:#161d2e;border:1px solid #1e2d45;"
                            f"border-radius:8px;padding:10px;text-align:center;'>"
                            f"<div style='font-size:10px;color:{clr_steps[i%len(clr_steps)]};font-weight:600;"
                            f"letter-spacing:.08em;text-transform:uppercase;'>{step}</div>"
                            f"<div style='font-size:24px;font-weight:800;color:#e2e8f0;'>{n_step}</div>"
                            f"<div style='font-size:10px;color:#64748b;'>titres</div></div>",
                            unsafe_allow_html=True)
                st.markdown("")

            # KPIs
            k1,k2,k3,k4,k5 = st.columns(5)
            k1.metric("Nb titres",         len(pw))
            k2.metric("Poids max",         f"{pw.max()*100:.2f}%")
            k3.metric("Poids min",         f"{pw.min()*100:.2f}%")
            k4.metric("HHI concentration", f"{(pw**2).sum():.4f}")
            k5.metric("Mode", "🤖 Auto" if mode_used=="auto" else "🖐️ Manuel")

            pl,pr = st.columns(2)
            with pl:
                st.markdown("**Répartition du portefeuille**")
                dpw = pw.copy()
                if len(dpw)>15:
                    dpw = pd.concat([dpw.head(15),pd.Series({"Autres":dpw.iloc[15:].sum()})])
                fig_p = go.Figure(go.Pie(
                    labels=dpw.index,values=dpw.values,hole=0.45,
                    textfont=dict(size=10),
                    marker=dict(line=dict(color="#0b0f1a",width=2)),
                    hovertemplate="<b>%{label}</b><br>%{percent:.2%}<extra></extra>"))
                fig_p.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)",height=380,
                    font=dict(color="#94a3b8",family="JetBrains Mono"),
                    legend=dict(font=dict(size=10),bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10,r=10,t=20,b=10),
                    annotations=[dict(text=f"<b>{len(pw)}</b><br>titres",
                        x=0.5,y=0.5,font_size=14,showarrow=False,
                        font=dict(color="#94a3b8"))])
                st.plotly_chart(fig_p,width="stretch")

            with pr:
                st.markdown("**Poids par titre**")
                pt = pw.head(25)
                fig_h = go.Figure(go.Bar(
                    x=pt.values*100,y=pt.index,orientation="h",
                    marker=dict(color=np.linspace(0.9,0.2,len(pt)),colorscale="Blues"),
                    text=[f"{v*100:.2f}%" for v in pt.values],
                    textposition="outside",textfont=dict(size=9,color="#94a3b8")))
                fig_h.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                    height=480,font=dict(color="#94a3b8",family="JetBrains Mono"),
                    margin=dict(l=10,r=80,t=20,b=30),
                    xaxis=dict(gridcolor="#1e2d45",ticksuffix="%"),
                    yaxis=dict(autorange="reversed"))
                st.plotly_chart(fig_h,width="stretch")

            # Table
            ranks_final = mf.reindex(pw.index).rank(ascending=False,method="min")
            alloc = pd.DataFrame({
                "Ticker":       pw.index,
                "Rang MF":      ranks_final.loc[pw.index].astype(int),
                "Score MF":     mf.loc[pw.index].round(6),
                "Poids α(T,t)": pw.values,
                "Poids (%)":    (pw.values*100).round(4),
            }).reset_index(drop=True)
            st.markdown("**Table d'allocation complète**")
            st.dataframe(
                alloc.style.format({"Score MF":"{:.6f}","Poids α(T,t)":"{:.6f}","Poids (%)":"{:.4f}%"})
                     .bar(subset=["Poids (%)"],color=["#1e3a5f","#3b82f6"]),
                width="stretch",hide_index=True)

            buf = io.BytesIO(); alloc.to_excel(buf,index=False)
            st.download_button("⬇️ Exporter le portefeuille (Excel)", buf.getvalue(),
                "portefeuille_BRVM.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # ── Portefeuille Charia ────────────────────────────────
            if CHARIA_AVAILABLE and st.session_state.charia_results:
                st.markdown("---")
                st.markdown(
                    "<span class='pill' style='background:#10b981;'>☪️</span>"
                    "<p class='sh'>Allocation Charia Compatible</p>",
                    unsafe_allow_html=True)
                st.markdown("""<div class='fbox'>
                Sous-ensemble du portefeuille filtré sur les titres ≥ 3/4 standards Charia<br>
                DJIM · FTSE · S&amp;P · AAOIFI — poids recalculés sur cet univers restreint
                </div>""", unsafe_allow_html=True)

                charia_tickers = get_charia_compatible_tickers(
                    st.session_state.charia_results)
                charia_in_ptf  = [t for t in pw.index if t in charia_tickers]

                ck1, ck2, ck3 = st.columns(3)
                ck1.metric("Compatibles Charia (BRVM)", len(charia_tickers))
                ck2.metric("Dans le portefeuille cible", len(charia_in_ptf))
                ck3.metric("Σα Charia", f"{compute_portfolio_weights(mf, included=charia_in_ptf).sum():.4f}"
                           if charia_in_ptf else "—")

                if charia_in_ptf:
                    pw_charia = compute_portfolio_weights(mf, included=charia_in_ptf)

                    ch_l, ch_r = st.columns(2)
                    with ch_l:
                        st.markdown("**Répartition Charia**")
                        dpwc = pw_charia.copy()
                        if len(dpwc) > 12:
                            dpwc = pd.concat([dpwc.head(12),
                                             pd.Series({"Autres": dpwc.iloc[12:].sum()})])
                        fig_pc = go.Figure(go.Pie(
                            labels=dpwc.index, values=dpwc.values, hole=0.45,
                            textfont=dict(size=10),
                            marker=dict(line=dict(color="#0b0f1a", width=2)),
                            hovertemplate="<b>%{label}</b><br>%{percent:.2%}<extra></extra>"
                        ))
                        fig_pc.update_layout(
                            paper_bgcolor="rgba(0,0,0,0)", height=320,
                            font=dict(color="#94a3b8", family="JetBrains Mono"),
                            legend=dict(font=dict(size=9), bgcolor="rgba(0,0,0,0)"),
                            margin=dict(l=10,r=10,t=20,b=10),
                            annotations=[dict(text=f"<b>{len(pw_charia)}</b><br>titres",
                                             x=0.5,y=0.5,font_size=13,showarrow=False,
                                             font=dict(color="#10b981"))]
                        )
                        st.plotly_chart(fig_pc, width="stretch")

                    with ch_r:
                        st.markdown("**Poids par titre**")
                        fig_hc = go.Figure(go.Bar(
                            x=pw_charia.values*100, y=pw_charia.index,
                            orientation="h",
                            marker=dict(color="#10b981", opacity=0.85),
                            text=[f"{v*100:.2f}%" for v in pw_charia.values],
                            textposition="outside",
                            textfont=dict(size=9, color="#94a3b8")
                        ))
                        fig_hc.update_layout(
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            height=360, font=dict(color="#94a3b8", family="JetBrains Mono"),
                            margin=dict(l=10,r=80,t=20,b=30),
                            xaxis=dict(gridcolor="#1e2d45", ticksuffix="%"),
                            yaxis=dict(autorange="reversed")
                        )
                        st.plotly_chart(fig_hc, width="stretch")

                    ranks_c = mf.reindex(pw_charia.index).rank(
                        ascending=False, method="min")
                    alloc_c = pd.DataFrame({
                        "Ticker":       pw_charia.index,
                        "Rang MF":      ranks_c.loc[pw_charia.index].astype(int),
                        "Score MF":     mf.loc[pw_charia.index].round(6),
                        "☪️ Charia":   [get_charia_label(t,st.session_state.charia_results)
                                         for t in pw_charia.index],
                        "Poids α(T,t)": pw_charia.values,
                        "Poids (%)":    (pw_charia.values*100).round(4),
                    }).reset_index(drop=True)

                    st.dataframe(
                        alloc_c.style.format({
                            "Score MF":     "{:.6f}",
                            "Poids α(T,t)": "{:.6f}",
                            "Poids (%)":    "{:.4f}%"
                        }), width="stretch", hide_index=True
                    )
                    buf_c = io.BytesIO(); alloc_c.to_excel(buf_c, index=False)
                    st.download_button("⬇️ Exporter portefeuille Charia (Excel)",
                        buf_c.getvalue(), "portefeuille_charia_BRVM.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                else:
                    st.info("☪️ Aucun titre Charia compatible dans le portefeuille cible. "
                            "Élargissez les filtres ou chargez le fichier de screening.")

            # ── Concentration sectorielle du portefeuille ──────
            if SECTOR_MAP:
                st.markdown("---")
                st.markdown("**🏭 Concentration sectorielle du portefeuille**")
                sect_pf = sector_allocation(pw, SECTOR_MAP, data["tickers"])
                sect_bm = benchmark_sector_weights(SECTOR_MAP, data["cours"])
                all_s   = sorted(set(sect_pf.index) | set(sect_bm.index))

                fig_sp2 = go.Figure()
                fig_sp2.add_trace(go.Bar(
                    name="Portefeuille", x=all_s,
                    y=[sect_pf.get(s,0)*100 for s in all_s],
                    marker_color="#3b82f6", opacity=0.85))
                fig_sp2.add_trace(go.Bar(
                    name="BRVM Composite", x=all_s,
                    y=[sect_bm.get(s,0)*100 for s in all_s],
                    marker_color="#f59e0b", opacity=0.6))
                fig_sp2.update_layout(barmode="group",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8", family="JetBrains Mono"), height=320,
                    legend=dict(orientation="h", y=1.08, bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10,r=10,t=40,b=80),
                    xaxis=dict(gridcolor="#1e2d45", tickangle=-30),
                    yaxis=dict(gridcolor="#1e2d45", ticksuffix="%"))
                st.plotly_chart(fig_sp2, width="stretch")
                st.caption("Pour une analyse complète → onglet 🔗 Corrélations & Diversification")


# ══ VALORISATION ═══════════════════════════════════════════════
with t8:
    st.markdown("<span class='pill'>Valorisation · Prix Cible</span><p class='sh'>Modèles de valorisation théorique</p>", unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    DDM · P/E relatif · P/B · DCF simplifié → Prix cible combiné par titre<br>
    Chargez votre fichier d'états financiers · La base est persistée entre les sessions
    </div>""", unsafe_allow_html=True)

    if not VALUATION_AVAILABLE:
        st.error("❌ Modules fs_parser.py et valuation_models.py introuvables.")
        st.code(_VALUATION_ERROR if '_VALUATION_ERROR' in dir() else "Erreur inconnue")
        st.info("Vérifiez que fs_parser.py et valuation_models.py sont à la racine du dépôt GitHub (même niveau que app.py).")
        st.stop()

    # ── Chargement du fichier états financiers ─────────────────
    st.markdown("**📁 États Financiers**")
    col_up, col_db = st.columns([2, 1])

    with col_up:
        fs_file = st.file_uploader(
            "Charger Etats_Financiers_BRVM.xlsx (nouveau chargement = mise à jour cumulative)",
            type=["xlsx","xls"], key="fs_uploader"
        )
        if fs_file:
            with st.spinner("Parsing des états financiers..."):
                new_data = parse_financial_file(fs_file)
                nb_titres_tmp = st.session_state.data.get("nb_titres", {}) \
                                if st.session_state.data else {}
                cours_tmp     = st.session_state.data.get("cours", pd.DataFrame()) \
                                if st.session_state.data else pd.DataFrame()
                moy_tmp       = st.session_state.data.get("moyenne_cours", pd.DataFrame()) \
                                if st.session_state.data else pd.DataFrame()
                if nb_titres_tmp and not cours_tmp.empty:
                    from fs_parser import validate_and_fix_units
                    new_data, corrections = validate_and_fix_units(
                        new_data, nb_titres_tmp, cours_tmp
                    )
                    if corrections:
                        st.warning(
                            f"⚠️ Correction automatique d'unités : "
                            + ", ".join(f"{t} (PE {v['pe_avant']}x→{v['pe_apres']}x)"
                                        for t, v in corrections.items())
                        )
                existing = load_financial_db(VALUATION_DB_PATH)
                merged   = merge_financial_data(existing, new_data)
                save_financial_db(merged, VALUATION_DB_PATH)
                st.session_state.fin_data = merged

                # Sauvegarde GitHub silencieuse
                if GITHUB_STORAGE and is_github_configured():
                    ok_fin = save_financial_db_github(merged)
                    ok_ch  = save_charia_results(st.session_state.charia_results)
                    if ok_fin:
                        st.caption("☁️ États financiers sauvegardés sur GitHub")

                # ── Screening Charia automatique depuis fin_data ──────
                if CHARIA_AVAILABLE and nb_titres_tmp:
                    charia_auto = screen_all_from_fin_data(
                        merged, nb_titres_tmp, moy_tmp
                    )
                    # Fusion avec screening existant (priorité au fichier statique)
                    existing_charia = st.session_state.charia_results
                    for t, r in charia_auto.items():
                        if t not in existing_charia:
                            existing_charia[t] = r
                    st.session_state.charia_results = existing_charia

            tickers_new = set(new_data.keys()) - set(existing.keys())
            years_new   = {yr for t in new_data.values() for yr in t.keys()}
            n_charia    = len(get_charia_compatible_tickers(
                              st.session_state.charia_results)) \
                          if CHARIA_AVAILABLE else 0
            st.success(
                f"✅ {len(new_data)} sociétés · Années : {sorted(years_new)} · "
                f"Base : {len(merged)} · ☪️ Charia compatibles : {n_charia}"
            )

    # ── Chargement fichier screening Charia statique ───────────
    if CHARIA_AVAILABLE:
        with st.expander("☪️ Charger le fichier de screening Charia (optionnel)", expanded=False):
            charia_file = st.file_uploader(
                "CHARIA_SCREENING_MODEL.xlsx — enrichit le screening automatique",
                type=["xlsx","xls"], key="charia_uploader"
            )
            if charia_file:
                with st.spinner("Parsing screening Charia..."):
                    charia_static = parse_charia(charia_file)
                    # Le fichier statique a priorité sur le calcul auto
                    merged_charia = {**st.session_state.charia_results, **charia_static}
                    st.session_state.charia_results = merged_charia
                n_ok = len(get_charia_compatible_tickers(merged_charia))
                st.success(f"✅ {len(charia_static)} tickers screenés · "
                           f"☪️ {n_ok} compatibles Charia")
                # Sauvegarde GitHub
                if GITHUB_STORAGE and is_github_configured():
                    with st.spinner("💾 Sauvegarde Charia GitHub..."):
                        save_charia_results(merged_charia)

    with col_db:
        fin_data = st.session_state.fin_data
        if fin_data:
            all_years = sorted({yr for t in fin_data.values() for yr in t.keys()})
            st.metric("Sociétés en base", len(fin_data))
            st.metric("Années disponibles", f"{min(all_years)}–{max(all_years)}")
            if os.path.exists(VALUATION_DB_PATH):
                db_data = json.load(open(VALUATION_DB_PATH))
                st.caption(f"Dernière mise à jour : {db_data.get('metadata',{}).get('last_updated','—')[:10]}")
        else:
            st.info("Aucune donnée en base. Chargez un fichier.")

    _has_fin_data = bool(fin_data)
    if not _has_fin_data:
        st.info("📂 Aucune donnée en base. Chargez votre fichier **Etats_Financiers_BRVM.xlsx** ci-dessus pour commencer.")

    if _has_fin_data:
        st.markdown("---")

        # ── Info calibrage automatique ─────────────────────────────
        st.markdown("""<div class='fbox'>
        🤖 <b>Calibrage 100% automatique</b> — Zéro saisie manuelle<br>
        β calculé depuis cours historiques · ke/WACC via CAPM · kd depuis états financiers<br>
        P/E et P/B calibrés sur les multiples historiques observés du titre · g_FCF = CAGR Rex/RN
        </div>""", unsafe_allow_html=True)

        st.markdown("**📅 Fenêtre de calcul du Beta (β)**")
        st.caption("Le beta mesure la sensibilité du titre au marché BRVM sur la période choisie. "
                   "Une fenêtre longue (3–5 ans) donne un beta stable · courte (6–12 mois) capture le beta récent.")

        c_min_v = data["cours"].index.min().date() if st.session_state.data else pd.Timestamp("2014-01-01").date()
        c_max_v = data["cours"].index.max().date() if st.session_state.data else pd.Timestamp("2026-01-01").date()

        bc1, bc2, bc3 = st.columns(3)
        with bc1:
            beta_start = st.date_input(
                "Début fenêtre β", key="beta_start",
                value=max(c_min_v, min(c_max_v, pd.Timestamp("2022-01-01").date())),
                min_value=c_min_v, max_value=max(c_max_v, datetime.date.today()),
                help="Date de début pour estimer le beta"
            )
        with bc2:
            beta_end = st.date_input(
                "Fin fenêtre β", key="beta_end",
                value=c_max_v,
                min_value=c_min_v, max_value=max(c_max_v, datetime.date.today()),
                help="Date de fin pour estimer le beta (en général = aujourd'hui)"
            )
        with bc3:
            if beta_start < beta_end and st.session_state.data:
                cours_tmp  = st.session_state.data["cours"]
                mask_beta  = (cours_tmp.index >= pd.to_datetime(beta_start)) & \
                             (cours_tmp.index <= pd.to_datetime(beta_end))
                n_jours = mask_beta.sum()
                st.markdown("<br>", unsafe_allow_html=True)
                st.success(f"✅ **{n_jours}** jours de trading\n\n"
                           f"≈ {n_jours//252} an(s) {(n_jours%252)//21} mois")
            elif beta_start >= beta_end:
                st.markdown("<br>", unsafe_allow_html=True)
                st.error("⚠️ Date début ≥ Date fin")

        st.markdown("---")

        ac1, ac2 = st.columns(2)
        with ac1:
            dcf_horizon = st.slider(
                "Horizon DCF (ans)", 3, 10, 5, 1, key="dcf_h",
                help="Nombre d'années de projection explicite. La valeur terminale couvre l'infini."
            )
        with ac2:
            st.markdown("**Poids des modèles (renormalisés si modèle absent)**")
            wp1, wp2, wp3, wp4 = st.columns(4)
            w_ddm = wp1.number_input("DDM",  0.0,1.0,0.20,0.05,key="w_ddm")
            w_pe  = wp2.number_input("P/E",  0.0,1.0,0.30,0.05,key="w_pe")
            w_pb  = wp3.number_input("P/B",  0.0,1.0,0.20,0.05,key="w_pb")
            w_dcf = wp4.number_input("DCF",  0.0,1.0,0.30,0.05,key="w_dcf")
        model_weights = {"DDM": w_ddm, "P/E": w_pe, "P/B": w_pb, "DCF": w_dcf}

        st.markdown("---")

        # ── Sélection des titres ────────────────────────────────────
        av_tickers = sorted(fin_data.keys())
        if st.session_state.data:
            cours_tickers = st.session_state.data.get("tickers", [])
            common   = [t for t in av_tickers if t in cours_tickers]
            only_fs  = [t for t in av_tickers if t not in cours_tickers]
        else:
            common, only_fs = av_tickers, []

        sel_tickers = st.multiselect(
            "Titres à valoriser",
            options=av_tickers,
            default=common[:15] if len(common) >= 15 else common,
            help="Titres disponibles dans les états financiers chargés"
        )

        if st.button("📈 Calculer les valorisations", type="primary"):
            if not sel_tickers:
                st.warning("Sélectionnez au moins un titre.")
            else:
                nb_titres  = st.session_state.data.get("nb_titres",  {}) if st.session_state.data else {}
                dividendes = st.session_state.data.get("dividendes", pd.DataFrame()) if st.session_state.data else pd.DataFrame()
                cours_df   = st.session_state.data.get("cours",      pd.DataFrame()) if st.session_state.data else pd.DataFrame()
                moy_cours  = st.session_state.data.get("moyenne_cours", pd.DataFrame()) if st.session_state.data else pd.DataFrame()

                # Dividendes → {ticker: {year: montant}}
                div_hist = {}
                if not dividendes.empty and "Date" in dividendes.columns:
                    for _, row in dividendes.iterrows():
                        yr_d = int(row["Date"])
                        for t in sel_tickers:
                            if t in dividendes.columns:
                                v = row.get(t)
                                if pd.notna(v) and v > 0:
                                    div_hist.setdefault(t, {})[yr_d] = float(v)

                results_all = {}
                params_all  = {}
                with st.spinner(f"Calibrage auto + valorisation de {len(sel_tickers)} titres..."):
                    for ticker in sel_tickers:
                        # ── Calibrage auto de tous les paramètres ──────
                        try:
                            from valuation_models import calibrate_params
                            p = calibrate_params(
                                ticker, fin_data, nb_titres,
                                cours_df, moy_cours,
                                div_history=div_hist.get(ticker),
                                beta_date_start=beta_start,
                                beta_date_end=beta_end,
                            )
                        except Exception:
                            p = None
                        params_all[ticker] = p

                        res = {}
                        res["DDM"] = valuation_ddm(
                            ticker, fin_data, div_hist.get(ticker, {}),
                            nb_titres, cours_df, params=p
                        )
                        res["P/E"] = valuation_pe(
                            ticker, fin_data, nb_titres,
                            cours_df, moy_cours, params=p
                        )
                        res["P/B"] = valuation_pb(
                            ticker, fin_data, nb_titres,
                            cours_df, moy_cours, params=p
                        )
                        res["DCF"] = valuation_dcf(
                            ticker, fin_data, nb_titres,
                            cours_df, horizon=dcf_horizon, params=p
                        )
                        p_comb, prices_ok = combined_price(res, model_weights)
                        results_all[ticker] = {
                            "modeles":      res,
                            "prix_cible":   p_comb,
                            "prix_modeles": prices_ok,
                            "params":       p,
                        }

                st.session_state.val_results = results_all
                st.success(f"✅ Valorisation terminée — {len(results_all)} titres")

        # ── Affichage des résultats ────────────────────────────────
        if "val_results" in st.session_state and st.session_state.val_results:
            val      = st.session_state.val_results
            cours_df = st.session_state.data.get("cours", pd.DataFrame()) if st.session_state.data else pd.DataFrame()

            st.markdown("---")
            st.markdown("**📋 Tableau de synthèse — Prix cibles et potentiels**")

            rows = []
            for ticker, v in val.items():
                p_comb = v.get("prix_cible")
                if not p_comb:
                    continue

                # Cours actuel
                cours_act = np.nan
                if not cours_df.empty and ticker in cours_df.columns:
                    s = cours_df[ticker].dropna()
                    if not s.empty:
                        cours_act = float(s.iloc[-1])

                potentiel = (p_comb - cours_act) / cours_act * 100 \
                            if not np.isnan(cours_act) and cours_act > 0 else np.nan
                signal = ("🟢 Achat" if potentiel > 10 else
                          "🔴 Vente" if potentiel < -10 else
                          "🟡 Neutre") if not np.isnan(potentiel) else "—"

                pm = v.get("prix_modeles", {})

                # Secteur
                sect = SECTOR_MAP.get(ticker.upper())
                if not sect:
                    fd = fin_data.get(ticker, {})
                    yr_last = max(fd.keys()) if fd else None
                    sect = fd.get(yr_last, {}).get("secteur", "—") \
                           if fd and yr_last else "—"

                pot_num = potentiel if not np.isnan(potentiel) else 0
                charia_lbl = get_charia_label(ticker, st.session_state.charia_results) \
                             if CHARIA_AVAILABLE else "—"

                rows.append({
                    "Ticker":      ticker,
                    "Secteur":     sect,
                    "☪️ Charia":   charia_lbl,
                    "Cours actuel": f"{cours_act:,.0f}" if not np.isnan(cours_act) else "—",
                    "DDM":          f"{pm.get('DDM',0):,.0f}"  if "DDM" in pm  else "—",
                    "P/E":          f"{pm.get('P/E',0):,.0f}"  if "P/E" in pm  else "—",
                    "P/B":          f"{pm.get('P/B',0):,.0f}"  if "P/B" in pm  else "—",
                    "DCF":          f"{pm.get('DCF',0):,.0f}"  if "DCF" in pm  else "—",
                    "Prix cible":   f"{p_comb:,.0f}",
                    "Potentiel":    f"{potentiel:+.1f}%" if not np.isnan(potentiel) else "—",
                    "Signal":       signal,
                    "_pot_num":     pot_num,
                })

            if not rows:
                st.warning("Aucun résultat disponible — vérifiez que les nb_titres sont chargés.")
            else:
                df_res = pd.DataFrame(rows).sort_values("_pot_num", ascending=False)
                display_cols = ["Ticker","Secteur","☪️ Charia","Cours actuel",
                            "DDM","P/E","P/B","DCF","Prix cible","Potentiel","Signal"]
                st.dataframe(df_res[display_cols], width="stretch", hide_index=True)

                # ── Téléchargement Excel détaillé par titre ────────────
                st.markdown("**📥 Fichier Excel détaillé par titre**")
                st.caption("Cliquez sur un ticker pour télécharger son fichier de valorisation détaillé")

                @st.cache_data(show_spinner=False)
                def build_valuation_excel_cached(ticker, _val_result, _fin_data, _cours_df, cache_key):
                    """Wrapper caché — cache_key force l'invalidation si les résultats changent."""
                    return build_valuation_excel(ticker, _val_result, _fin_data, _cours_df)

                def build_valuation_excel(ticker, val_result, fin_data, cours_df):
                    """
                    Génère un fichier Excel détaillé de valorisation pour un titre.
                    4 feuilles : Synthèse · P/E · P/B · DCF (+ DDM si disponible)
                    Avec les données brutes, les paramètres et les calculs explicités.
                    """
                    from openpyxl import Workbook
                    from openpyxl.styles import (Font, PatternFill, Alignment,
                                                  Border, Side, numbers)
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()

                    # ── Styles ──────────────────────────────────────────
                    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
                    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                    SUB_FILL = PatternFill("solid", fgColor="2D5A8E")
                    SUB_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                    BLU_FONT = Font(color="0000FF", name="Arial", size=9)  # Inputs
                    BLK_FONT = Font(color="000000", name="Arial", size=9)  # Formules
                    GRN_FONT = Font(color="006400", name="Arial", size=9)  # Liens
                    BLD_FONT = Font(bold=True, name="Arial", size=9)
                    NRM_FONT = Font(name="Arial", size=9)
                    YLW_FILL = PatternFill("solid", fgColor="FFFF00")  # Hypothèses
                    LBL_FILL = PatternFill("solid", fgColor="EBF5FB")
                    ALT_FILL = PatternFill("solid", fgColor="F8F9FA")
                    thin     = Side(style="thin", color="CCCCCC")
                    BORDER   = Border(left=thin, right=thin, top=thin, bottom=thin)
                    CNTR     = Alignment(horizontal="center", vertical="center")
                    RGHT     = Alignment(horizontal="right",  vertical="center")
                    LEFT     = Alignment(horizontal="left",   vertical="center")

                    def hdr(ws, row, col, val, w=None):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = HDR_FONT; c.fill = HDR_FILL
                        c.alignment = CNTR; c.border = BORDER
                        if w: ws.column_dimensions[get_column_letter(col)].width = w
                        return c

                    def sub(ws, row, col, val):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = SUB_FONT; c.fill = SUB_FILL
                        c.alignment = LEFT; c.border = BORDER
                        return c

                    def inp(ws, row, col, val, fmt=None):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = BLU_FONT; c.fill = YLW_FILL
                        c.border = BORDER; c.alignment = RGHT
                        if fmt: c.number_format = fmt
                        return c

                    def lbl(ws, row, col, val):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = BLD_FONT; c.fill = LBL_FILL
                        c.border = BORDER; c.alignment = LEFT
                        return c

                    def val_cell(ws, row, col, val, fmt=None, bold=False):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = BLD_FONT if bold else NRM_FONT
                        c.border = BORDER; c.alignment = RGHT
                        if fmt: c.number_format = fmt
                        return c

                    def note(ws, row, col, val):
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = Font(name="Arial", size=8,
                                      color="666666", italic=True)
                        c.alignment = LEFT
                        return c

                    # Paramètres du titre
                    p    = val_result.get("params", {}) or {}
                    mods = val_result.get("modeles", {})
                    pc   = val_result.get("prix_cible")
                    fd   = fin_data.get(ticker, {})
                    yr   = max(fd.keys()) if fd else "—"
                    postes = fd.get(yr, {}) if fd else {}

                    cours_act = np.nan
                    if cours_df is not None and ticker in cours_df.columns:
                        s = cours_df[ticker].dropna()
                        if not s.empty:
                            cours_act = float(s.iloc[-1])
                    potentiel = (pc - cours_act)/cours_act if (
                        pc and not np.isnan(cours_act) and cours_act > 0) else np.nan

                    # ══════════════════════════════════════════════
                    # FEUILLE 1 — SYNTHÈSE
                    # ══════════════════════════════════════════════
                    ws1 = wb.active
                    ws1.title = "Synthèse"
                    ws1.column_dimensions["A"].width = 32
                    ws1.column_dimensions["B"].width = 22
                    ws1.column_dimensions["C"].width = 22
                    ws1.column_dimensions["D"].width = 28

                    hdr(ws1, 1, 1, f"VALORISATION — {ticker}")
                    ws1.merge_cells("A1:D1")
                    ws1.cell(1,1).alignment = CNTR
                    ws1.row_dimensions[1].height = 22

                    lbl(ws1, 2, 1, "Ticker");     val_cell(ws1, 2, 2, ticker, bold=True)
                    lbl(ws1, 3, 1, "Secteur");    val_cell(ws1, 3, 2, SECTOR_MAP.get(
                                                    ticker.upper(), "—"))
                    lbl(ws1, 4, 1, "Année réf."); val_cell(ws1, 4, 2, yr)
                    lbl(ws1, 5, 1, "Cours actuel (FCFA)");
                    val_cell(ws1, 5, 2,
                             cours_act if not np.isnan(cours_act) else "—",
                             fmt='#,##0', bold=True)
                    lbl(ws1, 6, 1, "Prix cible combiné (FCFA)");
                    val_cell(ws1, 6, 2,
                             round(pc) if pc else "—",
                             fmt='#,##0', bold=True)
                    lbl(ws1, 7, 1, "Potentiel (%)")
                    c = val_cell(ws1, 7, 2,
                                 potentiel if not np.isnan(potentiel) else "—",
                                 fmt='0.00%')
                    if not np.isnan(potentiel) if isinstance(potentiel, float) else False:
                        c.font = Font(bold=True, color="006400" if potentiel > 0
                                      else "FF0000", name="Arial", size=9)

                    # Paramètres calibrés automatiquement
                    sub(ws1, 9, 1, "Paramètres calibrés automatiquement")
                    ws1.merge_cells("A9:D9")
                    hdr(ws1, 10, 1, "Paramètre", w=32)
                    hdr(ws1, 10, 2, "Valeur", w=22)
                    hdr(ws1, 10, 3, "Description", w=22)
                    hdr(ws1, 10, 4, "Source", w=28)

                    params_rows = [
                        ("Beta (β)",         p.get("beta",1),    "0.000",
                         "Régression OLS cours vs marché BRVM", "Cours historiques"),
                        ("ke (taux actua.)",  p.get("ke",0),      "0.0%",
                         "Rf + β × prime_risque", "CAPM / BCEAO"),
                        ("kd (coût dette)",   p.get("kd",0),      "0.0%",
                         "Intérêts / Dette financière", "États financiers"),
                        ("WACC",              p.get("wacc",0),    "0.0%",
                         "ke×E/(D+E) + kd×(1-t)×D/(D+E)", "Calculé"),
                        ("g FCF",             p.get("g_fcf",0),   "0.0%",
                         "CAGR Rex/RN historique", "États financiers"),
                        ("g dividendes",      p.get("g_div",0),   "0.0%",
                         "CAGR dividendes historiques", "Données de marché"),
                        ("P/E cible",         p.get("pe_target",0),"0.0x",
                         p.get("pe_method","—"), "Médiane historique"),
                        ("P/B cible",         p.get("pb_target",0),"0.00x",
                         p.get("pb_method","—"), "Médiane historique"),
                        ("Taux IS effectif",  p.get("tax_rate",0),"0.0%",
                         "Impôts / Rex observé", "États financiers"),
                    ]
                    for i, (pname, pval, fmt, desc, src) in enumerate(params_rows):
                        r = 11 + i
                        fill = ALT_FILL if i % 2 else PatternFill()
                        for col in range(1, 5):
                            ws1.cell(r, col).fill = fill
                            ws1.cell(r, col).border = BORDER
                        lbl(ws1, r, 1, pname)
                        v = pval if isinstance(pval, str) else round(pval, 6)
                        inp(ws1, r, 2, v,
                            fmt="0.000" if "beta" in pname.lower() else
                                "0.0%" if "%" in fmt else "0.00x")
                        note(ws1, r, 3, desc)
                        note(ws1, r, 4, src)

                    # Prix par modèle
                    r0 = 11 + len(params_rows) + 1
                    sub(ws1, r0, 1, "Prix cibles par modèle")
                    ws1.merge_cells(f"A{r0}:D{r0}")
                    hdr(ws1, r0+1, 1, "Modèle")
                    hdr(ws1, r0+1, 2, "Prix cible (FCFA)")
                    hdr(ws1, r0+1, 3, "Poids dans combiné")
                    hdr(ws1, r0+1, 4, "Potentiel vs cours")
                    WEIGHTS = {"DDM": 0.20, "P/E": 0.30, "P/B": 0.20, "DCF": 0.30}
                    pm = val_result.get("prix_modeles", {})
                    for i, (mod, mprice) in enumerate(pm.items()):
                        r = r0 + 2 + i
                        ws1.cell(r, 1).fill = ALT_FILL if i%2 else PatternFill()
                        lbl(ws1, r, 1, mod)
                        val_cell(ws1, r, 2, round(mprice), fmt='#,##0', bold=True)
                        val_cell(ws1, r, 3, WEIGHTS.get(mod, 0), fmt='0%')
                        pot_m = (mprice - cours_act)/cours_act if (
                            not np.isnan(cours_act) and cours_act > 0) else None
                        c = val_cell(ws1, r, 4,
                                     pot_m if pot_m is not None else "—",
                                     fmt='+0.0%;-0.0%;"-"')
                        if pot_m is not None:
                            c.font = Font(bold=True,
                                          color="006400" if pot_m > 0 else "FF0000",
                                          name="Arial", size=9)

                    # ══════════════════════════════════════════════
                    # FEUILLE 2 — P/E
                    # ══════════════════════════════════════════════
                    ws2 = wb.create_sheet("Valorisation P-E")
                    for col, width in [(1,35),(2,22),(3,25),(4,20)]:
                        ws2.column_dimensions[get_column_letter(col)].width = width

                    hdr(ws2, 1, 1, "MÉTHODE P/E RELATIF")
                    ws2.merge_cells("A1:D1"); ws2.cell(1,1).alignment = CNTR

                    pe_res = mods.get("P/E")
                    note(ws2, 2, 1, "Formule : Prix cible = EPS × P/E_cible")
                    note(ws2, 2, 2, "EPS = Résultat Net / Nombre d'actions")

                    sub(ws2, 4, 1, "Données de base"); ws2.merge_cells("A4:D4")
                    rows_pe = [
                        ("Résultat Net (FCFA)", postes.get("rn","—"), '#,##0',
                         "Source : États financiers"),
                        ("Nb actions", data.get("nb_titres",{}).get(ticker,"—")
                         if data else "—", '#,##0',
                         "Source : Base de données SMF"),
                        ("EPS (FCFA/action)", pe_res.get("eps","—")
                         if pe_res else "—", '#,##0.00',
                         "=Résultat Net / Nb actions"),
                        ("P/E cible (x)", pe_res.get("pe_cible","—")
                         if pe_res else "—", '0.0"x"',
                         f"Méthode : {p.get('pe_method','—')}"),
                        ("Prix cible P/E (FCFA)", pe_res.get("prix_cible","—")
                         if pe_res else "—", '#,##0',
                         "=EPS × P/E_cible"),
                        ("Cours actuel (FCFA)", cours_act
                         if not np.isnan(cours_act) else "—", '#,##0', "Cours de clôture"),
                        ("Potentiel (%)", (pe_res["prix_cible"]-cours_act)/cours_act
                         if pe_res and not np.isnan(cours_act) and cours_act>0 else "—",
                         '+0.0%;-0.0%;"-"', "=(Prix cible - Cours) / Cours"),
                    ]
                    hdr(ws2, 5, 1, "Poste"); hdr(ws2, 5, 2, "Valeur")
                    hdr(ws2, 5, 3, "Formule/Note"); hdr(ws2, 5, 4, "Source")
                    for i, (pname, pval, fmt, src) in enumerate(rows_pe):
                        r = 6 + i
                        ws2.cell(r,1).fill = ALT_FILL if i%2 else PatternFill()
                        lbl(ws2, r, 1, pname)
                        v = round(pval, 4) if isinstance(pval, float) else pval
                        inp(ws2, r, 2, v, fmt=fmt)
                        note(ws2, r, 3, src.split("=")[1] if "=" in src else src)
                        note(ws2, r, 4, src)

                    # ══════════════════════════════════════════════
                    # FEUILLE 3 — P/B
                    # ══════════════════════════════════════════════
                    ws3 = wb.create_sheet("Valorisation P-B")
                    for col, width in [(1,35),(2,22),(3,25),(4,20)]:
                        ws3.column_dimensions[get_column_letter(col)].width = width

                    hdr(ws3, 1, 1, "MÉTHODE P/B (PRICE-TO-BOOK)")
                    ws3.merge_cells("A1:D1"); ws3.cell(1,1).alignment = CNTR

                    pb_res = mods.get("P/B")
                    note(ws3, 2, 1, "Formule : Prix cible = BVPS × P/B_cible")
                    note(ws3, 2, 2, "BVPS = Capitaux propres / Nombre d'actions")

                    sub(ws3, 4, 1, "Données de base"); ws3.merge_cells("A4:D4")
                    rows_pb = [
                        ("Capitaux propres (FCFA)", postes.get("capitaux_propres",
                         postes.get("capitaux_propres_mere","—")), '#,##0',
                         "Source : États financiers"),
                        ("Nb actions", data.get("nb_titres",{}).get(ticker,"—")
                         if data else "—", '#,##0',
                         "Source : Base de données SMF"),
                        ("BVPS (FCFA/action)", pb_res.get("bvps","—")
                         if pb_res else "—", '#,##0.00',
                         "=Capitaux propres / Nb actions"),
                        ("P/B cible (x)", pb_res.get("pb_cible","—")
                         if pb_res else "—", '0.00"x"',
                         f"Méthode : {p.get('pb_method','—')}"),
                        ("Prix cible P/B (FCFA)", pb_res.get("prix_cible","—")
                         if pb_res else "—", '#,##0',
                         "=BVPS × P/B_cible"),
                        ("Cours actuel (FCFA)", cours_act
                         if not np.isnan(cours_act) else "—", '#,##0', "Cours de clôture"),
                        ("Potentiel (%)", (pb_res["prix_cible"]-cours_act)/cours_act
                         if pb_res and not np.isnan(cours_act) and cours_act>0 else "—",
                         '+0.0%;-0.0%;"-"', "=(Prix cible - Cours) / Cours"),
                    ]
                    hdr(ws3, 5, 1, "Poste"); hdr(ws3, 5, 2, "Valeur")
                    hdr(ws3, 5, 3, "Note"); hdr(ws3, 5, 4, "Source")
                    for i, (pname, pval, fmt, src) in enumerate(rows_pb):
                        r = 6 + i
                        ws3.cell(r,1).fill = ALT_FILL if i%2 else PatternFill()
                        lbl(ws3, r, 1, pname)
                        v = round(pval, 4) if isinstance(pval, float) else pval
                        inp(ws3, r, 2, v, fmt=fmt)
                        note(ws3, r, 3, "")
                        note(ws3, r, 4, src)

                    # ══════════════════════════════════════════════
                    # FEUILLE 4 — DCF
                    # ══════════════════════════════════════════════
                    ws4 = wb.create_sheet("Valorisation DCF")
                    for col, width in [(1,35),(2,18),(3,18),(4,18),(5,18),(6,18),(7,18),(8,20)]:
                        ws4.column_dimensions[get_column_letter(col)].width = width

                    hdr(ws4, 1, 1, "DISCOUNTED CASH FLOW (DCF)")
                    ws4.merge_cells("A1:H1"); ws4.cell(1,1).alignment = CNTR

                    dcf_res = mods.get("DCF")
                    note(ws4, 2, 1,
                         "Formule : EV = Σ FCF_t/(1+WACC)^t + TV/(1+WACC)^n  |  "
                         "TV = FCF_n×(1+g_TV)/(WACC-g_TV)  |  "
                         "Prix cible = (EV - Dettes + Cash) / Nb actions")

                    # Hypothèses
                    sub(ws4, 4, 1, "Hypothèses"); ws4.merge_cells("A4:H4")
                    hyp_rows = [
                        ("WACC",         p.get("wacc",0),    "0.0%",
                         "ke×E/(D+E) + kd×(1-t)×D/(D+E)"),
                        ("ke",           p.get("ke",0),      "0.0%",
                         "Rf(6%) + β×Prime(6%)"),
                        ("kd",           p.get("kd",0),      "0.0%",
                         "Intérêts / Dette — source : EF"),
                        ("Beta (β)",     p.get("beta",1),    "0.000",
                         "Régression OLS cours vs marché"),
                        ("g FCF",        p.get("g_fcf",0),   "0.0%",
                         "CAGR Rex/RN — source : EF"),
                        ("g Terminale",  p.get("g_tv",
                         dcf_res.get("g_tv",0.045) if dcf_res else 0.045), "0.0%",
                         "Croissance PIB UEMOA long terme"),
                        ("Horizon",      dcf_res.get("horizon",5)
                         if dcf_res else 5, "0",
                         "Années de projection explicite"),
                        ("FCF0 (FCFA)",  round(dcf_res.get("fcf0_m",0))
                         if dcf_res else "—", '#,##0',
                         "FCF de base : Rex×(1-t)+Amort-CAPEX"),
                    ]
                    hdr(ws4, 5, 1, "Hypothèse"); hdr(ws4, 5, 2, "Valeur")
                    hdr(ws4, 5, 3, "Description"); hdr(ws4, 5, 4, "Source")
                    for i, (hname, hval, fmt, hdesc) in enumerate(hyp_rows):
                        r = 6 + i
                        ws4.cell(r,1).fill = ALT_FILL if i%2 else PatternFill()
                        lbl(ws4, r, 1, hname)
                        v = round(hval, 6) if isinstance(hval, float) else hval
                        inp(ws4, r, 2, v, fmt=fmt)
                        note(ws4, r, 3, hdesc)
                        note(ws4, r, 4, "Calibrage auto")

                    # Projection FCF
                    if dcf_res and dcf_res.get("fcf_table"):
                        r_fcf = 6 + len(hyp_rows) + 2
                        sub(ws4, r_fcf, 1, "Projection des Free Cash Flows")
                        ws4.merge_cells(f"A{r_fcf}:H{r_fcf}")
                        hdr(ws4, r_fcf+1, 1, "Année")
                        hdr(ws4, r_fcf+1, 2, "FCF projeté (FCFA)")
                        hdr(ws4, r_fcf+1, 3, "Facteur actuali.")
                        hdr(ws4, r_fcf+1, 4, "Valeur actuelle (FCFA)")
                        for i, row_fcf in enumerate(dcf_res["fcf_table"]):
                            r = r_fcf + 2 + i
                            ws4.cell(r,1).fill = ALT_FILL if i%2 else PatternFill()
                            val_cell(ws4, r, 1, row_fcf["annee"])
                            val_cell(ws4, r, 2, round(row_fcf["fcf_m"]),
                                     fmt='#,##0')
                            val_cell(ws4, r, 3,
                                     round(1/(1+p.get("wacc",0.12))**(i+1), 6),
                                     fmt='0.000000')
                            val_cell(ws4, r, 4, round(row_fcf["pv_m"]),
                                     fmt='#,##0')

                        # Résumé DCF
                        r_sum = r_fcf + 2 + len(dcf_res["fcf_table"]) + 1
                        sub(ws4, r_sum, 1, "Résumé valorisation DCF")
                        ws4.merge_cells(f"A{r_sum}:D{r_sum}")
                        dcf_sum = [
                            ("PV des FCF projetés (FCFA)",
                             round(dcf_res["pv_fcf_m"]), '#,##0'),
                            ("Valeur terminale PV (FCFA)",
                             round(dcf_res["pv_tv_m"]),  '#,##0'),
                            ("Valeur d'entreprise EV (FCFA)",
                             round(dcf_res["ev_m"]),     '#,##0'),
                            ("Dette financière (FCFA)",
                             round(abs(postes.get("dette_financiere",0))), '#,##0'),
                            ("Prix cible DCF (FCFA)",
                             round(dcf_res["prix_cible"]),  '#,##0'),
                        ]
                        for i, (dname, dval, fmt) in enumerate(dcf_sum):
                            r = r_sum + 1 + i
                            lbl(ws4, r, 1, dname)
                            val_cell(ws4, r, 2, dval, fmt=fmt, bold=True)

                    # ══════════════════════════════════════════════
                    # FEUILLE 5 — DDM (si disponible)
                    # ══════════════════════════════════════════════
                    ddm_res = mods.get("DDM")
                    if ddm_res:
                        ws5 = wb.create_sheet("Valorisation DDM")
                        for col, width in [(1,35),(2,22),(3,28)]:
                            ws5.column_dimensions[get_column_letter(col)].width = width

                        hdr(ws5, 1, 1, "DIVIDEND DISCOUNT MODEL (DDM)")
                        ws5.merge_cells("A1:C1"); ws5.cell(1,1).alignment = CNTR
                        note(ws5, 2, 1,
                             "Formule Gordon-Shapiro : P = D1 / (ke - g)  |  "
                             "D1 = D0 × (1 + g)")

                        sub(ws5, 4, 1, "Paramètres DDM"); ws5.merge_cells("A4:C4")
                        ddm_rows = [
                            ("D0 — Dividende actuel (FCFA/action)",
                             round(ddm_res.get("d0",0)), '#,##0.00',
                             "Dernier dividende versé"),
                            ("g — Croissance dividendes",
                             ddm_res.get("g",0), '0.0%',
                             "CAGR dividendes historiques"),
                            ("D1 — Dividende projeté (FCFA)",
                             round(ddm_res.get("d1",0),2), '#,##0.00',
                             "=D0 × (1 + g)"),
                            ("ke — Taux d'actualisation",
                             ddm_res.get("ke",0), '0.0%',
                             "CAPM : Rf + β × prime"),
                            ("ke - g (spread)",
                             ddm_res.get("ke",0)-ddm_res.get("g",0), '0.0%',
                             "Doit être > 0"),
                            ("Prix cible DDM (FCFA)",
                             round(ddm_res.get("prix_cible",0)), '#,##0',
                             "=D1 / (ke - g)"),
                        ]
                        hdr(ws5, 5, 1, "Paramètre")
                        hdr(ws5, 5, 2, "Valeur")
                        hdr(ws5, 5, 3, "Note")
                        for i, (pname, pval, fmt, src) in enumerate(ddm_rows):
                            r = 6 + i
                            ws5.cell(r,1).fill = ALT_FILL if i%2 else PatternFill()
                            lbl(ws5, r, 1, pname)
                            v = round(pval, 6) if isinstance(pval, float) else pval
                            inp(ws5, r, 2, v, fmt=fmt)
                            note(ws5, r, 3, src)

                    # Sauvegarder
                    buf_xl = io.BytesIO()
                    wb.save(buf_xl)
                    buf_xl.seek(0)
                    return buf_xl.getvalue()

                # Génération à la demande — un seul fichier à la fois (performance)
                dl_c1, dl_c2 = st.columns([2, 1])
                with dl_c1:
                    ticker_to_dl = st.selectbox(
                        "Choisir un titre à télécharger",
                        options=[r["Ticker"] for r in rows],
                        key="ticker_excel_dl",
                        label_visibility="collapsed",
                    )
                with dl_c2:
                    if ticker_to_dl and ticker_to_dl in val:
                        # Clé de cache : ticker + prix cible (change si recalcul)
                        pc_key = val[ticker_to_dl].get("prix_cible", 0)
                        cache_key = f"{ticker_to_dl}_{pc_key:.2f}" if pc_key else ticker_to_dl
                        try:
                            xl_bytes = build_valuation_excel_cached(
                                ticker_to_dl, val[ticker_to_dl],
                                fin_data, cours_df, cache_key
                            )
                            st.download_button(
                                label=f"⬇️ Excel {ticker_to_dl}",
                                data=xl_bytes,
                                file_name=f"Valorisation_{ticker_to_dl}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument"
                                     ".spreadsheetml.sheet",
                                key=f"dl_btn_{ticker_to_dl}",
                                use_container_width=True,
                            )
                        except Exception as e:
                            st.error(f"Erreur : {e}")

                # Graphique potentiels
                st.markdown("**Potentiels de hausse / baisse par titre**")
                df_chart = df_res[df_res["_pot_num"] != 0].copy()
                colors_bar = ["#10b981" if v > 0 else "#ef4444" for v in df_chart["_pot_num"]]
                fig_pot = go.Figure(go.Bar(
                    x=df_chart["Ticker"],
                    y=df_chart["_pot_num"],
                    marker=dict(color=colors_bar, line=dict(width=0)),
                    text=[f"{v:+.1f}%" for v in df_chart["_pot_num"]],
                    textposition="outside",
                    textfont=dict(size=9, color="#94a3b8"),
                ))
                fig_pot.add_hline(y=0, line_width=1, line_color="#475569")
                fig_pot.add_hline(y=10,  line_dash="dash", line_color="#10b981", line_width=0.8)
                fig_pot.add_hline(y=-10, line_dash="dash", line_color="#ef4444", line_width=0.8)
                fig_pot.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8", family="JetBrains Mono"), height=380,
                    margin=dict(l=10, r=10, t=20, b=50),
                    xaxis=dict(gridcolor="#1e2d45", tickangle=-45),
                    yaxis=dict(gridcolor="#1e2d45", ticksuffix="%", title="Potentiel (%)"),
                )
                st.plotly_chart(fig_pot, width="stretch")

                # Répartition des signaux par secteur
                if SECTOR_MAP:
                    st.markdown("**📊 Potentiel moyen par secteur**")
                    df_sect = df_res[df_res["_pot_num"] != 0].copy()
                    df_sect["Secteur_map"] = df_sect["Ticker"].apply(
                        lambda t: SECTOR_MAP.get(t.upper(), "Autre")
                    )
                    sect_avg = df_sect.groupby("Secteur_map")["_pot_num"].mean().sort_values(ascending=False)
                    fig_sect = go.Figure(go.Bar(
                        x=sect_avg.index,
                        y=sect_avg.values,
                        marker=dict(
                            color=["#10b981" if v > 0 else "#ef4444" for v in sect_avg.values],
                            line=dict(width=0)
                        ),
                        text=[f"{v:+.1f}%" for v in sect_avg.values],
                        textposition="outside",
                        textfont=dict(size=10, color="#94a3b8"),
                    ))
                    fig_sect.add_hline(y=0, line_width=1, line_color="#475569")
                    fig_sect.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(color="#94a3b8", family="JetBrains Mono"), height=320,
                        margin=dict(l=10, r=10, t=20, b=80),
                        xaxis=dict(gridcolor="#1e2d45", tickangle=-30),
                        yaxis=dict(gridcolor="#1e2d45", ticksuffix="%"),
                    )
                    st.plotly_chart(fig_sect, width="stretch")

                # Tableau transparence des paramètres calibrés
                st.markdown("---")
                st.markdown("**🔬 Transparence — Paramètres calibrés automatiquement**")
                st.caption("Tous dérivés des données réelles · aucune saisie manuelle")
                param_rows = []
                for ticker_p, v_p in val.items():
                    p = v_p.get("params")
                    if not p:
                        continue
                    param_rows.append({
                        "Ticker":    ticker_p,
                        "Secteur":   SECTOR_MAP.get(ticker_p.upper(), p.get("secteur","—")),
                        "Type":      "Banque" if p.get("is_bank") else "Société",
                        "β":         f"{p.get('beta',1):.2f}",
                        "ke":        f"{p.get('ke',0)*100:.1f}%",
                        "kd":        f"{p.get('kd',0)*100:.1f}%",
                        "WACC":      f"{p.get('wacc',0)*100:.1f}%",
                        "g_FCF":     f"{p.get('g_fcf',0)*100:.1f}%",
                        "g_div":     f"{p.get('g_div',0)*100:.1f}%",
                        "P/E cible": f"{p.get('pe_target',0):.1f}x",
                        "P/E méth.": p.get("pe_method","—"),
                        "P/B cible": f"{p.get('pb_target',0):.2f}x",
                        "Taux IS":   f"{p.get('tax_rate',0)*100:.1f}%",
                        "Année":     str(p.get("annee_ref","—")),
                    })
                if param_rows:
                    st.dataframe(pd.DataFrame(param_rows), width="stretch", hide_index=True)

                # Détail par titre
                st.markdown("**🔍 Détail par titre — DCF**")
                ticker_detail = st.selectbox("Choisir un titre", [r["Ticker"] for r in rows])
                if ticker_detail and ticker_detail in val:
                    dcf_res = val[ticker_detail]["modeles"].get("DCF")
                    if dcf_res and dcf_res.get("fcf_table"):
                        st.markdown(f"**{ticker_detail} — Projection DCF** · WACC={dcf_res['wacc']:.2%} · g_FCF={dcf_res['g_fcf']:.2%} · g_TV={dcf_res['g_tv']:.2%}")
                        dcf_df = pd.DataFrame(dcf_res["fcf_table"])
                        dcf_df.columns = ["Année","FCF projeté (M FCFA)","PV (M FCFA)"]
                        dcf_df = dcf_df.round(0)

                        col_dcf1, col_dcf2 = st.columns(2)
                        with col_dcf1:
                            st.dataframe(dcf_df, width="stretch", hide_index=True)
                        with col_dcf2:
                            kd1, kd2, kd3 = st.columns(3)
                            kd1.metric("PV FCF (M FCFA)", f"{dcf_res['pv_fcf_m']:,.0f}")
                            kd2.metric("PV TV (M FCFA)", f"{dcf_res['pv_tv_m']:,.0f}")
                            kd3.metric("EV (M FCFA)", f"{dcf_res['ev_m']:,.0f}")
                            st.metric("Prix cible DCF (FCFA)", f"{dcf_res['prix_cible']:,.0f}")
                    else:
                        st.info("DCF non disponible pour ce titre.")

                # Export
                buf = io.BytesIO()
                df_res[display_cols].to_excel(buf, index=False)
                st.download_button("⬇️ Exporter valorisations (Excel)", buf.getvalue(),
                                   "valorisation_BRVM.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ BACKTESTING ════════════════════════════════════════════════
with t9:
    st.markdown("<span class='pill'>Backtesting</span>"
                "<p class='sh'>Performance historique de la stratégie multifactorielle</p>",
                unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    Simulation historique · Rebalancement trimestriel · Poids rang MF recalculés à chaque période<br>
    Benchmarks : BRVM Composite (équipondéré) · Portefeuille Charia compatible
    </div>""", unsafe_allow_html=True)

    if not data:
        st.info("👈 Chargez d'abord le fichier de données (sidebar).")
    else:
        c_min = data["cours"].index.min().date()
        c_max = data["cours"].index.max().date()

        # ── Paramètres ─────────────────────────────────────────
        st.markdown("**⚙️ Paramètres du backtesting**")
        bp1, bp2, bp3 = st.columns(3)
        with bp1:
            bt_start = st.date_input("Date de début",
                value=max(c_min, min(c_max, pd.Timestamp("2018-01-01").date())),
                min_value=c_min, max_value=max(c_max, datetime.date.today()), key="bt_start")
        with bp2:
            bt_end = st.date_input("Date de fin",
                value=c_max, min_value=c_min, max_value=max(c_max, datetime.date.today()), key="bt_end")
        with bp3:
            st.markdown("<br>", unsafe_allow_html=True)
            n_days_bt = (bt_end - bt_start).days
            n_years_bt = n_days_bt / 365
            st.success(f"✅ **{n_days_bt}** jours · **{n_years_bt:.1f}** ans")

        betas_bt = st.session_state.get("betas",
            {"Value":0.2,"Momentum":0.2,"Volatilité":0.2,"Dividende":0.2,"Liquidité":0.2})

        if st.button("🚀 Lancer le backtesting", type="primary"):
            if bt_start >= bt_end:
                st.error("Date de début ≥ Date de fin.")
            else:
                with st.spinner("Simulation en cours — calcul des performances historiques..."):
                    bt_result = run_backtest(
                        cours_df=data["cours"],
                        factor_results=st.session_state.factor_results,
                        betas=betas_bt,
                        start_date=bt_start,
                        end_date=bt_end,
                        rebal_freq="QS",   # trimestriel
                        charia_results=st.session_state.charia_results
                                       if CHARIA_AVAILABLE else None,
                        data=data,
                    )
                if bt_result is None:
                    st.error("Données insuffisantes pour la période sélectionnée.")
                else:
                    st.session_state.bt_result = bt_result
                    st.success("✅ Backtesting terminé")

        # ── Résultats ──────────────────────────────────────────
        if "bt_result" in st.session_state and st.session_state.bt_result:
            bt  = st.session_state.bt_result
            perf= bt["perf"]
            mets= bt["metrics"]

            st.markdown("---")

            # KPIs synthèse
            st.markdown("**📊 Métriques de performance**")
            PORTF_COLS = list(perf.columns)
            PORTF_CLRS = {
                "Portefeuille MF":     "#3b82f6",
                "Portefeuille Charia": "#10b981",
                "BRVM Composite":      "#f59e0b",
            }

            # Ligne de métriques
            m_rows = ["Rendement total","CAGR","Volatilité annuelle",
                      "Ratio de Sharpe","Max Drawdown","Alpha annualisé","Beta"]
            tbl_data = {"Métrique": m_rows}
            for col in PORTF_COLS:
                tbl_data[col] = [mets[col].get(r, "—") for r in m_rows]
            st.dataframe(pd.DataFrame(tbl_data), width="stretch", hide_index=True)

            st.markdown("---")

            # Graphique performance cumulée
            st.markdown("**📈 Performance cumulée (base 100)**")
            fig_perf = go.Figure()
            for col in PORTF_COLS:
                fig_perf.add_trace(go.Scatter(
                    x=perf.index, y=perf[col],
                    name=col,
                    line=dict(color=PORTF_CLRS.get(col,"#94a3b8"), width=2),
                    hovertemplate=f"<b>{col}</b><br>%{{x|%d/%m/%Y}}<br>%{{y:.2f}}<extra></extra>"
                ))
            # Marquer les dates de rebalancement
            for rd in bt["rebal_dates"][::4]:  # 1 sur 4 pour ne pas surcharger
                fig_perf.add_vline(x=rd, line_width=0.5,
                                   line_dash="dot", line_color="#475569")
            fig_perf.add_annotation(x=bt["rebal_dates"][0],
                                    y=perf.iloc[0].max()*0.95,
                                    text="▲ rebalancement",
                                    font=dict(size=9, color="#64748b"),
                                    showarrow=False)
            fig_perf.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8", family="JetBrains Mono"), height=420,
                legend=dict(orientation="h", y=1.05, bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=10, r=10, t=40, b=40),
                xaxis=dict(gridcolor="#1e2d45"),
                yaxis=dict(gridcolor="#1e2d45", title="Valeur (base 100)"),
                hovermode="x unified",
            )
            st.plotly_chart(fig_perf, width="stretch")

            # Drawdown
            st.markdown("**📉 Drawdown**")
            # Conversion hex → rgba pour fillcolor (Plotly n'accepte pas hex 8 chars)
            def hex_to_rgba(hex_color, alpha=0.15):
                h = hex_color.lstrip("#")
                r, g, b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
                return f"rgba({r},{g},{b},{alpha})"

            fig_dd = go.Figure()
            for col in PORTF_COLS:
                s   = perf[col]
                dd  = (s - s.cummax()) / s.cummax() * 100
                clr = PORTF_CLRS.get(col, "#94a3b8")
                fig_dd.add_trace(go.Scatter(
                    x=dd.index, y=dd,
                    name=col, fill="tozeroy",
                    line=dict(color=clr, width=1),
                    fillcolor=hex_to_rgba(clr, 0.15),
                    hovertemplate=f"<b>{col}</b><br>%{{y:.2f}}%<extra></extra>"
                ))
            fig_dd.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8", family="JetBrains Mono"), height=300,
                legend=dict(orientation="h", y=1.05, bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=10, r=10, t=40, b=40),
                xaxis=dict(gridcolor="#1e2d45"),
                yaxis=dict(gridcolor="#1e2d45", title="Drawdown (%)",
                           ticksuffix="%"),
                hovermode="x unified",
            )
            st.plotly_chart(fig_dd, width="stretch")

            # Rendements annuels
            st.markdown("**📅 Rendements annuels**")
            annual = {}
            for col in PORTF_COLS:
                yr_ret = perf[col].resample("YE").last().pct_change().dropna()
                annual[col] = yr_ret * 100
            annual_df = pd.DataFrame(annual)
            annual_df.index = annual_df.index.year

            fig_ann = go.Figure()
            for col in PORTF_COLS:
                if col in annual_df:
                    fig_ann.add_trace(go.Bar(
                        name=col,
                        x=annual_df.index.astype(str),
                        y=annual_df[col].round(2),
                        marker_color=PORTF_CLRS.get(col,"#94a3b8"),
                        opacity=0.85,
                        text=[f"{v:.1f}%" for v in annual_df[col]],
                        textposition="outside",
                        textfont=dict(size=9),
                    ))
            fig_ann.add_hline(y=0, line_width=1, line_color="#475569")
            fig_ann.update_layout(
                barmode="group",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8", family="JetBrains Mono"), height=340,
                legend=dict(orientation="h", y=1.05, bgcolor="rgba(0,0,0,0)"),
                margin=dict(l=10, r=10, t=40, b=40),
                xaxis=dict(gridcolor="#1e2d45", title="Année"),
                yaxis=dict(gridcolor="#1e2d45", ticksuffix="%"),
            )
            st.plotly_chart(fig_ann, width="stretch")

            # Rendements mensuels — heatmap Portefeuille MF
            st.markdown("**🗓️ Rendements mensuels — Portefeuille MF (%)**")
            monthly = perf["Portefeuille MF"].resample("ME").last().pct_change().dropna() * 100
            monthly_df = monthly.groupby([monthly.index.year, monthly.index.month]).first().unstack()
            monthly_df.columns = ["Jan","Fév","Mar","Avr","Mai","Jun",
                                   "Jul","Aoû","Sep","Oct","Nov","Déc"][:len(monthly_df.columns)]
            fig_hm = go.Figure(go.Heatmap(
                z=monthly_df.values,
                x=monthly_df.columns.tolist(),
                y=monthly_df.index.tolist(),
                colorscale=[[0,"#ef4444"],[0.5,"#1e2d45"],[1,"#10b981"]],
                zmid=0,
                text=[[f"{v:.1f}%" if not np.isnan(v) else ""
                       for v in row] for row in monthly_df.values],
                texttemplate="%{text}",
                textfont=dict(size=9),
                hovertemplate="<b>%{y} %{x}</b><br>%{z:.2f}%<extra></extra>",
            ))
            fig_hm.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8", family="JetBrains Mono"),
                height=max(250, len(monthly_df)*35+60),
                margin=dict(l=60, r=10, t=20, b=40),
            )
            st.plotly_chart(fig_hm, width="stretch")

            # Export
            buf_bt = io.BytesIO()
            with pd.ExcelWriter(buf_bt, engine="openpyxl") as writer:
                perf.to_excel(writer, sheet_name="Performance")
                pd.DataFrame(tbl_data).to_excel(writer, sheet_name="Métriques", index=False)
                annual_df.to_excel(writer, sheet_name="Rendements annuels")
            st.download_button("⬇️ Exporter le backtesting (Excel)",
                               buf_bt.getvalue(), "backtesting_BRVM.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ══ CORRÉLATIONS & DIVERSIFICATION ════════════════════════════
with t10:
    st.markdown(
        "<span class='pill'>Analyse</span>"
        "<p class='sh'>Corrélations & Diversification Sectorielle</p>",
        unsafe_allow_html=True)
    st.markdown("""<div class='fbox'>
    Matrice de corrélations · Clustering des titres · Concentration sectorielle · Performance par secteur<br>
    Basé sur l'historique des cours et la classification BRVM officielle (sectors.json)
    </div>""", unsafe_allow_html=True)

    if not data:
        st.info("👈 Chargez d'abord le fichier de données (sidebar).")
    else:
        c_min = data["cours"].index.min().date()
        c_max = data["cours"].index.max().date()

        # ── Paramètres ─────────────────────────────────────────
        st.markdown("**⚙️ Paramètres**")
        cp1, cp2, cp3 = st.columns(3)
        with cp1:
            corr_window = st.slider("Fenêtre corrélations (jours)", 60, 504, 252, 21,
                                    key="corr_win")
        with cp2:
            max_corr_cl = st.slider("Seuil clustering (corrélation max)", 0.3, 0.95, 0.70, 0.05,
                                    key="corr_cl")
        with cp3:
            perf_start = st.date_input("Début perf. sectorielle",
                value=max(c_min, min(c_max, pd.Timestamp("2020-01-01").date())),
                min_value=c_min, max_value=max(c_max, datetime.date.today()), key="perf_start")

        if st.button("🔗 Calculer les analyses", type="primary"):
            with st.spinner("Calcul des corrélations et diversification..."):
                corr_mat   = compute_correlation_matrix(data["cours"], corr_window)
                clusters, ordered_tickers, Z = cluster_tickers(corr_mat, max_corr_cl)
                sect_perf  = sector_performance(
                    data["cours"], SECTOR_MAP, perf_start, c_max)
            st.session_state.corr_mat  = corr_mat
            st.session_state.clusters  = clusters
            st.session_state.ord_tick  = ordered_tickers
            st.session_state.sect_perf = sect_perf
            st.success(f"✅ Matrice {corr_mat.shape[0]}×{corr_mat.shape[0]} · "
                       f"{len(set(clusters.values()))} clusters · "
                       f"{len(sect_perf.columns) if not sect_perf.empty else 0} secteurs")

        if "corr_mat" in st.session_state and st.session_state.corr_mat is not None:
            corr_mat = st.session_state.corr_mat
            clusters = st.session_state.clusters
            ordered  = st.session_state.ord_tick
            sect_perf= st.session_state.sect_perf

            # ── 1. Heatmap corrélations ────────────────────────
            st.markdown("---")
            st.markdown("**🌡️ Matrice de corrélations des rendements**")
            st.caption(f"Fenêtre : {corr_window} jours · "
                       f"Rouge = corrélation positive · Bleu = corrélation négative")

            # Réordonner par cluster pour regrouper les titres similaires
            ord_valid = [t for t in ordered if t in corr_mat.index]
            corr_ord  = corr_mat.loc[ord_valid, ord_valid]

            # Ajouter labels de cluster
            cl_labels = [f"{t}\n(C{clusters.get(t,'?')})" for t in ord_valid]

            fig_hm = go.Figure(go.Heatmap(
                z=corr_ord.values,
                x=cl_labels, y=cl_labels,
                colorscale=[[0,"#3b82f6"],[0.5,"#1e2d45"],[1,"#ef4444"]],
                zmid=0, zmin=-1, zmax=1,
                text=[[f"{v:.2f}" for v in row] for row in corr_ord.values],
                texttemplate="%{text}",
                textfont=dict(size=7),
                hovertemplate="<b>%{y} × %{x}</b><br>Corrélation : %{z:.3f}<extra></extra>",
                colorbar=dict(title="ρ", thickness=12),
            ))
            n_t = len(ord_valid)
            fig_hm.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#94a3b8", family="JetBrains Mono", size=8),
                height=max(500, n_t * 20 + 100),
                margin=dict(l=100, r=20, t=40, b=100),
                xaxis=dict(tickangle=-45),
            )
            st.plotly_chart(fig_hm, width="stretch")

            # ── 2. Clustering ──────────────────────────────────
            st.markdown("---")
            st.markdown("**🔵 Clusters de titres corrélés**")
            st.caption("Titres dans le même cluster évoluent de manière similaire → "
                       "en conserver un seul améliore la diversification")

            n_clusters = max(clusters.values()) if clusters else 0
            CLUST_CLRS = ["#3b82f6","#10b981","#f59e0b","#ef4444","#8b5cf6",
                          "#06b6d4","#ec4899","#84cc16","#f97316","#a855f7"]

            # Grille de clusters
            cluster_groups = {}
            for t, c in clusters.items():
                cluster_groups.setdefault(c, []).append(t)

            cols_cl = st.columns(min(4, n_clusters))
            for i, (cl_id, members) in enumerate(sorted(cluster_groups.items())):
                clr = CLUST_CLRS[i % len(CLUST_CLRS)]
                with cols_cl[i % len(cols_cl)]:
                    members_str = " · ".join(sorted(members))
                    # Corrélation intra-cluster
                    if len(members) > 1:
                        sub_corr = corr_mat.loc[
                            [m for m in members if m in corr_mat.index],
                            [m for m in members if m in corr_mat.index]
                        ]
                        # Moyenne des corrélations hors diagonale
                        mask = ~np.eye(len(sub_corr), dtype=bool)
                        intra = sub_corr.values[mask].mean() if mask.any() else 1.0
                        intra_str = f"ρ intra = {intra:.2f}"
                    else:
                        intra_str = "titre unique"
                    st.markdown(
                        f"<div style='background:#161d2e;border-left:3px solid {clr};"
                        f"border-radius:6px;padding:10px;margin-bottom:8px;'>"
                        f"<div style='color:{clr};font-weight:700;font-size:11px;'>"
                        f"Cluster {cl_id} · {len(members)} titre(s) · {intra_str}</div>"
                        f"<div style='color:#94a3b8;font-size:10px;margin-top:4px;'>"
                        f"{members_str}</div></div>",
                        unsafe_allow_html=True
                    )

            # ── 3. Concentration sectorielle ───────────────────
            st.markdown("---")
            st.markdown("**🏭 Concentration sectorielle**")

            # Portefeuille actuel si disponible
            pw_current = st.session_state.pw
            if pw_current is not None and not pw_current.empty:
                sect_pf = sector_allocation(pw_current, SECTOR_MAP,
                                            data["tickers"])
                sect_bm = benchmark_sector_weights(SECTOR_MAP, data["cours"])

                # Aligner les secteurs
                all_sects = sorted(set(sect_pf.index) | set(sect_bm.index))
                pf_vals = [sect_pf.get(s, 0) for s in all_sects]
                bm_vals = [sect_bm.get(s, 0) for s in all_sects]

                fig_sect = go.Figure()
                fig_sect.add_trace(go.Bar(
                    name="Portefeuille MF", x=all_sects, y=[v*100 for v in pf_vals],
                    marker_color="#3b82f6", opacity=0.85,
                    text=[f"{v*100:.1f}%" for v in pf_vals],
                    textposition="outside", textfont=dict(size=9)
                ))
                fig_sect.add_trace(go.Bar(
                    name="BRVM Composite", x=all_sects, y=[v*100 for v in bm_vals],
                    marker_color="#f59e0b", opacity=0.6,
                    text=[f"{v*100:.1f}%" for v in bm_vals],
                    textposition="outside", textfont=dict(size=9)
                ))
                fig_sect.update_layout(
                    barmode="group",
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8", family="JetBrains Mono"), height=380,
                    legend=dict(orientation="h", y=1.08, bgcolor="rgba(0,0,0,0)"),
                    margin=dict(l=10, r=10, t=50, b=80),
                    xaxis=dict(gridcolor="#1e2d45", tickangle=-30),
                    yaxis=dict(gridcolor="#1e2d45", ticksuffix="%",
                               title="Poids (%)"),
                )
                st.plotly_chart(fig_sect, width="stretch")

                # Tableau écarts surpoids/sous-poids
                st.markdown("**Écarts vs benchmark (surpoids / sous-poids)**")
                ecarts = []
                for s in all_sects:
                    pf_w = sect_pf.get(s, 0) * 100
                    bm_w = sect_bm.get(s, 0) * 100
                    diff = pf_w - bm_w
                    ecarts.append({
                        "Secteur":  s,
                        "Ptf MF (%)":  f"{pf_w:.1f}%",
                        "Benchmark (%)": f"{bm_w:.1f}%",
                        "Écart":   f"{'▲' if diff>0.5 else '▼' if diff<-0.5 else '≈'} "
                                   f"{diff:+.1f}%",
                        "_diff": diff,
                    })
                df_ecarts = pd.DataFrame(ecarts).sort_values("_diff",ascending=False)
                st.dataframe(df_ecarts.drop("_diff",axis=1),
                             width="stretch", hide_index=True)
            else:
                st.info("ℹ️ Construisez d'abord un portefeuille (onglet 📂) "
                        "pour voir la concentration sectorielle.")

            # ── 4. Performance sectorielle ─────────────────────
            st.markdown("---")
            st.markdown("**📈 Performance sectorielle historique**")

            if not sect_perf.empty:
                fig_sp = go.Figure()
                SECT_CLRS = ["#3b82f6","#10b981","#f59e0b","#ef4444",
                             "#8b5cf6","#06b6d4","#ec4899"]
                for i, col in enumerate(sect_perf.columns):
                    final_ret = (sect_perf[col].iloc[-1] - 1) * 100
                    fig_sp.add_trace(go.Scatter(
                        x=sect_perf.index, y=sect_perf[col] * 100 - 100,
                        name=f"{col} ({final_ret:+.1f}%)",
                        line=dict(color=SECT_CLRS[i % len(SECT_CLRS)], width=2),
                        hovertemplate=f"<b>{col}</b><br>"
                                      f"%{{x|%d/%m/%Y}}<br>"
                                      f"Perf : %{{y:+.2f}}%<extra></extra>"
                    ))
                fig_sp.add_hline(y=0, line_width=1, line_color="#475569")
                fig_sp.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#94a3b8", family="JetBrains Mono"), height=420,
                    legend=dict(orientation="h", y=1.05, bgcolor="rgba(0,0,0,0)",
                                font=dict(size=10)),
                    margin=dict(l=10, r=10, t=50, b=40),
                    xaxis=dict(gridcolor="#1e2d45"),
                    yaxis=dict(gridcolor="#1e2d45", ticksuffix="%",
                               title="Performance cumulée (%)"),
                    hovermode="x unified",
                )
                st.plotly_chart(fig_sp, width="stretch")

                # Tableau perf annuelle par secteur
                st.markdown("**Rendements annuels par secteur (%)**")
                ann_sect = {}
                for col in sect_perf.columns:
                    yr_ret = sect_perf[col].resample("YE").last().pct_change().dropna() * 100
                    ann_sect[col] = yr_ret
                ann_df = pd.DataFrame(ann_sect)
                ann_df.index = ann_df.index.year
                st.dataframe(ann_df.round(2).style.format("{:+.2f}%"),
                             width="stretch")

                # Export
                buf_cd = io.BytesIO()
                with pd.ExcelWriter(buf_cd, engine="openpyxl") as writer:
                    corr_mat.round(4).to_excel(writer, sheet_name="Corrélations")
                    sect_perf.to_excel(writer, sheet_name="Perf sectorielle")
                    if pw_current is not None:
                        df_ecarts.drop("_diff",axis=1).to_excel(
                            writer, sheet_name="Concentration sectorielle", index=False)
                st.download_button(
                    "⬇️ Exporter analyses (Excel)", buf_cd.getvalue(),
                    "correlations_diversification_BRVM.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("Données insuffisantes pour calculer la performance sectorielle.")

# ══ DONNÉES ════════════════════════════════════════════════════
with t11:
    st.markdown("<span class='pill'>Sources</span><p class='sh'>Aperçu des données chargées</p>", unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    c1.metric("Observations cours", len(data["cours"]))
    c2.metric("Première date", str(data["cours"].index[0].date()))
    c3.metric("Dernière date",  str(data["cours"].index[-1].date()))
    st.markdown(f"**{len(data['tickers'])} titres :** {', '.join(data['tickers'])}")
    st.markdown("---")

    ti1, ti2, ti3 = st.tabs(["📈 Cours historiques", "💸 Dividendes", "🏢 Résultats nets"])
    with ti1:
        sel = st.multiselect("Titres", data["tickers"], default=data["tickers"][:6])
        if sel:
            fig_c = go.Figure()
            for t in sel:
                s = data["cours"][t].dropna()
                fig_c.add_trace(go.Scatter(x=s.index, y=s, name=t, mode="lines", line=dict(width=1.5)))
            fig_c.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                                 font=dict(color="#94a3b8",family="JetBrains Mono"), height=400,
                                 legend=dict(bgcolor="rgba(0,0,0,0)"),
                                 xaxis=dict(gridcolor="#1e2d45"), yaxis=dict(gridcolor="#1e2d45"),
                                 hovermode="x unified")
            st.plotly_chart(fig_c, width="stretch")
    with ti2:
        st.dataframe(data["dividendes"].set_index("Date"), width="stretch")
    with ti3:
        rn = data["resultat_net"]
        rows2 = [{"Ticker":t,"Année":y,"Résultat net (MFCFA)":v/1e6}
                 for t,yrs in rn.items() for y,v in yrs.items()]
        if rows2:
            df_rn = pd.DataFrame(rows2).pivot(index="Ticker",columns="Année",values="Résultat net (MFCFA)")
            st.dataframe(df_rn.style.format("{:,.0f}"), width="stretch")
