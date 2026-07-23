"""
charia_screening.py — Screening Chariatique BRVM
4 standards : DJIM · FTSE · S&P · Malaysia
Compatibilité : ≥ 3 standards sur 4 passés

Les ratios sont calculés depuis le fichier de screening (Feuil1)
ou recalculés automatiquement depuis les états financiers chargés.

Exclusions sectorielles :
  - Banques conventionnelles : incompatibles structurellement (modèle Riba)
  - Revenus illicites : revenus_financiers / CA > SEUIL_ILLICITE (5%)
"""
import pandas as pd
import numpy as np

MIN_STANDARDS_PASS = 3
SEUIL_ILLICITE     = 0.05   # 5% — seuil de tolérance AAOIFI

# Tickers BRVM du secteur bancaire conventionnel — exclus automatiquement
BANK_TICKERS = {
    "SGBC","SIBC","NSBC","ECOC","BICC","BOAB","BOAS",
    "BOABF","BOAM","BOAC","BOAN","CBIBF","ETI","ETIT",
    "ORGT","SAFC","BICB",
}

# Tickers exclus pour activités sectorielles illicites
ILLICIT_SECTOR_TICKERS = {
    "STBC": "Industrie du tabac",
    "LNBB": "Jeux de hasard / loterie",
    "SLBC": "Production / distribution de boissons alcoolisées",
}

# Seuils par standard (définitions exactes)
SEUILS = {
    "FTSE":   {"RE": 0.33, "RC_actif": 0.50, "RL": 0.33},   # base ACTIF TOTAL
    "DJIM":   {"RE_cap24": 0.33, "RC_cap24": 0.33, "RL_cap24": 0.33},  # base CAP MOY 24 MOIS
    "S&P":    {"RE_cap36": 0.33, "RC_cap36": 0.49, "RL_cap36": 0.33}, # base CAP MOY 36 MOIS
    "Malaysia": {"RE": 0.33, "RC_cash": 0.33},  # base ACTIF TOTAL — RC=Cash/Actif
}


def _r(n, d):
    """Ratio sécurisé."""
    try:
        n, d = float(n), float(d)
        if np.isnan(n) or np.isnan(d) or d == 0:
            return None
        return n / d
    except Exception:
        return None


def _chk(v, seuil):
    """True si v < seuil (None → True : donnée manquante ne pénalise pas)."""
    return v is None or v < seuil


def _screen_ratios(re, rc_a, rl, re_c24, rc_c24, rl_c24,
                   re_c36, rc_c36, rl_c36, halal):
    """
    Évalue les 4 standards et retourne (résultats dict, n_pass).

    FTSE     : base ACTIF TOTAL      → RE<33% · RC<50% · RL<33%
    DJIM     : base CAP MOY 24 MOIS  → RE<33% · RC<33% · RL<33%
    S&P      : base CAP MOY 36 MOIS  → RE<33% · RC<49% · RL<33%
    Malaysia : base ACTIF TOTAL      → RE<33% · RC(=Cash)<33%
    """
    ftse     = halal and _chk(re,0.33)     and _chk(rc_a,0.50)   and _chk(rl,0.33)
    djim     = halal and _chk(re_c24,0.33) and _chk(rc_c24,0.33) and _chk(rl_c24,0.33)
    sp       = halal and _chk(re_c36,0.33) and _chk(rc_c36,0.49) and _chk(rl_c36,0.33)
    malaysia = halal and _chk(re,0.33)     and _chk(rl,0.33)

    n_pass = sum([ftse, djim, sp, malaysia])
    return {
        "FTSE":     {"pass": ftse},
        "DJIM":     {"pass": djim},
        "S&P":      {"pass": sp},
        "Malaysia": {"pass": malaysia},
    }, n_pass


def parse_screening_file(filepath):
    """
    Parse le fichier Excel de screening — feuille 'TEST SCREENING UMOA'.

    Colonnes lues (format officiel CGF) :
      AA : Symbole (ticker)
      J,K,L   : FTSE     — RE, RC, RL  (base ACTIF TOTAL)
      N,O,P   : DJIM     — RE, RC, RL  (base CAP MOY 24 MOIS)
      R,S,T   : S&P      — RE, RC, RL  (base CAP MOY 36 MOIS)
      V,W     : Malaysia — RE, RC(cash) (base ACTIF TOTAL)
      Y       : RR = Revenus portant intérêt / CA total (filtre préalable ≤5%)

    Seuils réels appliqués (formatage conditionnel du fichier) :
      FTSE     : RE>33% · RC>50% · RL>48%  → fail
      DJIM     : RE>33% · RC>33% · RL>33%  → fail
      S&P      : RE>33% · RC>49% · RL>33%  → fail
      Malaysia : RE>33% · RC>48%           → fail

    Un titre doit d'abord avoir Y ≤ 5%, puis qualifier sur ≥ 3 standards / 4.
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string

    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "TEST SCREENING UMOA" not in wb.sheetnames:
        return {}
    ws = wb["TEST SCREENING UMOA"]

    COLS = {c: column_index_from_string(c)
            for c in ["AA","J","K","L","N","O","P","R","S","T","V","W","Y"]}

    def cell_num(row, col_letter):
        v = ws.cell(row, COLS[col_letter]).value
        if v is None or isinstance(v, str):
            return None  # #DIV/0! ou vide
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    results = {}
    for row in range(1, ws.max_row + 1):
        ticker = ws.cell(row, COLS["AA"]).value
        if not ticker or not isinstance(ticker, str) or not ticker.strip():
            continue
        ticker = ticker.strip().upper()
        if not (2 <= len(ticker) <= 6 and ticker.isalnum()):
            continue

        j, k, l = cell_num(row,"J"), cell_num(row,"K"), cell_num(row,"L")
        n, o, p = cell_num(row,"N"), cell_num(row,"O"), cell_num(row,"P")
        r, s, t = cell_num(row,"R"), cell_num(row,"S"), cell_num(row,"T")
        v, w    = cell_num(row,"V"), cell_num(row,"W")
        y       = cell_num(row,"Y")

        # Si toutes les valeurs sont None → ligne vide/erreur → ignorer
        if all(x is None for x in [j,k,l,n,o,p,r,s,t,v,w]):
            continue

        # Si toutes les valeurs sont à 0 → absence réelle de données
        # (pas de dette/créances/cash renseignés) → non évaluable
        all_vals = [x for x in [j,k,l,n,o,p,r,s,t,v,w] if x is not None]
        if all_vals and all(x == 0 for x in all_vals):
            results[ticker] = {
                "compatible":    False,
                "n_standards":   0,
                "halal_sector":  None,
                "excluded":      False,
                "donnees_manquantes": True,
                "standards": {s: {"pass": False}
                             for s in ["FTSE","DJIM","S&P","Malaysia"]},
                "source": "screening_file",
                "ratios": {},
            }
            continue

        def under(x, seuil):
            """True si x <= seuil ou donnée manquante (ne pénalise pas)."""
            return x is None or x <= seuil

        ftse     = under(j,0.33) and under(k,0.50) and under(l,0.48)
        djim     = under(n,0.33) and under(o,0.33) and under(p,0.33)
        sp       = under(r,0.33) and under(s,0.49) and under(t,0.33)
        malaysia = under(v,0.33) and under(w,0.48)

        # Filtre préalable : revenus portant intérêt / CA ≤ 5%
        halal = under(y, SEUIL_ILLICITE)

        n_pass = sum([ftse, djim, sp, malaysia]) if halal else 0
        compatible = halal and n_pass >= MIN_STANDARDS_PASS

        results[ticker] = {
            "compatible":    compatible,
            "n_standards":   n_pass,
            "halal_sector":  halal,
            "excluded":      False,
            "illicit_ratio": round(y, 4) if y is not None else 0,
            "standards": {
                "FTSE":     {"pass": ftse},
                "DJIM":     {"pass": djim},
                "S&P":      {"pass": sp},
                "Malaysia": {"pass": malaysia},
            },
            "source": "screening_file",
            "ratios": {
                "RE_FTSE": round(j,4) if j is not None else None,
                "RC_FTSE": round(k,4) if k is not None else None,
                "RL_FTSE": round(l,4) if l is not None else None,
                "RE_DJIM": round(n,4) if n is not None else None,
                "RC_DJIM": round(o,4) if o is not None else None,
                "RL_DJIM": round(p,4) if p is not None else None,
            },
        }

    return results


def screen_from_fin_data(ticker, fin_data, nb_titres, cours_moy=None):
    """
    Recalcule le screening depuis les états financiers chargés.

    Exclusions :
      1. Banques conventionnelles → non compatible (halal=False)
      2. Revenus d'intérêts > 5% du CA → non compatible (Riba)

    cap24 = moyenne des CP sur les 2 dernières années disponibles
    cap36 = moyenne des CP sur les 3 dernières années disponibles
    """
    d = fin_data.get(ticker, {})
    if not d:
        return None

    yr    = max(d.keys())
    p     = d[yr]
    actif = abs(p.get("total_actif") or 0)
    if actif == 0:
        return None

    # ── 1. Exclusion banques conventionnelles ─────────────────
    is_bank = (p.get("type") == "banque") or (ticker.upper() in BANK_TICKERS)
    if is_bank:
        return {
            "compatible":    False,
            "n_standards":   0,
            "halal_sector":  False,
            "excluded":      True,
            "raison":        "Banque conventionnelle — incompatible Charia (modèle Riba)",
            "standards":     {s: {"pass": False} for s in ["DJIM","FTSE","S&P","Malaysia"]},
            "annee":         yr,
            "source":        "fin_data",
            "ratios":        {},
        }

    # ── 2. Exclusion secteurs illicites (tabac, jeux, alcool) ─
    illicit_sector = ILLICIT_SECTOR_TICKERS.get(ticker.upper())
    if illicit_sector:
        return {
            "compatible":    False,
            "n_standards":   0,
            "halal_sector":  False,
            "excluded":      True,
            "raison":        f"Secteur illicite — {illicit_sector}",
            "standards":     {s: {"pass": False} for s in ["DJIM","FTSE","S&P","Malaysia"]},
            "annee":         yr,
            "source":        "fin_data",
            "ratios":        {},
        }

    dette  = abs(p.get("dette_financiere") or 0)
    crean  = abs(p.get("creances_clientele") or 0)
    cash   = abs(p.get("tresorerie") or 0)

    # ── 2. Détection revenus illicites (Riba sur sociétés) ────
    ca           = abs(p.get("ca") or p.get("pnb") or 0)
    rev_fin      = abs(p.get("revenus_financiers") or 0)
    interets     = abs(p.get("interets_charges") or 0)
    # Proxy : revenus financiers (intérêts reçus) / CA total
    illicit_ratio = rev_fin / ca if ca > 0 else 0
    halal = illicit_ratio <= SEUIL_ILLICITE

    # ── 3. Capitaux propres moyens ────────────────────────────
    cp_hist = {}
    for yr_h, postes_h in d.items():
        cp_h = abs(postes_h.get("capitaux_propres") or
                   postes_h.get("capitaux_propres_mere") or 0)
        if cp_h > 0:
            cp_hist[yr_h] = cp_h

    cp_sorted = [cp_hist[y] for y in sorted(cp_hist.keys())]
    cap24 = float(np.mean(cp_sorted[-2:])) if len(cp_sorted) >= 2 else \
            (cp_sorted[-1] if cp_sorted else None)
    cap36 = float(np.mean(cp_sorted[-3:])) if len(cp_sorted) >= 3 else \
            (float(np.mean(cp_sorted[-2:])) if len(cp_sorted) >= 2 else
             (cp_sorted[-1] if cp_sorted else None))

    # ── 4. Ratios financiers ──────────────────────────────────
    re     = _r(dette, actif)
    rc_a   = _r(crean, actif)
    rl     = _r(cash,  actif)
    re_c24 = _r(dette, cap24) if cap24 else re
    rc_c24 = _r(crean, cap24) if cap24 else rc_a
    rl_c24 = _r(cash,  cap24) if cap24 else rl
    re_c36 = _r(dette, cap36) if cap36 else re
    rc_c36 = _r(crean, cap36) if cap36 else rc_a
    rl_c36 = _r(cash,  cap36) if cap36 else rl

    stds, n_pass = _screen_ratios(
        re, rc_a, rl,
        re_c24, rc_c24, rl_c24,
        re_c36, rc_c36, rl_c36,
        halal=halal
    )
    return {
        "compatible":    n_pass >= MIN_STANDARDS_PASS,
        "n_standards":   n_pass,
        "halal_sector":  halal,
        "excluded":      False,
        "illicit_ratio": round(illicit_ratio, 4),
        "standards":     stds,
        "annee":         yr,
        "source":        "fin_data",
        "cap24_source":  f"moy CP {len(cp_sorted[-2:])} ans" if cap24 else "actif (fallback)",
        "cap36_source":  f"moy CP {min(len(cp_sorted),3)} ans" if cap36 else "actif (fallback)",
        "ratios": {
            "RE_actif":  round(re,4)     if re     is not None else None,
            "RC_actif":  round(rc_a,4)   if rc_a   is not None else None,
            "RL_actif":  round(rl,4)     if rl     is not None else None,
            "RE_cap24":  round(re_c24,4) if re_c24 is not None else None,
            "RE_cap36":  round(re_c36,4) if re_c36 is not None else None,
            "Rev_fin/CA":round(illicit_ratio,4),
        },
    }


def screen_all_from_fin_data(fin_data, nb_titres, cours_moy=None):
    """Screening de tous les tickers depuis fin_data."""
    return {t: screen_from_fin_data(t, fin_data, nb_titres, cours_moy)
            for t in fin_data
            if screen_from_fin_data(t, fin_data, nb_titres, cours_moy)}


def get_charia_label(ticker, screening_results):
    """Label court pour affichage tableau."""
    r = screening_results.get(ticker)
    if r is None:
        return "—"
    if r.get("donnees_manquantes"):
        return "❓ Données insuffisantes"
    if r.get("excluded"):
        raison = r.get("raison", "")
        if "Banque" in raison:
            return "🏦 Exclu (banque)"
        if "tabac" in raison.lower():
            return "🚬 Exclu (tabac)"
        if "jeux" in raison.lower():
            return "🎲 Exclu (jeux)"
        if "alcool" in raison.lower():
            return "🍺 Exclu (alcool)"
        return f"⛔ Exclu"
    n = r.get("n_standards", 0)
    if not r.get("halal_sector"):
        return f"☽ Riba {n}/4"
    return f"☪️ {n}/4" if r.get("compatible") else f"✗ {n}/4"


def get_charia_compatible_tickers(screening_results):
    """Retourne la liste des tickers compatibles Charia."""
    return [t for t, r in screening_results.items() if r and r.get("compatible")]
